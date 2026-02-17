"""Excel export service for assets."""

from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill

from django.db.models import Sum

from ..models import Asset

# Threshold above which .iterator() is used for memory efficiency
ITERATOR_THRESHOLD = 1000
ITERATOR_CHUNK_SIZE = 1000


def export_assets_xlsx(queryset=None) -> BytesIO:
    """Export assets to an Excel workbook.

    Returns a BytesIO containing the .xlsx file.
    Uses .iterator() for exports exceeding 1,000 assets to
    avoid memory pressure.
    """
    if queryset is None:
        queryset = Asset.objects.select_related(
            "category",
            "category__department",
            "current_location",
            "checked_out_to",
            "created_by",
        ).prefetch_related("tags")

    wb = openpyxl.Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="F59E0B", end_color="F59E0B", fill_type="solid"
    )

    from django.conf import settings

    total_count = queryset.count()

    ws_summary.append([f"{settings.SITE_NAME} Asset Export"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([])
    ws_summary.append(["Total Assets", total_count])
    ws_summary.append(["Active", queryset.filter(status="active").count()])
    ws_summary.append(["Draft", queryset.filter(status="draft").count()])
    ws_summary.append(
        [
            "Checked Out",
            queryset.filter(checked_out_to__isnull=False).count(),
        ]
    )

    ws_summary.append([])
    # Calculate totals using database aggregation (avoids loading
    # all rows into Python memory)
    agg = queryset.aggregate(
        total_purchase=Sum("purchase_price"),
        total_estimated=Sum("estimated_value"),
    )
    total_purchase = float(agg["total_purchase"] or 0)
    total_estimated = float(agg["total_estimated"] or 0)
    ws_summary.append(["Total Purchase Price", f"${total_purchase:,.2f}"])
    ws_summary.append(["Total Estimated Value", f"${total_estimated:,.2f}"])

    # Assets sheet
    ws_assets = wb.create_sheet("Assets")
    headers = [
        "Name",
        "Description",
        "Barcode",
        "Category",
        "Department",
        "Location",
        "Condition",
        "Status",
        "Purchase Price",
        "Estimated Value",
        "Tags",
        "Quantity",
        "Created Date",
        "Last Updated",
        "Checked Out To",
    ]
    ws_assets.append(headers)
    for col_idx, _header in enumerate(headers, 1):
        cell = ws_assets.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    # Use .iterator() for large datasets to reduce memory pressure
    use_iterator = total_count > ITERATOR_THRESHOLD
    asset_iter = (
        queryset.iterator(chunk_size=ITERATOR_CHUNK_SIZE)
        if use_iterator
        else queryset
    )

    for asset in asset_iter:
        location_display = ""
        if asset.checked_out_to:
            location_display = (
                f"Checked out to {asset.checked_out_to.get_display_name()}"
            )
        elif asset.current_location:
            location_display = str(asset.current_location)
        elif asset.status == "active":
            location_display = "Unknown"

        ws_assets.append(
            [
                asset.name,
                asset.description or "",
                asset.barcode,
                asset.category.name if asset.category else "",
                (
                    asset.category.department.name
                    if asset.category and asset.category.department
                    else ""
                ),
                location_display,
                asset.get_condition_display(),
                asset.get_status_display(),
                float(asset.purchase_price) if asset.purchase_price else "",
                float(asset.estimated_value) if asset.estimated_value else "",
                ", ".join(t.name for t in asset.tags.all()),
                asset.quantity,
                (
                    asset.created_at.strftime("%Y-%m-%dT%H:%M:%S")
                    if asset.created_at
                    else ""
                ),
                (
                    asset.updated_at.strftime("%Y-%m-%dT%H:%M:%S")
                    if asset.updated_at
                    else ""
                ),
                (
                    asset.checked_out_to.get_full_name()
                    if asset.checked_out_to
                    else ""
                ),
            ]
        )

    # Auto-size columns
    for ws in [ws_summary, ws_assets]:
        for column_cells in ws.columns:
            max_length = max(
                len(str(cell.value or "")) for cell in column_cells
            )
            ws.column_dimensions[column_cells[0].column_letter].width = min(
                max_length + 2, 50
            )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
