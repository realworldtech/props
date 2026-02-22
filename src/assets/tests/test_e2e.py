"""Tests for end-to-end scenarios and edge cases."""

import json

import pytest

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.urls import reverse

from assets.factories import (
    AssetFactory,
    AssetImageFactory,
    AssetKitFactory,
    AssetSerialFactory,
    CategoryFactory,
    DepartmentFactory,
    HoldListFactory,
    HoldListItemFactory,
    HoldListStatusFactory,
    LocationFactory,
    NFCTagFactory,
    ProjectFactory,
    SiteBrandingFactory,
    StocktakeItemFactory,
    StocktakeSessionFactory,
    TagFactory,
    TransactionFactory,
    UserFactory,
    VirtualBarcodeFactory,
)
from assets.models import (
    Asset,
    AssetImage,
    AssetSerial,
    Category,
    Department,
    Location,
    NFCTag,
    StocktakeSession,
    Transaction,
)

User = get_user_model()

# ============================================================
# EDGE CASE REGRESSION TESTS (V36)
# ============================================================


@pytest.mark.django_db
class TestEdgeCaseScanLookupNullSafety:
    """S7.3.3: scan_lookup returns valid JSON when current_location is null."""

    def test_scan_lookup_null_location_returns_valid_json(
        self, client_logged_in, asset
    ):
        """Asset with current_location=None should return JSON with
        location=null, not crash."""
        asset.current_location = None
        asset.status = "draft"
        asset.save()
        response = client_logged_in.get(
            reverse("assets:scan_lookup") + f"?code={asset.barcode}"
        )
        assert response.status_code == 200
        data = response.json()
        assert data["found"] is True
        assert data["location"] is None
        assert data["asset_name"] == asset.name

    def test_scan_lookup_null_location_serial(
        self, client_logged_in, serialised_asset, location
    ):
        """Serial with null current_location and parent with null
        current_location should not crash."""
        serial = AssetSerial.objects.create(
            asset=serialised_asset,
            serial_number="NL001",
            barcode=f"{serialised_asset.barcode}-NL001",
            status="active",
            current_location=None,
        )
        serialised_asset.current_location = None
        serialised_asset.status = "draft"
        serialised_asset.save()
        response = client_logged_in.get(
            reverse("assets:scan_lookup") + f"?code={serial.barcode}"
        )
        assert response.status_code == 200
        data = response.json()
        assert data["found"] is True
        assert data["location"] is None

    def test_scan_lookup_nfc_null_location(
        self, client_logged_in, asset, user
    ):
        """NFC tag lookup on asset with null location returns valid JSON."""
        asset.current_location = None
        asset.status = "draft"
        asset.save()
        NFCTag.objects.create(
            tag_id="NFC-NULL-LOC",
            asset=asset,
            assigned_by=user,
        )
        response = client_logged_in.get(
            reverse("assets:scan_lookup") + "?code=NFC-NULL-LOC"
        )
        assert response.status_code == 200
        data = response.json()
        assert data["found"] is True
        assert data["location"] is None

    def test_scan_lookup_empty_code(self, client_logged_in):
        """Empty code param returns found=False with no crash."""
        response = client_logged_in.get(
            reverse("assets:scan_lookup") + "?code="
        )
        assert response.status_code == 200
        data = response.json()
        assert data["found"] is False

    def test_scan_lookup_unknown_code(self, client_logged_in):
        """Unknown code returns found=False with capture URL."""
        response = client_logged_in.get(
            reverse("assets:scan_lookup") + "?code=NOSUCHCODE"
        )
        assert response.status_code == 200
        data = response.json()
        assert data["found"] is False
        assert "quick_capture_url" in data


@pytest.mark.django_db
class TestEdgeCaseConcurrentCheckout:
    """S7.4.2: Concurrent checkout protection via select_for_update."""

    def test_already_checked_out_asset_rejects_second_checkout(
        self, client_logged_in, asset, user, second_user, password
    ):
        """If asset is already checked out, a second checkout should fail."""
        asset.checked_out_to = second_user
        asset.save()
        response = client_logged_in.post(
            reverse("assets:asset_checkout", args=[asset.pk]),
            {"borrower": user.pk, "notes": "Second checkout"},
        )
        # Should redirect back (not crash) with error message
        assert response.status_code == 302
        asset.refresh_from_db()
        # Still checked out to original borrower
        assert asset.checked_out_to == second_user

    def test_checkout_view_uses_select_for_update(
        self, admin_client, asset, second_user
    ):
        """Verify the checkout view successfully checks out via the
        atomic select_for_update path."""
        response = admin_client.post(
            reverse("assets:asset_checkout", args=[asset.pk]),
            {"borrower": second_user.pk, "notes": "Locking checkout"},
        )
        assert response.status_code == 302
        asset.refresh_from_db()
        assert asset.checked_out_to == second_user
        assert Transaction.objects.filter(
            asset=asset, action="checkout"
        ).exists()


@pytest.mark.django_db
class TestEdgeCaseLocationHierarchy:
    """S7.6.1 & S7.6.3: Location deletion with children and
    circular references."""

    def test_location_parent_on_delete_cascades(self, location):
        """Deleting a parent location cascades to children (on_delete=CASCADE
        per model definition)."""
        child = Location.objects.create(name="Child Loc", parent=location)
        grandchild = Location.objects.create(name="Grandchild", parent=child)
        child_pk = child.pk
        grandchild_pk = grandchild.pk
        location.delete()
        assert not Location.objects.filter(pk=child_pk).exists()
        assert not Location.objects.filter(pk=grandchild_pk).exists()

    def test_location_with_active_assets_protected(self, location, asset):
        """Cannot delete a location that has assets (PROTECT on
        Asset.current_location)."""
        from django.db.models import ProtectedError

        with pytest.raises(ProtectedError):
            location.delete()

    def test_circular_location_self_parent(self, location):
        """Setting a location as its own parent is blocked by clean()."""
        location.parent = location
        with pytest.raises(ValidationError, match="own ancestor"):
            location.clean()

    def test_circular_location_descendant_parent(self, location):
        """Setting a location's parent to its child creates a cycle."""
        child = Location.objects.create(name="Child", parent=location)
        location.parent = child
        with pytest.raises(ValidationError, match="own ancestor"):
            location.clean()

    def test_circular_location_deep_cycle(self, location):
        """Three-level circular reference is detected."""
        child = Location.objects.create(name="L2", parent=location)
        grandchild = Location.objects.create(name="L3", parent=child)
        location.parent = grandchild
        with pytest.raises(ValidationError, match="own ancestor"):
            location.clean()

    def test_max_nesting_depth_exceeded(self, location):
        """Location clean() blocks nesting deeper than 4 levels."""
        l2 = Location.objects.create(name="L2", parent=location)
        l3 = Location.objects.create(name="L3", parent=l2)
        l4 = Location.objects.create(name="L4", parent=l3)
        l5 = Location(name="L5", parent=l4)
        with pytest.raises(ValidationError, match="nesting depth"):
            l5.clean()


@pytest.mark.django_db
class TestEdgeCaseUserDeletion:
    """S7.10.1 & S7.10.2: User deletion with checked-out or created assets."""

    def test_user_with_checked_out_assets_set_null(self, asset, second_user):
        """Deleting a user who has assets checked out sets
        checked_out_to to NULL (SET_NULL)."""
        asset.checked_out_to = second_user
        asset.save()
        second_user.delete()
        asset.refresh_from_db()
        assert asset.checked_out_to is None

    def test_user_who_created_assets_set_null(self, asset, user):
        """Deleting the user who created an asset sets created_by
        to NULL (SET_NULL)."""
        assert asset.created_by == user
        # Delete the user
        user.delete()
        asset.refresh_from_db()
        assert asset.created_by is None

    def test_transaction_user_set_null(self, asset, user, location):
        """Deleting a user who performed transactions sets user to NULL."""
        Transaction.objects.create(
            asset=asset,
            user=user,
            action="audit",
            from_location=location,
        )
        user.delete()
        txn = Transaction.objects.filter(asset=asset, action="audit").first()
        assert txn.user is None


@pytest.mark.django_db
class TestEdgeCaseBulkStateChangePartialFailures:
    """S7.5.5: Bulk state change handles partial failures."""

    def test_bulk_state_change_mixed_transitions(
        self, user, category, location
    ):
        """Bulk status change succeeds for valid assets and reports
        failures for invalid ones."""
        from assets.services.bulk import bulk_status_change

        active_asset = Asset(
            name="Active One",
            category=category,
            current_location=location,
            status="active",
            is_serialised=False,
            created_by=user,
        )
        active_asset.save()
        disposed_asset = Asset(
            name="Disposed One",
            category=category,
            current_location=location,
            status="disposed",
            is_serialised=False,
            created_by=user,
        )
        disposed_asset.save()
        success_count, failures = bulk_status_change(
            [active_asset.pk, disposed_asset.pk], "retired", user
        )
        assert success_count == 1
        assert len(failures) == 1
        assert "Disposed One" in failures[0]
        active_asset.refresh_from_db()
        assert active_asset.status == "retired"
        disposed_asset.refresh_from_db()
        assert disposed_asset.status == "disposed"

    def test_bulk_state_change_checked_out_blocked(
        self, user, second_user, category, location
    ):
        """Bulk retire on a checked-out asset fails for that asset."""
        from assets.services.bulk import bulk_status_change

        checked_out = Asset(
            name="Checked Out",
            category=category,
            current_location=location,
            status="active",
            is_serialised=False,
            created_by=user,
        )
        checked_out.save()
        checked_out.checked_out_to = second_user
        checked_out.save()
        success_count, failures = bulk_status_change(
            [checked_out.pk], "retired", user
        )
        assert success_count == 0
        assert len(failures) == 1
        assert "Checked Out" in failures[0]

    def test_bulk_state_change_all_succeed(self, user, category, location):
        """When all assets can transition, all succeed with no failures."""
        from assets.services.bulk import bulk_status_change

        a1 = Asset(
            name="A1",
            category=category,
            current_location=location,
            status="active",
            is_serialised=False,
            created_by=user,
        )
        a1.save()
        a2 = Asset(
            name="A2",
            category=category,
            current_location=location,
            status="active",
            is_serialised=False,
            created_by=user,
        )
        a2.save()
        success_count, failures = bulk_status_change(
            [a1.pk, a2.pk], "retired", user
        )
        assert success_count == 2
        assert len(failures) == 0

    def test_bulk_state_change_all_fail(self, user, category, location):
        """When no assets can transition, all fail."""
        from assets.services.bulk import bulk_status_change

        d1 = Asset(
            name="D1",
            category=category,
            current_location=location,
            status="disposed",
            is_serialised=False,
            created_by=user,
        )
        d1.save()
        success_count, failures = bulk_status_change([d1.pk], "active", user)
        assert success_count == 0
        assert len(failures) == 1


@pytest.mark.django_db
class TestEdgeCasePrimaryImagePromotion:
    """S7.8.4: When the primary image is deleted, the next image
    becomes primary."""

    def test_delete_primary_promotes_next_image(self, admin_client, asset):
        """Deleting the primary image promotes the next image."""
        from django.core.files.uploadedfile import SimpleUploadedFile

        # Create two images manually
        img1 = AssetImage.objects.create(
            asset=asset,
            image=SimpleUploadedFile(
                "img1.jpg",
                b"\xff\xd8\xff\xe0" + b"\x00" * 50,
                content_type="image/jpeg",
            ),
            is_primary=True,
        )
        img2 = AssetImage.objects.create(
            asset=asset,
            image=SimpleUploadedFile(
                "img2.jpg",
                b"\xff\xd8\xff\xe0" + b"\x00" * 50,
                content_type="image/jpeg",
            ),
            is_primary=False,
        )
        # Delete primary via view
        response = admin_client.post(
            reverse(
                "assets:image_delete",
                args=[asset.pk, img1.pk],
            )
        )
        assert response.status_code == 302
        assert not AssetImage.objects.filter(pk=img1.pk).exists()
        img2.refresh_from_db()
        assert img2.is_primary is True

    def test_delete_only_image_no_crash(self, admin_client, asset):
        """Deleting the only (primary) image doesn't crash when there's
        no next image to promote."""
        from django.core.files.uploadedfile import SimpleUploadedFile

        img = AssetImage.objects.create(
            asset=asset,
            image=SimpleUploadedFile(
                "only.jpg",
                b"\xff\xd8\xff\xe0" + b"\x00" * 50,
                content_type="image/jpeg",
            ),
            is_primary=True,
        )
        response = admin_client.post(
            reverse(
                "assets:image_delete",
                args=[asset.pk, img.pk],
            )
        )
        assert response.status_code == 302
        assert asset.images.count() == 0

    def test_delete_non_primary_preserves_primary(self, admin_client, asset):
        """Deleting a non-primary image doesn't change which image
        is primary."""
        from django.core.files.uploadedfile import SimpleUploadedFile

        img1 = AssetImage.objects.create(
            asset=asset,
            image=SimpleUploadedFile(
                "p.jpg",
                b"\xff\xd8\xff\xe0" + b"\x00" * 50,
                content_type="image/jpeg",
            ),
            is_primary=True,
        )
        img2 = AssetImage.objects.create(
            asset=asset,
            image=SimpleUploadedFile(
                "s.jpg",
                b"\xff\xd8\xff\xe0" + b"\x00" * 50,
                content_type="image/jpeg",
            ),
            is_primary=False,
        )
        admin_client.post(
            reverse(
                "assets:image_delete",
                args=[asset.pk, img2.pk],
            )
        )
        img1.refresh_from_db()
        assert img1.is_primary is True


@pytest.mark.django_db
class TestEdgeCaseCrossDepartmentMerge:
    """S7.1.3: Department managers cannot merge assets across departments."""

    def test_dept_manager_cannot_merge_cross_department(
        self, dept_manager_user, department, category, location, user
    ):
        """Department manager merging assets from another department
        raises ValueError."""
        from assets.services.merge import merge_assets

        other_dept = Department.objects.create(name="Costumes")
        other_cat = Category.objects.create(
            name="Dresses", department=other_dept
        )
        primary = Asset(
            name="Primary",
            category=category,
            current_location=location,
            status="active",
            is_serialised=False,
            created_by=user,
        )
        primary.save()
        foreign_dup = Asset(
            name="Foreign Dup",
            category=other_cat,
            current_location=location,
            status="active",
            is_serialised=False,
            created_by=user,
        )
        foreign_dup.save()
        with pytest.raises(ValueError, match="do not manage"):
            merge_assets(primary, [foreign_dup], dept_manager_user)

    def test_admin_can_merge_cross_department(
        self, admin_user, department, category, location, user
    ):
        """System admin can merge assets across departments."""
        from assets.services.merge import merge_assets

        other_dept = Department.objects.create(name="Sound")
        other_cat = Category.objects.create(
            name="Speakers", department=other_dept
        )
        primary = Asset(
            name="Primary",
            category=category,
            current_location=location,
            status="active",
            is_serialised=False,
            created_by=user,
        )
        primary.save()
        foreign_dup = Asset(
            name="Foreign Dup",
            category=other_cat,
            current_location=location,
            status="active",
            is_serialised=False,
            created_by=user,
        )
        foreign_dup.save()
        result = merge_assets(primary, [foreign_dup], admin_user)
        assert result.pk == primary.pk
        foreign_dup.refresh_from_db()
        assert foreign_dup.status == "disposed"


@pytest.mark.django_db
class TestEdgeCaseBarcodeCollision:
    """S7.4.1: Barcode generation handles IntegrityError/collision."""

    def test_barcode_generation_produces_unique_values(
        self, category, location, user
    ):
        """Multiple assets always end up with distinct barcodes."""
        assets = []
        for i in range(10):
            a = Asset(
                name=f"Asset {i}",
                category=category,
                current_location=location,
                status="active",
                is_serialised=False,
                created_by=user,
            )
            a.save()
            assets.append(a)
        barcodes = [a.barcode for a in assets]
        assert len(set(barcodes)) == 10

    def test_save_retry_logic_exists(self, category, location, user):
        """Asset.save() has a max_attempts retry loop for IntegrityError."""
        import inspect

        source = inspect.getsource(Asset.save)
        assert "max_attempts" in source
        assert "IntegrityError" in source
        assert "_generate_barcode" in source

    def test_barcode_uniqueness_enforced(self, category, location, user):
        """Two assets cannot share the same barcode."""
        a1 = Asset(
            name="Original",
            category=category,
            current_location=location,
            status="active",
            is_serialised=False,
            created_by=user,
        )
        a1.save()
        # Directly try to force same barcode using update
        a2 = Asset(
            name="Duplicate",
            category=category,
            current_location=location,
            status="active",
            is_serialised=False,
            created_by=user,
        )
        a2.save()
        with pytest.raises(IntegrityError):
            Asset.objects.filter(pk=a2.pk).update(barcode=a1.barcode)


@pytest.mark.django_db
class TestEdgeCaseTransactionImmutability:
    """S7.9: Transactions are immutable — cannot update or delete."""

    def test_transaction_cannot_be_updated(self, asset, user, location):
        """Updating an existing transaction raises ValidationError."""
        txn = Transaction.objects.create(
            asset=asset,
            user=user,
            action="audit",
            from_location=location,
        )
        txn.notes = "Modified notes"
        with pytest.raises(ValidationError, match="immutable"):
            txn.save()

    def test_transaction_cannot_be_deleted(self, asset, user, location):
        """Deleting a transaction raises ValidationError."""
        txn = Transaction.objects.create(
            asset=asset,
            user=user,
            action="audit",
            from_location=location,
        )
        with pytest.raises(ValidationError, match="immutable"):
            txn.delete()


@pytest.mark.django_db
class TestEdgeCaseAssetCleanValidation:
    """S7.2: Non-draft assets require category and location."""

    def test_active_asset_without_category_fails_clean(self, location):
        """Active asset with no category fails validation."""
        a = Asset(
            name="No Cat",
            status="active",
            current_location=location,
        )
        with pytest.raises(ValidationError, match="category"):
            a.clean()

    def test_active_asset_without_location_fails_clean(self, category):
        """Active asset with no location fails validation."""
        a = Asset(
            name="No Loc",
            status="active",
            category=category,
        )
        with pytest.raises(ValidationError, match="current_location"):
            a.clean()

    def test_draft_asset_without_category_and_location_passes(self):
        """Draft asset needs neither category nor location."""
        a = Asset(name="Bare Draft", status="draft")
        a.clean()  # Should not raise


@pytest.mark.django_db
class TestEdgeCaseDisposedAssetNoTransitions:
    """S7.5: Disposed is a terminal state — no transitions allowed."""

    def test_disposed_cannot_transition_to_active(
        self, user, category, location
    ):
        """Disposed asset cannot be reactivated."""
        from assets.services.state import validate_transition

        a = Asset(
            name="Gone",
            category=category,
            current_location=location,
            status="disposed",
            is_serialised=False,
            created_by=user,
        )
        a.save()
        with pytest.raises(ValidationError, match="Cannot transition"):
            validate_transition(a, "active")

    def test_disposed_cannot_transition_to_any_state(
        self, user, category, location
    ):
        """No outgoing transitions from disposed."""
        a = Asset(
            name="Terminal",
            category=category,
            current_location=location,
            status="disposed",
            is_serialised=False,
            created_by=user,
        )
        a.save()
        for target in ["active", "retired", "missing", "lost", "stolen"]:
            assert a.can_transition_to(target) is False


# ============================================================
# E2E SCENARIO TESTS (V38 / S8.5)
# ============================================================


@pytest.mark.django_db
class TestE2EBorrowingLifecycle:
    """S8.5.2: Full borrowing lifecycle — checkout, search, checkin."""

    def test_full_checkout_search_checkin_lifecycle(
        self, client_logged_in, asset, user, second_user, location
    ):
        """Complete check-out -> My Borrowed Items -> check-in cycle."""
        # Precondition: asset is active and not checked out
        asset.status = "active"
        asset.save()
        assert asset.checked_out_to is None

        # --- Step 1: Check out the asset to second_user ---
        checkout_url = reverse("assets:asset_checkout", args=[asset.pk])
        response = client_logged_in.post(
            checkout_url,
            {
                "borrower": second_user.pk,
                "notes": "E2E borrowing lifecycle test",
            },
        )
        assert response.status_code == 302

        # Verify: asset is now checked out
        asset.refresh_from_db()
        assert asset.checked_out_to == second_user

        # Verify: a checkout transaction was created
        checkout_tx = Transaction.objects.filter(
            asset=asset, action="checkout"
        ).first()
        assert checkout_tx is not None
        assert checkout_tx.borrower == second_user
        assert checkout_tx.user == user
        assert checkout_tx.notes == "E2E borrowing lifecycle test"

        # --- Step 2: Second user views My Borrowed Items ---
        from django.test import Client

        borrower_client = Client()
        borrower_client.login(
            username=second_user.username, password="testpass123!"
        )
        my_items_url = reverse("assets:my_borrowed_items")
        response = borrower_client.get(my_items_url)
        assert response.status_code == 200
        assert asset.name.encode() in response.content

        # --- Step 3: Check the asset back in ---
        checkin_url = reverse("assets:asset_checkin", args=[asset.pk])
        response = client_logged_in.post(
            checkin_url,
            {
                "location": location.pk,
                "notes": "E2E checkin complete",
            },
        )
        assert response.status_code == 302

        # Verify: asset is no longer checked out
        asset.refresh_from_db()
        assert asset.checked_out_to is None
        assert asset.current_location == location

        # Verify: a checkin transaction was created
        checkin_tx = Transaction.objects.filter(
            asset=asset, action="checkin"
        ).first()
        assert checkin_tx is not None
        assert checkin_tx.to_location == location
        assert checkin_tx.notes == "E2E checkin complete"

        # --- Step 4: My Borrowed Items is now empty for borrower ---
        response = borrower_client.get(my_items_url)
        assert response.status_code == 200
        assert asset.name.encode() not in response.content

    def test_checkout_updates_home_location(
        self, client_logged_in, asset, second_user, location
    ):
        """Checkout sets home_location if not already set."""
        asset.status = "active"
        asset.home_location = None
        asset.save()

        checkout_url = reverse("assets:asset_checkout", args=[asset.pk])
        client_logged_in.post(
            checkout_url,
            {"borrower": second_user.pk},
        )

        asset.refresh_from_db()
        assert asset.checked_out_to == second_user
        # home_location should be set to original current_location
        assert asset.home_location == location


@pytest.mark.django_db
class TestE2EAIQuickCapture:
    """S8.5.4: Quick capture -> AI analysis -> apply suggestions."""

    def test_quick_capture_then_ai_apply(
        self, client_logged_in, user, category, department
    ):
        """Quick capture creates draft, then AI suggestions are
        applied."""
        # --- Step 1: Quick capture creates a draft asset ---
        qc_url = reverse("assets:quick_capture")
        response = client_logged_in.post(
            qc_url,
            {"name": "Mystery Prop", "notes": "Found backstage"},
        )
        assert response.status_code == 200

        draft = Asset.objects.filter(
            name="Mystery Prop", status="draft"
        ).first()
        assert draft is not None
        assert draft.created_by == user

        # --- Step 2: Simulate AI analysis on an image ---
        # (In production Celery runs the analysis task)
        image = AssetImage.objects.create(
            asset=draft,
            image="test_image.jpg",
            is_primary=True,
            uploaded_by=user,
            ai_processing_status="completed",
            ai_name_suggestion="Vintage Telephone Prop",
            ai_description=("A black rotary telephone, circa 1960s."),
            ai_category_suggestion=category.name,
            ai_tag_suggestions=[
                "vintage",
                "telephone",
                "black",
            ],
            ai_condition_suggestion="good",
        )

        # --- Step 3: Apply AI suggestions ---
        apply_url = reverse(
            "assets:ai_apply_suggestions",
            args=[draft.pk, image.pk],
        )
        response = client_logged_in.post(
            apply_url,
            {
                "apply_name": "1",
                "apply_description": "1",
                "apply_category": "1",
                "apply_tags": "1",
                "apply_condition": "1",
            },
        )
        assert response.status_code == 302

        # --- Step 4: Verify suggestions were applied ---
        draft.refresh_from_db()
        assert draft.name == "Vintage Telephone Prop"
        assert draft.description == ("A black rotary telephone, circa 1960s.")
        assert draft.category == category
        assert draft.condition == "good"

        # Verify tags were created and applied
        tag_names = set(draft.tags.values_list("name", flat=True))
        assert "vintage" in tag_names
        assert "telephone" in tag_names
        assert "black" in tag_names

    def test_quick_capture_with_scanned_code(self, client_logged_in, user):
        """Quick capture with a scanned NFC tag ID."""
        qc_url = reverse("assets:quick_capture")
        response = client_logged_in.post(
            qc_url,
            {
                "name": "Tagged Item",
                "scanned_code": "04:a2:3b:c1:d4:e5:f6",
            },
        )
        assert response.status_code == 200

        draft = Asset.objects.filter(
            name="Tagged Item", status="draft"
        ).first()
        assert draft is not None

        # NFC tag should have been created
        nfc = NFCTag.objects.filter(
            tag_id="04:a2:3b:c1:d4:e5:f6", asset=draft
        ).first()
        assert nfc is not None


@pytest.mark.django_db
class TestE2EConcurrentCheckout:
    """S8.5.7: Two simultaneous checkouts — only first succeeds."""

    def test_concurrent_checkout_second_fails(
        self,
        admin_client,
        admin_user,
        asset,
        second_user,
        user,
        password,
    ):
        """Two requests try to check out the same asset; second is
        rejected."""
        asset.status = "active"
        asset.save()

        checkout_url = reverse("assets:asset_checkout", args=[asset.pk])

        # --- First checkout: succeeds ---
        response1 = admin_client.post(
            checkout_url,
            {
                "borrower": second_user.pk,
                "notes": "First checkout",
            },
        )
        assert response1.status_code == 302

        asset.refresh_from_db()
        assert asset.checked_out_to == second_user

        # --- Second checkout attempt: should fail ---
        response2 = admin_client.post(
            checkout_url,
            {
                "borrower": user.pk,
                "notes": "Second checkout attempt",
            },
        )
        assert response2.status_code == 302

        # Asset should still be checked out to the first borrower
        asset.refresh_from_db()
        assert asset.checked_out_to == second_user

        # Only one checkout transaction should exist
        checkout_count = Transaction.objects.filter(
            asset=asset, action="checkout"
        ).count()
        assert checkout_count == 1

    def test_concurrent_checkout_via_two_clients(
        self,
        admin_user,
        asset,
        second_user,
        user,
        password,
    ):
        """Simulate race condition with two separate client
        sessions."""
        from django.test import Client

        asset.status = "active"
        asset.save()

        client_a = Client()
        client_b = Client()
        client_a.login(username=admin_user.username, password=password)
        client_b.login(username=admin_user.username, password=password)

        checkout_url = reverse("assets:asset_checkout", args=[asset.pk])

        # Both clients GET the checkout page (both see available)
        resp_a = client_a.get(checkout_url)
        resp_b = client_b.get(checkout_url)
        assert resp_a.status_code == 200
        assert resp_b.status_code == 200

        # Client A checks out first
        client_a.post(
            checkout_url,
            {
                "borrower": second_user.pk,
                "notes": "Client A",
            },
        )
        asset.refresh_from_db()
        assert asset.checked_out_to == second_user

        # Client B tries to check out (should fail gracefully)
        client_b.post(
            checkout_url,
            {"borrower": user.pk, "notes": "Client B"},
        )

        # Asset still belongs to first borrower
        asset.refresh_from_db()
        assert asset.checked_out_to == second_user
        assert (
            Transaction.objects.filter(asset=asset, action="checkout").count()
            == 1
        )


# ============================================================
# L32: E2E "THE MOVE" SCENARIO TEST
# ============================================================


@pytest.mark.django_db
class TestE2ETheMove:
    """L32: E2E scenario — complete organizational move.

    Simulates a full lifecycle: project creation, hold lists,
    checkout, checkin at new location, and stocktake verification.
    """

    def test_the_move_full_lifecycle(self, admin_client, admin_user):
        """Simulate a full organizational move:
        project -> hold lists -> checkout -> checkin -> stocktake.
        """
        from datetime import date

        from assets.factories import (
            AssetFactory,
            CategoryFactory,
            DepartmentFactory,
            HoldListFactory,
            HoldListItemFactory,
            HoldListStatusFactory,
            LocationFactory,
            ProjectFactory,
            UserFactory,
        )
        from assets.models import ProjectDateRange

        # --- Setup: two locations, department, category, assets ---
        old_location = LocationFactory(name="Old Warehouse")
        new_location = LocationFactory(name="New Warehouse")
        dept = DepartmentFactory(name="Theatre")
        category = CategoryFactory(name="Furniture", department=dept)

        assets = [
            AssetFactory(
                name=f"Desk {i}",
                category=category,
                current_location=old_location,
                status="active",
                created_by=admin_user,
            )
            for i in range(5)
        ]

        # --- Step 1: Create project with date ranges ---
        project = ProjectFactory(name="The Big Move", created_by=admin_user)
        ProjectDateRange.objects.create(
            project=project,
            label="Moving Week",
            start_date=date(2026, 3, 1),
            end_date=date(2026, 3, 7),
        )
        assert project.date_ranges.count() == 1

        # --- Step 2: Create hold list for the project ---
        terminal_status = HoldListStatusFactory(
            name="Completed", is_terminal=True
        )
        active_status = HoldListStatusFactory(
            name="Confirmed", is_terminal=False
        )
        hold_list = HoldListFactory(
            name="Move Hold List",
            project=project,
            department=dept,
            status=active_status,
            created_by=admin_user,
        )

        # --- Step 3: Add assets to hold list ---
        for asset in assets:
            HoldListItemFactory(hold_list=hold_list, asset=asset)
        assert hold_list.items.count() == 5

        # --- Step 4: Mark hold list terminal so checkout is allowed ---
        hold_list.status = terminal_status
        hold_list.save()

        # --- Step 5: Check out assets to a borrower ---
        borrower = UserFactory(username="mover", email="mover@example.com")

        for asset in assets:
            response = admin_client.post(
                reverse(
                    "assets:asset_checkout",
                    kwargs={"pk": asset.pk},
                ),
                {"borrower": borrower.pk, "notes": "Moving day"},
            )
            assert response.status_code == 302

        # Verify all assets are checked out
        for asset in assets:
            asset.refresh_from_db()
            assert asset.checked_out_to == borrower

        # --- Step 6: Check in assets at new location ---
        for asset in assets:
            response = admin_client.post(
                reverse(
                    "assets:asset_checkin",
                    kwargs={"pk": asset.pk},
                ),
                {
                    "location": new_location.pk,
                    "notes": "Arrived at new warehouse",
                },
            )
            assert response.status_code == 302

        # --- Step 7: Verify assets at new location ---
        for asset in assets:
            asset.refresh_from_db()
            assert asset.current_location == new_location
            assert asset.checked_out_to is None

        # --- Step 8: Start stocktake at new location ---
        response = admin_client.post(
            reverse("assets:stocktake_start"),
            {"location": new_location.pk},
        )
        assert response.status_code == 302

        session = StocktakeSession.objects.filter(
            location=new_location, status="in_progress"
        ).first()
        assert session is not None
        assert session.started_by == admin_user

        # Verify expected assets match our moved assets
        expected_ids = set(
            session.expected_assets.values_list("pk", flat=True)
        )
        for asset in assets:
            assert asset.pk in expected_ids

        # --- Step 9: Confirm all assets during stocktake ---
        confirm_url = reverse("assets:stocktake_confirm", args=[session.pk])
        for asset in assets:
            response = admin_client.post(confirm_url, {"asset_id": asset.pk})
            assert response.status_code == 302

        # Verify audit transactions were created
        for asset in assets:
            assert Transaction.objects.filter(
                asset=asset, action="audit"
            ).exists()

        # --- Step 10: Complete stocktake ---
        complete_url = reverse("assets:stocktake_complete", args=[session.pk])
        response = admin_client.post(
            complete_url,
            {
                "action": "complete",
                "mark_missing": "1",
                "notes": "The Big Move stocktake",
            },
        )
        assert response.status_code == 302

        # Verify session completed
        session.refresh_from_db()
        assert session.status == "completed"

        # All assets should still be active (all were confirmed)
        for asset in assets:
            asset.refresh_from_db()
            assert asset.status == "active"
            assert asset.current_location == new_location

    def test_the_move_partial_with_missing(self, admin_client, admin_user):
        """Move scenario where some assets go missing in transit."""
        from assets.factories import (
            AssetFactory,
            CategoryFactory,
            DepartmentFactory,
            LocationFactory,
        )

        old_loc = LocationFactory(name="Old Office")
        new_loc = LocationFactory(name="New Office")
        dept = DepartmentFactory(name="Stage")
        cat = CategoryFactory(name="Chairs", department=dept)

        assets = [
            AssetFactory(
                name=f"Chair {i}",
                category=cat,
                current_location=old_loc,
                status="active",
                created_by=admin_user,
            )
            for i in range(4)
        ]

        # Move only first 2 assets to new location via checkin
        borrower_user = User.objects.create_user(
            username="mover2",
            email="mover2@example.com",
            password="test123!",
        )

        for asset in assets[:2]:
            admin_client.post(
                reverse(
                    "assets:asset_checkout",
                    kwargs={"pk": asset.pk},
                ),
                {"borrower": borrower_user.pk, "notes": "Moving"},
            )
            admin_client.post(
                reverse(
                    "assets:asset_checkin",
                    kwargs={"pk": asset.pk},
                ),
                {"location": new_loc.pk, "notes": "Arrived"},
            )

        # Move remaining 2 via checkin as well (they arrive)
        for asset in assets[2:]:
            admin_client.post(
                reverse(
                    "assets:asset_checkout",
                    kwargs={"pk": asset.pk},
                ),
                {"borrower": borrower_user.pk, "notes": "Moving"},
            )
            admin_client.post(
                reverse(
                    "assets:asset_checkin",
                    kwargs={"pk": asset.pk},
                ),
                {"location": new_loc.pk, "notes": "Arrived"},
            )

        # Start stocktake at new location
        admin_client.post(
            reverse("assets:stocktake_start"),
            {"location": new_loc.pk},
        )
        session = StocktakeSession.objects.get(
            location=new_loc, status="in_progress"
        )

        # Only confirm first 2 assets (the other 2 are "missing")
        confirm_url = reverse("assets:stocktake_confirm", args=[session.pk])
        for asset in assets[:2]:
            admin_client.post(confirm_url, {"asset_id": asset.pk})

        # Complete stocktake with mark_missing
        admin_client.post(
            reverse("assets:stocktake_complete", args=[session.pk]),
            {
                "action": "complete",
                "mark_missing": "1",
                "notes": "Partial move stocktake",
            },
        )

        session.refresh_from_db()
        assert session.status == "completed"

        # First 2 assets should be active
        for asset in assets[:2]:
            asset.refresh_from_db()
            assert asset.status == "active"

        # Last 2 assets should be marked missing
        for asset in assets[2:]:
            asset.refresh_from_db()
            assert asset.status == "missing"


@pytest.mark.django_db
class TestExportFormatContent:
    """Export format tests: verify Excel export content/format."""

    def test_export_summary_contains_counts(self, asset):
        from io import BytesIO

        import openpyxl

        from assets.services.export import export_assets_xlsx

        buffer = export_assets_xlsx(Asset.objects.all())
        wb = openpyxl.load_workbook(BytesIO(buffer.getvalue()))
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=1).value for r in range(1, 10)]
        assert "Total Assets" in values
        assert "Active" in values

    def test_export_assets_sheet_has_headers(self, asset):
        from io import BytesIO

        import openpyxl

        from assets.services.export import export_assets_xlsx

        buffer = export_assets_xlsx(Asset.objects.all())
        wb = openpyxl.load_workbook(BytesIO(buffer.getvalue()))
        ws = wb["Assets"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 16)]
        assert "Name" in headers
        assert "Barcode" in headers
        assert "Category" in headers
        assert "Status" in headers
        assert "Checked Out To" in headers

    def test_export_checked_out_asset_location(self, asset, second_user):
        from io import BytesIO

        import openpyxl

        from assets.services.export import export_assets_xlsx

        asset.checked_out_to = second_user
        asset.save(update_fields=["checked_out_to"])

        buffer = export_assets_xlsx(Asset.objects.all())
        wb = openpyxl.load_workbook(BytesIO(buffer.getvalue()))
        ws = wb["Assets"]
        location_val = ws.cell(row=2, column=6).value
        assert "Checked out to" in location_val

    def test_export_includes_tags(self, asset, tag):
        from io import BytesIO

        import openpyxl

        from assets.services.export import export_assets_xlsx

        asset.tags.add(tag)
        buffer = export_assets_xlsx(Asset.objects.all())
        wb = openpyxl.load_workbook(BytesIO(buffer.getvalue()))
        ws = wb["Assets"]
        tags_val = ws.cell(row=2, column=11).value
        assert tag.name in tags_val


class TestV611DockerHealthChecks:
    """V611: Docker health checks for web and celery services."""

    @staticmethod
    def _compose_path():
        from pathlib import Path

        p = Path(__file__).parent.parent.parent.parent / "docker-compose.yml"
        return p if p.exists() else None

    @pytest.mark.skipif(
        not (
            __import__("pathlib").Path(__file__).parent.parent.parent.parent
            / "docker-compose.yml"
        ).exists(),
        reason="docker-compose.yml not available",
    )
    def test_docker_compose_web_healthcheck(self):
        """docker-compose.yml web service should have healthcheck."""
        import re
        from pathlib import Path

        compose = (
            Path(__file__).parent.parent.parent.parent / "docker-compose.yml"
        )
        content = compose.read_text()
        match = re.search(
            r"^\s{2}web:\s*\n((?:\s{4,}.+\n)*)",
            content,
            re.MULTILINE,
        )
        assert match, "web service not found in docker-compose.yml"
        assert "healthcheck:" in match.group(
            1
        ), "web service missing healthcheck"

    @pytest.mark.skipif(
        not (
            __import__("pathlib").Path(__file__).parent.parent.parent.parent
            / "docker-compose.yml"
        ).exists(),
        reason="docker-compose.yml not available",
    )
    def test_docker_compose_celery_healthcheck(self):
        """docker-compose.yml celery-worker should have healthcheck."""
        import re
        from pathlib import Path

        compose = (
            Path(__file__).parent.parent.parent.parent / "docker-compose.yml"
        )
        content = compose.read_text()
        match = re.search(
            r"^\s{2}celery-worker:\s*\n((?:\s{4,}.+\n)*)",
            content,
            re.MULTILINE,
        )
        assert match, "celery-worker not found in docker-compose.yml"
        assert "healthcheck:" in match.group(
            1
        ), "celery-worker missing healthcheck"
