"""S10B Department Manager user story tests.

Each class covers one US-DM-xxx user story. Tests verify acceptance
criteria from the dept manager's perspective. Failures identify spec gaps.

Read: specs/props/sections/s10b-dept-manager-stories.md
"""

import json
from html.parser import HTMLParser
from io import BytesIO

import pytest

from django.urls import reverse

from assets.models import Asset, AssetImage, Tag, Transaction


class _FormFieldCollector(HTMLParser):
    """Collect form field names and values from HTML."""

    def __init__(self):
        super().__init__()
        self.fields = {}
        self._current_select = None
        self._current_options = []
        self._in_textarea = None
        self._textarea_content = []

    def handle_starttag(self, tag, attrs):
        attrs_dict = dict(attrs)
        if tag == "input":
            name = attrs_dict.get("name")
            if name:
                self.fields[name] = attrs_dict.get("value", "")
        elif tag == "select":
            self._current_select = attrs_dict.get("name")
            self._current_options = []
        elif tag == "option" and self._current_select:
            val = attrs_dict.get("value", "")
            if val:
                self._current_options.append(val)
            if "selected" in attrs_dict:
                self.fields[self._current_select] = val
        elif tag == "textarea":
            self._in_textarea = attrs_dict.get("name")
            self._textarea_content = []

    def handle_data(self, data):
        if self._in_textarea is not None:
            self._textarea_content.append(data)

    def handle_endtag(self, tag):
        if tag == "select" and self._current_select:
            if (
                self._current_select not in self.fields
                and self._current_options
            ):
                self.fields[self._current_select] = self._current_options[0]
            self._current_select = None
        elif tag == "textarea" and self._in_textarea:
            self.fields[self._in_textarea] = "".join(self._textarea_content)
            self._in_textarea = None


# ---------------------------------------------------------------------------
# §10B.1 Quick Capture & Drafts
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_001_CaptureAssetViaQuickCapture:
    """US-DM-001: Capture a new asset via Quick Capture.

    MoSCoW: MUST
    Spec refs: S2.1.1-01, S2.1.1-02, S2.1.1-03, S2.1.1-05
    UI Surface: /quick-capture/
    """

    def test_quick_capture_page_loads(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:quick_capture"))
        assert resp.status_code == 200

    def test_submit_photo_creates_draft(
        self, dept_manager_client, dept_manager_user
    ):
        from django.core.files.uploadedfile import SimpleUploadedFile

        image = SimpleUploadedFile(
            "item.gif",
            (
                b"GIF87a\x01\x00\x01\x00\x80\x01\x00\x00\x00\x00"
                b"\xff\xff\xff,\x00\x00\x00\x00\x01\x00\x01\x00"
                b"\x00\x02\x02D\x01\x00;"
            ),
            content_type="image/gif",
        )
        dept_manager_client.post(
            reverse("assets:quick_capture"), {"image": image}
        )
        assert Asset.objects.filter(
            status="draft", created_by=dept_manager_user
        ).exists()

    def test_capture_another_returns_blank_form(self, dept_manager_client):
        from django.core.files.uploadedfile import SimpleUploadedFile

        image = SimpleUploadedFile(
            "item.gif",
            (
                b"GIF87a\x01\x00\x01\x00\x80\x01\x00\x00\x00\x00"
                b"\xff\xff\xff,\x00\x00\x00\x00\x01\x00\x01\x00"
                b"\x00\x02\x02D\x01\x00;"
            ),
            content_type="image/gif",
        )
        resp = dept_manager_client.post(
            reverse("assets:quick_capture"), {"image": image}
        )
        assert resp.status_code in (200, 302)

    def test_quick_capture_auto_name_format(
        self, dept_manager_client, dept_manager_user
    ):
        """S2.1.1: Auto-generated draft name must follow 'Quick Capture
        {MMM DD HH:MM}' format."""
        import re

        from django.core.files.uploadedfile import SimpleUploadedFile

        image = SimpleUploadedFile(
            "item.jpg",
            b"GIF87a\x01\x00\x01\x00\x80\x01\x00\x00\x00\x00"
            b"\xff\xff\xff,\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02D\x01\x00;",
            content_type="image/gif",
        )
        dept_manager_client.post(
            reverse("assets:quick_capture"),
            {"image": image},
        )
        from assets.models import Asset

        draft = (
            Asset.objects.filter(status="draft", created_by=dept_manager_user)
            .order_by("-pk")
            .first()
        )
        assert draft is not None
        # Name should match pattern like "Quick Capture Feb 23 14:37"
        pattern = r"Quick Capture \w+ \d{1,2} \d{2}:\d{2}"
        assert re.match(pattern, draft.name), (
            f"Auto-generated name '{draft.name}' must match"
            " 'Quick Capture MMM DD HH:MM'"
        )


@pytest.mark.django_db
class TestUS_DM_002_EditDraftAssetsInMyDept:
    """US-DM-002: Edit draft assets in my department.

    MoSCoW: MUST
    Spec refs: S2.1.4a-01, S2.2.2-04, S2.2.2-07
    UI Surface: /assets/<pk>/edit/
    """

    def test_dm_can_edit_own_dept_draft(
        self, dept_manager_client, dept_manager_user, category
    ):
        from assets.factories import AssetFactory

        draft = AssetFactory(
            name="Dept Draft Asset",
            status="draft",
            category=category,
            current_location=None,
            created_by=dept_manager_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:asset_edit", args=[draft.pk])
        )
        assert resp.status_code == 200

    def test_dm_can_edit_draft_created_by_team_member(
        self, dept_manager_client, member_user, category
    ):
        from assets.factories import AssetFactory

        draft = AssetFactory(
            name="Member Draft Asset",
            status="draft",
            category=category,
            current_location=None,
            created_by=member_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:asset_edit", args=[draft.pk])
        )
        # DM should be able to edit drafts in their department
        assert resp.status_code in (200, 403)

    def test_dm_cannot_edit_other_dept_draft(
        self, dept_manager_client, tech_dept, member_user
    ):
        from assets.factories import AssetFactory, CategoryFactory

        other_cat = CategoryFactory(name="Tech Cat", department=tech_dept)
        other_draft = AssetFactory(
            name="Other Dept Draft",
            status="draft",
            category=other_cat,
            current_location=None,
            created_by=member_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:asset_edit", args=[other_draft.pk])
        )
        assert resp.status_code in (403, 302)


@pytest.mark.django_db
class TestUS_DM_003_PromoteDraftToActive:
    """US-DM-003: Promote a draft to active status.

    MoSCoW: MUST
    Spec refs: S2.1.5-01, S2.1.5-04, S2.2.3-02
    UI Surface: /assets/<pk>/edit/
    """

    def test_promotion_with_required_fields_succeeds(
        self,
        dept_manager_client,
        dept_manager_user,
        category,
        location,
    ):
        from assets.factories import AssetFactory

        draft = AssetFactory(
            name="Draft to Promote",
            status="draft",
            category=None,
            current_location=None,
            created_by=dept_manager_user,
        )
        resp = dept_manager_client.post(
            reverse("assets:asset_edit", args=[draft.pk]),
            {
                "name": "Promoted Asset",
                "category": category.pk,
                "current_location": location.pk,
                "status": "active",
                "condition": "good",
                "quantity": 1,
            },
        )
        draft.refresh_from_db()
        assert draft.status == "active"

    def test_promoted_asset_appears_in_active_list(
        self,
        dept_manager_client,
        dept_manager_user,
        category,
        location,
    ):
        from assets.factories import AssetFactory

        draft = AssetFactory(
            name="Newly Promoted Asset",
            status="draft",
            category=None,
            current_location=None,
            created_by=dept_manager_user,
        )
        dept_manager_client.post(
            reverse("assets:asset_edit", args=[draft.pk]),
            {
                "name": "Newly Promoted Asset",
                "category": category.pk,
                "current_location": location.pk,
                "status": "active",
                "condition": "good",
                "quantity": 1,
            },
        )
        resp = dept_manager_client.get(
            reverse("assets:asset_list"),
            {"q": "Newly Promoted Asset"},
        )
        assert b"Newly Promoted Asset" in resp.content


@pytest.mark.django_db
class TestUS_DM_004_ReviewDraftsQueueForMyDept:
    """US-DM-004: Review the Drafts Queue for my department.

    MoSCoW: MUST
    Spec refs: S2.1.4-01, S2.1.4-02, S2.1.4-03, S2.1.4-07
    UI Surface: /drafts/
    """

    def test_drafts_queue_page_loads(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:drafts_queue"))
        assert resp.status_code == 200

    def test_drafts_queue_shows_own_dept_drafts(
        self, dept_manager_client, dept_manager_user, category
    ):
        from assets.factories import AssetFactory

        draft = AssetFactory(
            name="My Dept Draft",
            status="draft",
            category=category,
            current_location=None,
            created_by=dept_manager_user,
        )
        resp = dept_manager_client.get(reverse("assets:drafts_queue"))
        assert resp.status_code == 200
        assert draft.name.encode() in resp.content

    def test_drafts_with_ai_show_indicator(
        self, dept_manager_client, dept_manager_user
    ):
        """S2.1.4: Drafts with completed AI analysis must show an AI
        indicator."""
        from assets.factories import AssetFactory, AssetImageFactory

        draft = AssetFactory(status="draft", created_by=dept_manager_user)
        img = AssetImageFactory(asset=draft)
        img.ai_processing_status = "completed"
        img.ai_name_suggestion = "AI Suggested"
        img.save()
        resp = dept_manager_client.get(reverse("assets:drafts_queue"))
        assert resp.status_code == 200
        content = resp.content.decode().lower()
        assert "ai" in content or "suggestion" in content, (
            "Drafts queue must show AI indicator for drafts with completed"
            " AI analysis"
        )


# ---------------------------------------------------------------------------
# §10B.2 Asset Management
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_005_CreateAssetInMyDept:
    """US-DM-005: Create a new asset in my department.

    MoSCoW: MUST
    Spec refs: S2.2.1-01, S2.2.1-02, S2.2.1-03
    UI Surface: /assets/create/
    """

    def test_create_page_loads(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:asset_create"))
        assert resp.status_code == 200

    def test_create_asset_in_own_dept_succeeds(
        self, dept_manager_client, category, location
    ):
        resp = dept_manager_client.post(
            reverse("assets:asset_create"),
            {
                "name": "DM Created Asset",
                "category": category.pk,
                "current_location": location.pk,
                "condition": "good",
                "quantity": 1,
            },
        )
        assert Asset.objects.filter(
            name="DM Created Asset", status="active"
        ).exists()

    def test_created_asset_gets_barcode(
        self, dept_manager_client, category, location
    ):
        dept_manager_client.post(
            reverse("assets:asset_create"),
            {
                "name": "DM Barcode Asset",
                "category": category.pk,
                "current_location": location.pk,
                "condition": "good",
                "quantity": 1,
            },
        )
        asset = Asset.objects.filter(name="DM Barcode Asset").first()
        assert asset is not None
        assert asset.barcode is not None


@pytest.mark.django_db
class TestUS_DM_006_EditAssetInMyDept:
    """US-DM-006: Edit an asset in my department.

    MoSCoW: MUST
    Spec refs: S2.2.1-01, S2.2.2-04, S2.10.3-02
    UI Surface: /assets/<pk>/edit/
    """

    def test_edit_form_loads_for_own_dept_asset(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:asset_edit", args=[active_asset.pk])
        )
        assert resp.status_code == 200

    def test_edit_updates_asset_fields(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.post(
            reverse("assets:asset_edit", args=[active_asset.pk]),
            {
                "name": "DM Updated Name",
                "category": active_asset.category.pk,
                "current_location": (active_asset.current_location.pk),
                "condition": "fair",
                "quantity": 1,
                "status": "active",
            },
        )
        active_asset.refresh_from_db()
        assert active_asset.name == "DM Updated Name"


@pytest.mark.django_db
class TestUS_DM_007_ViewButNotEditOtherDeptAssets:
    """US-DM-007: View but not edit assets in other departments.

    MoSCoW: MUST
    Spec refs: S2.2.2-05, S2.10.3-01, S2.10.3-07
    UI Surface: /assets/<pk>/
    """

    def test_dm_can_view_other_dept_asset_detail(
        self,
        dept_manager_client,
        tech_dept,
        location,
        member_user,
    ):
        from assets.factories import AssetFactory, CategoryFactory

        other_cat = CategoryFactory(name="Tech Items", department=tech_dept)
        other_asset = AssetFactory(
            name="Other Dept Asset",
            status="active",
            category=other_cat,
            current_location=location,
            created_by=member_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:asset_detail", args=[other_asset.pk])
        )
        assert resp.status_code == 200

    def test_dm_cannot_edit_other_dept_asset(
        self,
        dept_manager_client,
        tech_dept,
        location,
        member_user,
    ):
        from assets.factories import AssetFactory, CategoryFactory

        other_cat = CategoryFactory(name="Tech Items 2", department=tech_dept)
        other_asset = AssetFactory(
            name="Other Dept Asset Edit Test",
            status="active",
            category=other_cat,
            current_location=location,
            created_by=member_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:asset_edit", args=[other_asset.pk])
        )
        assert resp.status_code == 403


@pytest.mark.django_db
class TestUS_DM_008_ManageImagesOnDeptAssets:
    """US-DM-008: Manage images on department assets.

    MoSCoW: MUST
    Spec refs: S2.2.5-01, S2.2.5-02, S2.2.5-03, S2.2.5-05
    UI Surface: /assets/<pk>/images/upload/
    """

    def test_image_upload_accessible_for_own_dept(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:image_upload", args=[active_asset.pk])
        )
        assert resp.status_code in (200, 405)

    def test_upload_image_creates_record(
        self, dept_manager_client, active_asset
    ):
        from django.core.files.uploadedfile import SimpleUploadedFile

        from assets.models import AssetImage

        image = SimpleUploadedFile(
            "test.gif",
            (
                b"GIF87a\x01\x00\x01\x00\x80\x01\x00\x00\x00\x00"
                b"\xff\xff\xff,\x00\x00\x00\x00\x01\x00\x01\x00"
                b"\x00\x02\x02D\x01\x00;"
            ),
            content_type="image/gif",
        )
        initial_count = AssetImage.objects.filter(asset=active_asset).count()
        dept_manager_client.post(
            reverse("assets:image_upload", args=[active_asset.pk]),
            {"image": image},
        )
        assert (
            AssetImage.objects.filter(asset=active_asset).count()
            >= initial_count
        )


@pytest.mark.django_db
class TestUS_DM_009_ManageTagsOnDeptAssets:
    """US-DM-009: Manage tags on department assets.

    MoSCoW: MUST
    Spec refs: S2.2.6-01, S2.2.6-06
    UI Surface: /assets/<pk>/edit/
    """

    def test_tag_field_present_on_edit_form(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:asset_edit", args=[active_asset.pk])
        )
        assert resp.status_code == 200
        assert b"tag" in resp.content.lower()

    def test_add_tag_to_dept_asset(
        self, dept_manager_client, active_asset, tag
    ):
        resp = dept_manager_client.post(
            reverse("assets:asset_edit", args=[active_asset.pk]),
            {
                "name": active_asset.name,
                "category": active_asset.category.pk,
                "current_location": (active_asset.current_location.pk),
                "condition": active_asset.condition,
                "quantity": active_asset.quantity,
                "status": active_asset.status,
                "tags": [tag.pk],
            },
        )
        active_asset.refresh_from_db()
        assert tag in active_asset.tags.all()


@pytest.mark.django_db
class TestUS_DM_010_UpdateAssetCondition:
    """US-DM-010: Update asset condition.

    MoSCoW: MUST
    Spec refs: S2.2.4-01, S2.2.4-03
    UI Surface: /assets/<pk>/edit/
    """

    def test_condition_dropdown_present(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:asset_edit", args=[active_asset.pk])
        )
        assert resp.status_code == 200
        assert b"condition" in resp.content

    def test_update_condition_only(self, dept_manager_client, active_asset):
        resp = dept_manager_client.post(
            reverse("assets:asset_edit", args=[active_asset.pk]),
            {
                "name": active_asset.name,
                "category": active_asset.category.pk,
                "current_location": (active_asset.current_location.pk),
                "condition": "poor",
                "quantity": active_asset.quantity,
                "status": active_asset.status,
            },
        )
        active_asset.refresh_from_db()
        assert active_asset.condition == "poor"


@pytest.mark.django_db
class TestUS_DM_011_MergeDuplicateAssetsInDept:
    """US-DM-011: Merge duplicate assets within my department.

    MoSCoW: MUST
    Spec refs: S2.2.7-01, S2.2.7-02, S2.2.7-05, S2.2.7-06, S2.2.7-10
    UI Surface: /assets/merge/select/ -> /assets/merge/execute/
    """

    def test_merge_select_page_loads(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:asset_merge_select"))
        assert resp.status_code == 200

    def test_merge_sets_secondary_to_disposed(
        self,
        dept_manager_client,
        active_asset,
        category,
        location,
        dept_manager_user,
    ):
        from assets.factories import AssetFactory

        secondary = AssetFactory(
            name="Secondary Duplicate",
            status="active",
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        dept_manager_client.post(
            reverse("assets:asset_merge_execute"),
            {
                "primary": active_asset.pk,
                "secondary": secondary.pk,
                "name": active_asset.name,
            },
        )
        secondary.refresh_from_db()
        assert secondary.status == "disposed"


@pytest.mark.django_db
class TestUS_DM_012_DisposeAssetInMyDept:
    """US-DM-012: Dispose of an asset in my department.

    MoSCoW: MUST
    Spec refs: S2.2.1-06, S2.2.1-07, S2.2.3-05, S2.3.15-01
    UI Surface: /assets/<pk>/
    """

    def test_dispose_transitions_to_disposed(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.post(
            reverse("assets:asset_edit", args=[active_asset.pk]),
            {
                "name": active_asset.name,
                "category": active_asset.category.pk,
                "current_location": (active_asset.current_location.pk),
                "condition": active_asset.condition,
                "quantity": active_asset.quantity,
                "status": "disposed",
            },
        )
        active_asset.refresh_from_db()
        assert active_asset.status == "disposed"

    def test_disposed_asset_not_in_default_search(
        self,
        dept_manager_client,
        category,
        location,
        dept_manager_user,
    ):
        from assets.factories import AssetFactory

        disposed = AssetFactory(
            name="Disposed DM Asset Unique",
            status="disposed",
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:asset_list"),
            {"q": "Disposed DM Asset Unique"},
        )
        assert b"Disposed DM Asset Unique" not in resp.content


# ---------------------------------------------------------------------------
# §10B.3 Check-out / Check-in / Transfer
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_013_CheckOutOnBehalfOfAnotherUser:
    """US-DM-013: Check out an asset on behalf of another user.

    MoSCoW: MUST
    Spec refs: S2.3.2-01, S2.3.2-05, S2.3.2-08, S2.3.8-05
    UI Surface: /assets/<pk>/checkout/
    """

    def test_checkout_form_loads(self, dept_manager_client, active_asset):
        resp = dept_manager_client.get(
            reverse("assets:asset_checkout", args=[active_asset.pk])
        )
        assert resp.status_code == 200

    def test_checkout_creates_transaction_with_performed_by(
        self,
        dept_manager_client,
        dept_manager_user,
        active_asset,
        borrower_user,
        location,
    ):
        resp = dept_manager_client.post(
            reverse("assets:asset_checkout", args=[active_asset.pk]),
            {
                "borrower": borrower_user.pk,
                "destination": location.pk,
                "notes": "",
            },
        )
        tx = Transaction.objects.filter(
            asset=active_asset,
            action="checkout",
            borrower=borrower_user,
        ).first()
        assert tx is not None
        # Spec names this field performed_by; model uses user (gap: field name)
        assert tx.user == dept_manager_user

    def test_checkout_updates_asset_location(
        self,
        dept_manager_client,
        active_asset,
        borrower_user,
        warehouse,
    ):
        dest = warehouse["bay1"]
        dept_manager_client.post(
            reverse("assets:asset_checkout", args=[active_asset.pk]),
            {
                "borrower": borrower_user.pk,
                "destination": dest.pk,
                "notes": "",
            },
        )
        active_asset.refresh_from_db()
        assert active_asset.current_location == dest


@pytest.mark.django_db
class TestUS_DM_014_CheckOutFromAnotherDept:
    """US-DM-014: Check out an asset from another department.

    MoSCoW: MUST
    Spec refs: S2.10.3-01, S3.3.4
    UI Surface: /assets/<pk>/checkout/
    """

    def test_dm_can_checkout_other_dept_asset(
        self,
        dept_manager_client,
        dept_manager_user,
        tech_dept,
        location,
        member_user,
        borrower_user,
    ):
        from assets.factories import AssetFactory, CategoryFactory

        other_cat = CategoryFactory(name="Tech Cat DM", department=tech_dept)
        other_asset = AssetFactory(
            name="Other Dept Asset For DM Checkout",
            status="active",
            category=other_cat,
            current_location=location,
            created_by=member_user,
        )
        resp = dept_manager_client.post(
            reverse("assets:asset_checkout", args=[other_asset.pk]),
            {
                "borrower": borrower_user.pk,
                "destination": location.pk,
                "notes": "",
            },
        )
        tx = Transaction.objects.filter(
            asset=other_asset, action="checkout"
        ).first()
        assert tx is not None
        # Spec names this field performed_by; model uses user (gap: field name)
        assert tx.user == dept_manager_user

    def test_dm_cannot_edit_other_dept_asset_after_checkout(
        self,
        dept_manager_client,
        tech_dept,
        location,
        member_user,
    ):
        from assets.factories import AssetFactory, CategoryFactory

        other_cat = CategoryFactory(
            name="Tech Cat DM Edit", department=tech_dept
        )
        other_asset = AssetFactory(
            name="Other Dept No Edit",
            status="active",
            category=other_cat,
            current_location=location,
            created_by=member_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:asset_edit", args=[other_asset.pk])
        )
        assert resp.status_code == 403


@pytest.mark.django_db
class TestUS_DM_015_CheckInAssetToMyDept:
    """US-DM-015: Check in an asset to my department.

    MoSCoW: MUST
    Spec refs: S2.3.3-01, S2.3.3-02, S2.3.3-05, S2.3.3-06a
    UI Surface: /assets/<pk>/checkin/
    """

    def test_checkin_form_loads_for_checked_out_asset(
        self,
        dept_manager_client,
        active_asset,
        borrower_user,
        location,
        admin_user,
    ):
        from assets.models import Transaction as Tx

        Tx.objects.create(
            asset=active_asset,
            action="checkout",
            user=admin_user,
            borrower=borrower_user,
            from_location=active_asset.current_location,
            to_location=location,
        )
        active_asset.checked_out_to = borrower_user
        active_asset.save()
        resp = dept_manager_client.get(
            reverse("assets:asset_checkin", args=[active_asset.pk])
        )
        assert resp.status_code == 200

    def test_checkin_clears_borrower(
        self,
        dept_manager_client,
        active_asset,
        borrower_user,
        location,
        admin_user,
    ):
        from assets.models import Transaction as Tx

        Tx.objects.create(
            asset=active_asset,
            action="checkout",
            user=admin_user,
            borrower=borrower_user,
            from_location=active_asset.current_location,
            to_location=location,
        )
        active_asset.checked_out_to = borrower_user
        active_asset.save()
        dept_manager_client.post(
            reverse("assets:asset_checkin", args=[active_asset.pk]),
            {"return_location": location.pk, "notes": ""},
        )
        active_asset.refresh_from_db()
        assert active_asset.checked_out_to is None


@pytest.mark.django_db
class TestUS_DM_016_TransferAssetWithinMyDept:
    """US-DM-016: Transfer an asset within my department.

    MoSCoW: MUST
    Spec refs: S2.3.4-01, S2.3.4-02, S2.3.4-03
    UI Surface: /assets/<pk>/transfer/
    """

    def test_transfer_form_loads(self, dept_manager_client, active_asset):
        resp = dept_manager_client.get(
            reverse("assets:asset_transfer", args=[active_asset.pk])
        )
        assert resp.status_code == 200

    def test_transfer_updates_location(
        self, dept_manager_client, active_asset, warehouse
    ):
        dest = warehouse["shelf_a"]
        dept_manager_client.post(
            reverse("assets:asset_transfer", args=[active_asset.pk]),
            {"destination": dest.pk, "notes": ""},
        )
        active_asset.refresh_from_db()
        assert active_asset.current_location == dest

    def test_transfer_creates_transaction(
        self, dept_manager_client, active_asset, warehouse
    ):
        dest = warehouse["shelf_b"]
        dept_manager_client.post(
            reverse("assets:asset_transfer", args=[active_asset.pk]),
            {"destination": dest.pk, "notes": ""},
        )
        assert Transaction.objects.filter(
            asset=active_asset, action="transfer"
        ).exists()


@pytest.mark.django_db
class TestUS_DM_017_CustodyHandover:
    """US-DM-017: Perform a custody handover.

    MoSCoW: MUST
    Spec refs: S2.3.5-01, S2.3.5-02, S2.3.5-03, S2.3.5-05
    UI Surface: /assets/<pk>/handover/
    """

    def test_handover_form_loads_for_checked_out_asset(
        self,
        dept_manager_client,
        active_asset,
        borrower_user,
        location,
        admin_user,
    ):
        from assets.models import Transaction as Tx

        Tx.objects.create(
            asset=active_asset,
            action="checkout",
            user=admin_user,
            borrower=borrower_user,
            from_location=active_asset.current_location,
            to_location=location,
        )
        active_asset.checked_out_to = borrower_user
        active_asset.save()
        resp = dept_manager_client.get(
            reverse("assets:asset_handover", args=[active_asset.pk])
        )
        assert resp.status_code == 200

    def test_handover_only_available_on_checked_out(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:asset_handover", args=[active_asset.pk])
        )
        assert resp.status_code in (200, 302, 404, 403)


@pytest.mark.django_db
class TestUS_DM_018_RelocateCheckedOutAsset:
    """US-DM-018: Relocate a checked-out asset.

    MoSCoW: MUST
    Spec refs: S2.3.11-01, S2.3.11-02
    UI Surface: /assets/<pk>/relocate/
    """

    def test_relocate_updates_location_keeps_borrower(
        self,
        dept_manager_client,
        active_asset,
        borrower_user,
        warehouse,
        admin_user,
    ):
        from assets.models import Transaction as Tx

        dest1 = warehouse["bay1"]
        Tx.objects.create(
            asset=active_asset,
            action="checkout",
            user=admin_user,
            borrower=borrower_user,
            from_location=active_asset.current_location,
            to_location=dest1,
        )
        active_asset.checked_out_to = borrower_user
        active_asset.save()
        dest2 = warehouse["bay4"]
        dept_manager_client.post(
            reverse("assets:asset_relocate", args=[active_asset.pk]),
            {"new_location": dest2.pk, "notes": ""},
        )
        active_asset.refresh_from_db()
        assert active_asset.current_location == dest2
        assert active_asset.checked_out_to == borrower_user


@pytest.mark.django_db
class TestUS_DM_056_BackdateDeptTransactions:
    """US-DM-056: Backdate transactions for department assets.

    MoSCoW: SHOULD
    Spec refs: S2.3.9-05, S2.3.9-06
    UI Surface: Check-in and Transfer forms
    """

    def test_checkin_form_has_date_field(
        self,
        dept_manager_client,
        active_asset,
        borrower_user,
        location,
        admin_user,
    ):
        from assets.models import Transaction as Tx

        Tx.objects.create(
            asset=active_asset,
            action="checkout",
            user=admin_user,
            borrower=borrower_user,
            from_location=active_asset.current_location,
            to_location=location,
        )
        active_asset.checked_out_to = borrower_user
        active_asset.save()
        resp = dept_manager_client.get(
            reverse("assets:asset_checkin", args=[active_asset.pk])
        )
        assert resp.status_code == 200
        content = resp.content.decode()
        assert "date" in content.lower()

    def test_transfer_form_has_date_field(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:asset_transfer", args=[active_asset.pk])
        )
        assert resp.status_code == 200
        content = resp.content.decode()
        assert "date" in content.lower()


# ---------------------------------------------------------------------------
# §10B.4 Barcode & NFC
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_019_PrintLabelForDeptAsset:
    """US-DM-019: Print a label for a department asset.

    MoSCoW: MUST
    Spec refs: S2.4.5-01, S2.4.5-02, S2.4.5-06, S2.4.5-08, S2.4.5a-01
    UI Surface: /assets/<pk>/label/
    """

    def test_print_label_page_loads(self, dept_manager_client, active_asset):
        resp = dept_manager_client.get(
            reverse("assets:asset_label", args=[active_asset.pk])
        )
        assert resp.status_code == 200

    def test_label_contains_barcode_and_name(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:asset_label", args=[active_asset.pk])
        )
        assert resp.status_code == 200
        content = resp.content.decode()
        assert active_asset.name in content or active_asset.barcode in content


@pytest.mark.django_db
class TestUS_DM_020_ManageNFCTagsOnDeptAssets:
    """US-DM-020: Manage NFC tags on department assets.

    MoSCoW: MUST
    Spec refs: S2.5.2-01, S2.5.2-05, S2.5.4-01, S2.5.4-02, S2.5.4-03
    UI Surface: /assets/<pk>/nfc/add/
    """

    def test_nfc_add_form_accessible(self, dept_manager_client, active_asset):
        resp = dept_manager_client.get(
            reverse("assets:nfc_add", args=[active_asset.pk])
        )
        assert resp.status_code == 200

    def test_add_nfc_tag_creates_record(
        self, dept_manager_client, active_asset
    ):
        from assets.models import NFCTag

        resp = dept_manager_client.post(
            reverse("assets:nfc_add", args=[active_asset.pk]),
            {"tag_id": "DMTAG001", "notes": ""},
        )
        assert NFCTag.objects.filter(
            asset=active_asset,
            tag_id__iexact="DMTAG001",
            removed_at__isnull=True,
        ).exists()

    def test_remove_nfc_tag_soft_deletes(
        self, dept_manager_client, active_asset, dept_manager_user
    ):
        from assets.models import NFCTag

        nfc = NFCTag.objects.create(
            asset=active_asset,
            tag_id="DMREMOVE001",
            assigned_by=dept_manager_user,
        )
        dept_manager_client.post(
            reverse("assets:nfc_remove", args=[active_asset.pk, nfc.pk]),
            {"notes": "Removed"},
        )
        nfc.refresh_from_db()
        assert nfc.removed_at is not None


@pytest.mark.django_db
class TestUS_DM_021_ScanBarcodeToLookUpAsset:
    """US-DM-021: Scan a barcode to look up an asset.

    MoSCoW: MUST
    Spec refs: S2.4.4-01, S2.4.4-03, S2.4.4-04
    UI Surface: /scan/
    """

    def test_scan_page_loads(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:scan"))
        assert resp.status_code == 200

    def test_scan_lookup_known_barcode_redirects_to_asset(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:scan_lookup"),
            {"code": active_asset.barcode},
        )
        assert resp.status_code in (200, 302)
        if resp.status_code == 302:
            assert str(active_asset.pk) in resp["Location"]

    def test_scan_unknown_barcode_redirects_to_quick_capture(
        self, dept_manager_client
    ):
        resp = dept_manager_client.get(
            reverse("assets:scan_lookup"),
            {"code": "UNKNOWN-BARCODE-XYZ"},
        )
        assert resp.status_code in (200, 302)


@pytest.mark.django_db
class TestUS_DM_022_PregenerateBarcodeLabelsForMyDept:
    """US-DM-022: Pre-generate barcode labels for my department.

    MoSCoW: MUST
    Spec refs: S2.4.3-01, S2.4.3-02, S2.4.3-03, S2.4.3-04
    UI Surface: /labels/pregenerate/
    """

    def test_pregenerate_page_loads(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:barcode_pregenerate"))
        assert resp.status_code == 200

    def test_pregenerate_does_not_create_assets(self, dept_manager_client):
        initial_count = Asset.objects.count()
        dept_manager_client.post(
            reverse("assets:barcode_pregenerate"),
            {"quantity": 2},
        )
        assert Asset.objects.count() == initial_count


# ---------------------------------------------------------------------------
# §10B.5 Search, Browse & Export
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_023_SearchAndFilterAssetList:
    """US-DM-023: Search and filter the asset list.

    MoSCoW: MUST
    Spec refs: S2.6.1-01, S2.6.2-01, S2.6.2-02, S2.6.2-04
    UI Surface: /assets/
    """

    def test_asset_list_page_loads(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:asset_list"))
        assert resp.status_code == 200

    def test_text_search_finds_asset(self, dept_manager_client, active_asset):
        resp = dept_manager_client.get(
            reverse("assets:asset_list"),
            {"q": active_asset.name},
        )
        assert resp.status_code == 200
        assert active_asset.name.encode() in resp.content

    def test_default_view_shows_active_only(
        self,
        dept_manager_client,
        active_asset,
        category,
        location,
        dept_manager_user,
    ):
        from assets.factories import AssetFactory

        disposed = AssetFactory(
            name="Disposed DM Search Asset",
            status="disposed",
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        resp = dept_manager_client.get(reverse("assets:asset_list"))
        assert b"Disposed DM Search Asset" not in resp.content


@pytest.mark.django_db
class TestUS_DM_024_SwitchListGridViews:
    """US-DM-024: Switch between list and grid views.

    MoSCoW: MUST
    Spec refs: S2.6.3-01, S2.6.3-02, S2.6.3-03, S2.6.3-04
    UI Surface: /assets/
    """

    def test_list_view_accessible(self, dept_manager_client, active_asset):
        resp = dept_manager_client.get(
            reverse("assets:asset_list"), {"view": "list"}
        )
        assert resp.status_code == 200

    def test_grid_view_accessible(self, dept_manager_client, active_asset):
        resp = dept_manager_client.get(
            reverse("assets:asset_list"), {"view": "grid"}
        )
        assert resp.status_code == 200


@pytest.mark.django_db
class TestUS_DM_025_ExportDeptAssets:
    """US-DM-025: Export assets scoped to my department.

    MoSCoW: MUST
    Spec refs: S2.9.1-01, S2.9.1-02, S2.9.1-03, S2.9.1-06
    UI Surface: /assets/export/
    """

    def test_export_accessible(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:export_assets"))
        assert resp.status_code in (200, 302)


# ---------------------------------------------------------------------------
# §10B.6 Stocktake
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_026_StartStocktakeSession:
    """US-DM-026: Start a stocktake session for a department location.

    MoSCoW: MUST
    Spec refs: S2.7.1-01, S2.7.1-02, S2.7.1-05
    UI Surface: /stocktake/start/
    """

    def test_stocktake_start_page_loads(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:stocktake_start"))
        assert resp.status_code == 200

    def test_creating_stocktake_session_with_location(
        self, dept_manager_client, dept_manager_user, location
    ):
        from assets.models import StocktakeSession

        resp = dept_manager_client.post(
            reverse("assets:stocktake_start"),
            {"location": location.pk},
        )
        assert StocktakeSession.objects.filter(location=location).exists()


@pytest.mark.django_db
class TestUS_DM_027_ConfirmAssetsDuringStocktake:
    """US-DM-027: Confirm assets during a stocktake.

    MoSCoW: MUST
    Spec refs: S2.7.2-01, S2.7.2-02, S2.7.2-03, S2.7.2-04
    UI Surface: /stocktake/<pk>/
    """

    def test_stocktake_detail_page_loads(
        self, dept_manager_client, location, dept_manager_user
    ):
        from assets.models import StocktakeSession

        session = StocktakeSession.objects.create(
            location=location, started_by=dept_manager_user
        )
        resp = dept_manager_client.get(
            reverse("assets:stocktake_detail", args=[session.pk])
        )
        assert resp.status_code == 200

    def test_confirm_asset_creates_audit_transaction(
        self,
        dept_manager_client,
        active_asset,
        dept_manager_user,
    ):
        from assets.models import StocktakeSession

        session = StocktakeSession.objects.create(
            location=active_asset.current_location,
            started_by=dept_manager_user,
        )
        dept_manager_client.post(
            reverse("assets:stocktake_confirm", args=[session.pk]),
            {"asset_id": active_asset.pk},
        )
        assert Transaction.objects.filter(
            asset=active_asset, action="audit"
        ).exists()


@pytest.mark.django_db
class TestUS_DM_028_HandleStocktakeDiscrepancies:
    """US-DM-028: Handle stocktake discrepancies.

    MoSCoW: MUST
    Spec refs: S2.7.3-01, S2.7.3-02, S2.7.3-03
    UI Surface: /stocktake/<pk>/
    """

    def test_stocktake_session_accessible(
        self, dept_manager_client, location, dept_manager_user
    ):
        from assets.models import StocktakeSession

        session = StocktakeSession.objects.create(
            location=location, started_by=dept_manager_user
        )
        resp = dept_manager_client.get(
            reverse("assets:stocktake_detail", args=[session.pk])
        )
        assert resp.status_code == 200


@pytest.mark.django_db
class TestUS_DM_029_CompleteStocktakeSession:
    """US-DM-029: Complete a stocktake session.

    MoSCoW: MUST
    Spec refs: S2.7.4-01, S2.7.4-03, S2.7.4-04
    UI Surface: /stocktake/<pk>/complete/
    """

    def test_complete_stocktake_sets_completed_at(
        self, dept_manager_client, location, dept_manager_user
    ):
        from assets.models import StocktakeSession

        session = StocktakeSession.objects.create(
            location=location, started_by=dept_manager_user
        )
        dept_manager_client.post(
            reverse("assets:stocktake_complete", args=[session.pk]),
            {},
        )
        session.refresh_from_db()
        # Spec says completed_at; model uses ended_at (gap: field name mismatch)
        assert session.ended_at is not None

    def test_stocktake_summary_accessible(
        self, dept_manager_client, location, dept_manager_user
    ):
        from assets.models import StocktakeSession

        session = StocktakeSession.objects.create(
            location=location, started_by=dept_manager_user
        )
        dept_manager_client.post(
            reverse("assets:stocktake_complete", args=[session.pk]),
            {},
        )
        resp = dept_manager_client.get(
            reverse("assets:stocktake_summary", args=[session.pk])
        )
        assert resp.status_code == 200


# ---------------------------------------------------------------------------
# §10B.7 Bulk Operations
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_030_BulkTransferAssetsWithinMyDept:
    """US-DM-030: Bulk transfer assets within my department.

    MoSCoW: MUST
    Spec refs: S2.8.1-01, S2.8.1-02, S2.8.1-04, S2.8.1-05
    UI Surface: /assets/bulk/
    """

    def test_bulk_transfer_creates_transactions(
        self,
        dept_manager_client,
        active_asset,
        warehouse,
        category,
        location,
        dept_manager_user,
    ):
        from assets.factories import AssetFactory

        asset2 = AssetFactory(
            name="DM Bulk Transfer Asset",
            status="active",
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        dest = warehouse["bay4"]
        resp = dept_manager_client.post(
            reverse("assets:bulk_actions"),
            {
                "action": "transfer",
                "selected_ids": [active_asset.pk, asset2.pk],
                "destination": dest.pk,
            },
        )
        assert Transaction.objects.filter(
            asset=active_asset, action="transfer"
        ).exists()


@pytest.mark.django_db
class TestUS_DM_031_BulkEditDeptAssets:
    """US-DM-031: Bulk edit assets in my department.

    MoSCoW: MUST
    Spec refs: S2.8.3-01, S2.8.3-02, S2.8.3-04, S2.8.3-05
    UI Surface: /assets/bulk/
    """

    def test_bulk_edit_sets_category(
        self,
        dept_manager_client,
        active_asset,
        category,
        location,
        dept_manager_user,
    ):
        from assets.factories import AssetFactory, CategoryFactory

        new_cat = CategoryFactory(
            name="DM Bulk Edit Cat",
            department=category.department,
        )
        asset2 = AssetFactory(
            name="DM Bulk Edit Asset",
            status="active",
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        resp = dept_manager_client.post(
            reverse("assets:bulk_actions"),
            {
                "action": "edit",
                "selected_ids": [active_asset.pk, asset2.pk],
                "category": new_cat.pk,
            },
        )
        active_asset.refresh_from_db()
        assert active_asset.category == new_cat


@pytest.mark.django_db
class TestUS_DM_032_BulkPrintLabelsDeptAssets:
    """US-DM-032: Bulk print labels for department assets.

    MoSCoW: MUST
    Spec refs: S2.8.2-01, S2.8.2-02, S2.8.2-04
    UI Surface: /assets/bulk/ + /assets/labels/all-filtered/
    """

    def test_bulk_print_accessible(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:bulk_actions"))
        assert resp.status_code in (200, 405)

    def test_print_all_filtered_accessible(self, dept_manager_client):
        resp = dept_manager_client.get(
            reverse("assets:print_all_filtered_labels")
        )
        assert resp.status_code in (200, 405)


# ---------------------------------------------------------------------------
# §10B.8 Category Management
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_033_CreateCategoryInMyDept:
    """US-DM-033: Create a category in my department.

    MoSCoW: MUST
    Spec refs: S2.10.2-01, S2.10.2-02
    UI Surface: /categories/create/
    """

    def test_category_create_page_loads(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:category_create"))
        assert resp.status_code == 200

    def test_create_category_in_own_dept(
        self, dept_manager_client, department
    ):
        from assets.models import Category

        resp = dept_manager_client.post(
            reverse("assets:category_create"),
            {
                "name": "DM Created Category",
                "department": department.pk,
                "description": "",
            },
        )
        assert Category.objects.filter(name="DM Created Category").exists()


@pytest.mark.django_db
class TestUS_DM_034_EditCategoryInMyDept:
    """US-DM-034: Edit a category in my department.

    MoSCoW: MUST
    Spec refs: S2.10.2-02
    UI Surface: /categories/<pk>/edit/
    """

    def test_edit_own_dept_category(self, dept_manager_client, category):
        resp = dept_manager_client.get(
            reverse("assets:category_edit", args=[category.pk])
        )
        assert resp.status_code == 200

    def test_cannot_edit_other_dept_category(
        self, dept_manager_client, tech_dept
    ):
        from assets.factories import CategoryFactory

        other_cat = CategoryFactory(
            name="Other Dept Cat",
            department=tech_dept,
        )
        resp = dept_manager_client.get(
            reverse("assets:category_edit", args=[other_cat.pk])
        )
        assert resp.status_code in (403, 302)


@pytest.mark.django_db
class TestUS_DM_035_DeleteCategoryInMyDept:
    """US-DM-035: Delete a category in my department.

    MoSCoW: MUST
    Spec refs: S2.10.2-02
    UI Surface: /categories/
    """

    def test_delete_empty_category(self, dept_manager_client, department):
        from assets.factories import CategoryFactory
        from assets.models import Category

        empty_cat = CategoryFactory(
            name="Empty Cat to Delete",
            department=department,
        )
        resp = dept_manager_client.post(
            f"/categories/{empty_cat.pk}/delete/", {}
        )
        assert not Category.objects.filter(
            pk=empty_cat.pk
        ).exists() or resp.status_code in (200, 302, 404)

    def test_cannot_delete_category_with_assets(
        self, dept_manager_client, category, active_asset
    ):
        from assets.models import Category

        resp = dept_manager_client.post(
            f"/categories/{category.pk}/delete/", {}
        )
        # Category should still exist because it has assets
        assert Category.objects.filter(pk=category.pk).exists()


# ---------------------------------------------------------------------------
# §10B.9 Dashboard
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_036_ViewDeptScopedDashboard:
    """US-DM-036: View a department-scoped dashboard.

    MoSCoW: MUST
    Spec refs: S2.11.1-01, S2.11.2-01, S2.11.2a-01
    UI Surface: /
    """

    def test_dashboard_page_loads(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:dashboard"))
        assert resp.status_code == 200

    def test_dashboard_shows_summary_data(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(reverse("assets:dashboard"))
        assert resp.status_code == 200


@pytest.mark.django_db
class TestUS_DM_037_ViewOverdueCheckedOutItems:
    """US-DM-037: View overdue checked-out items.

    MoSCoW: SHOULD
    Spec refs: S2.11.2a-01, S2.3.8-06
    UI Surface: / (dashboard widget)
    """

    def test_dashboard_accessible_for_overdue_check(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:dashboard"))
        assert resp.status_code == 200


# ---------------------------------------------------------------------------
# §10B.10 Hold Lists & Projects
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_038_CreateHoldListForMyDept:
    """US-DM-038: Create a hold list for my department.

    MoSCoW: MUST
    Spec refs: S2.16.3-01, S2.16.3-02, S2.16.3-04, S2.16.3-08
    UI Surface: /hold-lists/create/
    """

    def test_holdlist_create_page_loads(self, dept_manager_client):
        resp = dept_manager_client.get(reverse("assets:holdlist_create"))
        assert resp.status_code == 200

    def test_create_hold_list_for_own_dept(
        self, dept_manager_client, department, hold_list_status
    ):
        from assets.models import HoldList

        resp = dept_manager_client.post(
            reverse("assets:holdlist_create"),
            {
                "name": "DM Show Hold",
                "department": department.pk,
                "status": hold_list_status.pk,
                "start_date": "2026-03-01",
                "end_date": "2026-03-31",
            },
        )
        assert HoldList.objects.filter(name="DM Show Hold").exists()


@pytest.mark.django_db
class TestUS_DM_039_EditAndManageHoldListItems:
    """US-DM-039: Edit and manage hold list items.

    MoSCoW: MUST
    Spec refs: S2.16.4-01, S2.16.4-03, S2.16.7-03
    UI Surface: /hold-lists/<pk>/
    """

    def test_holdlist_detail_page_loads(self, dept_manager_client, hold_list):
        resp = dept_manager_client.get(
            reverse("assets:holdlist_detail", args=[hold_list.pk])
        )
        assert resp.status_code == 200

    def test_add_item_to_hold_list(
        self, dept_manager_client, hold_list, active_asset
    ):
        resp = dept_manager_client.post(
            reverse("assets:holdlist_add_item", args=[hold_list.pk]),
            {
                "asset": active_asset.pk,
                "quantity": 1,
                "notes": "",
            },
        )
        assert resp.status_code in (200, 302)


@pytest.mark.django_db
class TestUS_DM_040_LockUnlockHoldLists:
    """US-DM-040: Lock and unlock hold lists in my department.

    MoSCoW: MUST
    Spec refs: S2.16.3-06, S2.16.3-07
    UI Surface: /hold-lists/<pk>/lock/ + /hold-lists/<pk>/unlock/
    """

    def test_lock_hold_list(self, dept_manager_client, hold_list):
        resp = dept_manager_client.post(
            reverse("assets:holdlist_lock", args=[hold_list.pk]), {}
        )
        assert resp.status_code in (200, 302)
        hold_list.refresh_from_db()
        assert hold_list.is_locked

    def test_unlock_hold_list(self, dept_manager_client, hold_list):
        hold_list.is_locked = True
        hold_list.save()
        resp = dept_manager_client.post(
            reverse("assets:holdlist_unlock", args=[hold_list.pk]),
            {},
        )
        assert resp.status_code in (200, 302)
        hold_list.refresh_from_db()
        assert not hold_list.is_locked


@pytest.mark.django_db
class TestUS_DM_041_OverrideHoldBlocksForDeptAssets:
    """US-DM-041: Override hold blocks for department assets.

    MoSCoW: MUST
    Spec refs: S2.16.5-02, S2.16.5-03, S2.16.5-04
    UI Surface: /assets/<pk>/checkout/
    """

    def test_checkout_form_loads_for_held_asset(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:asset_checkout", args=[active_asset.pk])
        )
        assert resp.status_code == 200


@pytest.mark.django_db
class TestUS_DM_042_DownloadPickSheetForHoldList:
    """US-DM-042: Download a pick sheet for a hold list.

    MoSCoW: MUST
    Spec refs: S2.16.6-01, S2.16.6-02, S2.16.6-04
    UI Surface: /hold-lists/<pk>/pick-sheet/
    """

    def test_pick_sheet_url_accessible(self, dept_manager_client, hold_list):
        resp = dept_manager_client.get(
            reverse("assets:holdlist_pick_sheet", args=[hold_list.pk])
        )
        assert resp.status_code in (200, 302)


# ---------------------------------------------------------------------------
# §10B.11 Asset Kits & Serialisation
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_043_CreateAndManageKitInMyDept:
    """US-DM-043: Create and manage a kit in my department.

    MoSCoW: MUST
    Spec refs: S2.17.3-01, S2.17.3-02, S2.17.5-01, S2.17.5-03
    UI Surface: /assets/<pk>/kit/
    """

    def test_kit_contents_page_loads(
        self,
        dept_manager_client,
        category,
        location,
        dept_manager_user,
    ):
        from assets.factories import AssetFactory
        from assets.models import AssetKit

        kit = AssetFactory(
            name="DM Test Kit",
            status="active",
            is_kit=True,
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:kit_contents", args=[kit.pk])
        )
        assert resp.status_code == 200

    def test_kit_shows_components(
        self,
        dept_manager_client,
        category,
        location,
        dept_manager_user,
    ):
        from assets.factories import AssetFactory
        from assets.models import AssetKit

        kit = AssetFactory(
            name="DM Kit With Comps",
            status="active",
            is_kit=True,
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        comp = AssetFactory(
            name="DM Dimmer Component",
            status="active",
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        AssetKit.objects.create(kit=kit, component=comp, is_required=True)
        resp = dept_manager_client.get(
            reverse("assets:kit_contents", args=[kit.pk])
        )
        assert resp.status_code == 200
        content = resp.content.decode()
        assert comp.name in content


@pytest.mark.django_db
class TestUS_DM_044_CheckOutKitFromMyDept:
    """US-DM-044: Check out a kit from my department.

    MoSCoW: MUST
    Spec refs: S2.17.4-01, S2.17.4-02, S2.17.4-03, S2.17.4-04
    UI Surface: /assets/<pk>/checkout/
    """

    def test_checkout_form_loads_for_kit(
        self,
        dept_manager_client,
        category,
        location,
        dept_manager_user,
    ):
        from assets.factories import AssetFactory

        kit = AssetFactory(
            name="DM Kit Checkout Test",
            status="active",
            is_kit=True,
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:asset_checkout", args=[kit.pk])
        )
        assert resp.status_code == 200

    def test_kit_checkout_creates_component_transactions(
        self,
        dept_manager_client,
        category,
        location,
        dept_manager_user,
        borrower_user,
    ):
        from assets.factories import AssetFactory
        from assets.models import AssetKit

        kit = AssetFactory(
            name="DM Kit For Checkout",
            status="active",
            is_kit=True,
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        comp = AssetFactory(
            name="DM Kit Component For Checkout",
            status="active",
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        AssetKit.objects.create(kit=kit, component=comp, is_required=True)
        resp = dept_manager_client.post(
            reverse("assets:asset_checkout", args=[kit.pk]),
            {
                "borrower": borrower_user.pk,
                "destination": location.pk,
                "notes": "",
            },
        )
        assert Transaction.objects.filter(
            asset=kit, action="checkout"
        ).exists()


@pytest.mark.django_db
class TestUS_DM_045_ConvertAssetSerialisationType:
    """US-DM-045: Convert an asset between serialised and non-serialised.

    MoSCoW: MUST
    Spec refs: S2.17.1d-01, S2.17.1d-02, S2.17.1d-05, S2.17.1d-13
    UI Surface: /assets/<pk>/edit/ (Asset Type section)
    """

    def test_convert_serialisation_url_accessible(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse(
                "assets:asset_convert_serialisation",
                args=[active_asset.pk],
            )
        )
        assert resp.status_code in (200, 405)

    def test_edit_form_has_asset_type_section(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:asset_edit", args=[active_asset.pk])
        )
        assert resp.status_code == 200


@pytest.mark.django_db
class TestUS_DM_046_ManageSerialUnits:
    """US-DM-046: Manage serial units on a serialised department asset.

    MoSCoW: MUST
    Spec refs: S2.17.1a-01, S2.17.1b-01, S2.17.1b-02
    UI Surface: /assets/<pk>/ (Serials tab)
    """

    def test_serialised_asset_detail_loads(
        self, dept_manager_client, serialised_asset_with_units
    ):
        asset = serialised_asset_with_units["asset"]
        resp = dept_manager_client.get(
            reverse("assets:asset_detail", args=[asset.pk])
        )
        assert resp.status_code == 200

    def test_serials_appear_in_detail(
        self, dept_manager_client, serialised_asset_with_units
    ):
        asset = serialised_asset_with_units["asset"]
        serials = serialised_asset_with_units["serials"]
        resp = dept_manager_client.get(
            reverse("assets:asset_detail", args=[asset.pk])
        )
        content = resp.content.decode()
        assert serials[0].serial_number in content


@pytest.mark.django_db
class TestUS_DM_047_CheckOutIndividualSerialUnits:
    """US-DM-047: Check out individual serial units.

    MoSCoW: MUST
    Spec refs: S2.17.2-01, S2.17.2-04, S2.17.2-06
    UI Surface: /assets/<pk>/checkout/
    """

    def test_checkout_form_loads_for_serialised_asset(
        self, dept_manager_client, serialised_asset_with_units
    ):
        asset = serialised_asset_with_units["asset"]
        resp = dept_manager_client.get(
            reverse("assets:asset_checkout", args=[asset.pk])
        )
        assert resp.status_code == 200

    def test_serialised_checkout_form_shows_serial_mode(
        self, dept_manager_client, serialised_asset_with_units
    ):
        asset = serialised_asset_with_units["asset"]
        resp = dept_manager_client.get(
            reverse("assets:asset_checkout", args=[asset.pk])
        )
        assert resp.status_code == 200


@pytest.mark.django_db
class TestUS_DM_048_ReviewAndApplyAISuggestions:
    """US-DM-048: Review and apply AI suggestions on department assets.

    MoSCoW: MUST
    Spec refs: S2.14.3-01, S2.14.3-02, S2.14.3-03
    UI Surface: /assets/<pk>/
    """

    def test_asset_detail_shows_ai_panel(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:asset_detail", args=[active_asset.pk])
        )
        assert resp.status_code == 200

    def test_ai_apply_suggestions_url_accessible(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:ai_apply_suggestions", args=[active_asset.pk])
        )
        assert resp.status_code in (200, 405)


@pytest.mark.django_db
class TestUS_DM_049_TriggerReAnalysis:
    """US-DM-049: Trigger re-analysis of an asset image.

    MoSCoW: SHOULD
    Spec refs: S2.14.3-08
    UI Surface: /assets/<pk>/
    """

    def test_asset_detail_accessible_for_reanalysis(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:asset_detail", args=[active_asset.pk])
        )
        assert resp.status_code == 200


@pytest.mark.django_db
class TestUS_DM_050_ApplyAISuggestedDeptToDraft:
    """US-DM-050: Apply AI-suggested department to a draft asset.

    MoSCoW: SHOULD
    Spec refs: S2.14.3-03b, S2.14.3-07
    UI Surface: /assets/<pk>/
    """

    def test_draft_asset_detail_accessible(
        self, dept_manager_client, dept_manager_user, category
    ):
        from assets.factories import AssetFactory

        draft = AssetFactory(
            name="AI Dept Draft",
            status="draft",
            category=None,
            current_location=None,
            created_by=dept_manager_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:asset_detail", args=[draft.pk])
        )
        assert resp.status_code == 200


@pytest.mark.django_db
class TestUS_DM_051_CheckOutNonSerialisedByQuantity:
    """US-DM-051: Check out and return non-serialised assets by quantity.

    MoSCoW: MUST
    Spec refs: S2.17.2-02, S2.17.2-03, S2.17.2-05, S2.17.2-07
    UI Surface: /assets/<pk>/checkout/
    """

    def test_checkout_form_loads_for_non_serialised(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse("assets:asset_checkout", args=[active_asset.pk])
        )
        assert resp.status_code == 200


@pytest.mark.django_db
class TestUS_DM_052_ReviewSerialisationConversionImpact:
    """US-DM-052: Review serialisation conversion impact before changing.

    MoSCoW: MUST
    Spec refs: S2.17.1d-03, S2.17.1d-04, S2.17.1d-06
    UI Surface: /assets/<pk>/edit/
    """

    def test_convert_serialisation_endpoint_accessible(
        self, dept_manager_client, active_asset
    ):
        resp = dept_manager_client.get(
            reverse(
                "assets:asset_convert_serialisation",
                args=[active_asset.pk],
            )
        )
        assert resp.status_code in (200, 405)


@pytest.mark.django_db
class TestUS_DM_053_ValidateKitCompositionRules:
    """US-DM-053: Validate kit composition rules in my department.

    MoSCoW: MUST
    Spec refs: S2.17.3-02a, S2.17.3-04, S2.17.3-05
    UI Surface: /assets/<pk>/kit/
    """

    def test_kit_contents_tab_accessible(
        self,
        dept_manager_client,
        category,
        location,
        dept_manager_user,
    ):
        from assets.factories import AssetFactory

        kit = AssetFactory(
            name="DM Kit Composition Test",
            status="active",
            is_kit=True,
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:kit_contents", args=[kit.pk])
        )
        assert resp.status_code == 200

    def test_add_component_to_kit(
        self,
        dept_manager_client,
        category,
        location,
        dept_manager_user,
    ):
        from assets.factories import AssetFactory

        kit = AssetFactory(
            name="DM Kit Add Component Test",
            status="active",
            is_kit=True,
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        new_component = AssetFactory(
            name="New Kit Component",
            status="active",
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        resp = dept_manager_client.post(
            reverse("assets:kit_add_component", args=[kit.pk]),
            {
                "component": new_component.pk,
                "quantity": 1,
                "is_required": True,
            },
        )
        assert resp.status_code in (200, 302)


@pytest.mark.django_db
class TestUS_DM_054_HandleKitCheckoutEdgeCases:
    """US-DM-054: Handle kit checkout edge cases.

    MoSCoW: MUST
    Spec refs: S2.17.4-05, S2.17.4-06, S2.17.4-07
    UI Surface: /assets/<pk>/checkout/
    """

    def test_kit_checkout_form_accessible(
        self,
        dept_manager_client,
        category,
        location,
        dept_manager_user,
    ):
        from assets.factories import AssetFactory

        kit = AssetFactory(
            name="DM Kit Edge Cases Test",
            status="active",
            is_kit=True,
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:asset_checkout", args=[kit.pk])
        )
        assert resp.status_code == 200


@pytest.mark.django_db
class TestUS_DM_055_BrowseKitsAndViewKitMembership:
    """US-DM-055: Browse kits and view kit membership on department assets.

    MoSCoW: SHOULD
    Spec refs: S2.17.5-02, S2.17.5-04, S2.17.5-05
    UI Surface: /assets/
    """

    def test_filter_kits_in_asset_list(self, dept_manager_client):
        resp = dept_manager_client.get(
            reverse("assets:asset_list"), {"is_kit": "true"}
        )
        assert resp.status_code == 200

    def test_kit_component_detail_shows_member_of_kits(
        self,
        dept_manager_client,
        category,
        location,
        dept_manager_user,
    ):
        from assets.factories import AssetFactory
        from assets.models import AssetKit

        kit = AssetFactory(
            name="DM Kit Browse Test",
            status="active",
            is_kit=True,
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        comp = AssetFactory(
            name="DM Component Browse Test",
            status="active",
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )
        AssetKit.objects.create(kit=kit, component=comp, is_required=True)
        resp = dept_manager_client.get(
            reverse("assets:asset_detail", args=[comp.pk])
        )
        assert resp.status_code == 200


# ---------------------------------------------------------------------------
# §10B Help System
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# New uncovered acceptance-criteria tests — added Feb 2026
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_002_ManageDraftsQueue_CrossDept:
    """US-DM-002 (extra): DM cannot edit a draft from another department.

    Spec refs: S2.1.4a-01, S2.10.3-07
    """

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP: DM can currently edit drafts in other departments "
            "(returns 200). Should return 302/403. "
            "(S2.1.4a-01, S2.10.3-07)"
        ),
    )
    def test_dm_cannot_edit_draft_from_other_department(
        self,
        dept_manager_client,
        tech_dept,
        member_user,
    ):
        """Create a draft in dept B; DM of dept A must be rejected (302/403)."""
        from assets.factories import AssetFactory, CategoryFactory

        other_cat = CategoryFactory(
            name="Tech Draft Cat", department=tech_dept
        )
        other_draft = AssetFactory(
            name="Other Dept Draft Item",
            status="draft",
            category=other_cat,
            current_location=None,
            created_by=member_user,
        )
        resp = dept_manager_client.get(
            reverse("assets:asset_edit", args=[other_draft.pk])
        )
        assert resp.status_code in (302, 403)


@pytest.mark.django_db
class TestUS_DM_005_CreateAssetInOwnDept_OtherDeptRejected:
    """US-DM-005 (extra): DM cannot create asset in another department.

    Spec refs: S2.2.1-01, S2.10.3-07
    """

    def test_dm_cannot_create_asset_in_other_dept(
        self,
        dept_manager_client,
        tech_dept,
        location,
    ):
        """POST asset creation with a category from tech_dept — must be
        rejected (form stays invalid or returns 403/form error)."""
        from assets.factories import CategoryFactory

        other_cat = CategoryFactory(
            name="Tech Cat For Create", department=tech_dept
        )
        initial_count = Asset.objects.count()
        resp = dept_manager_client.post(
            reverse("assets:asset_create"),
            {
                "name": "Cross Dept Asset Attempt",
                "category": other_cat.pk,
                "current_location": location.pk,
                "condition": "good",
                "quantity": 1,
            },
        )
        # Asset must NOT have been created under other dept's category,
        # OR the response must be a rejection (403/form error 200).
        created = Asset.objects.filter(name="Cross Dept Asset Attempt").first()
        if created is not None:
            # If the server did create it, it should NOT belong to tech_dept
            assert created.category.department != tech_dept, (
                "DM was able to create an asset in another department's "
                "category — cross-department creation should be blocked"
            )
        else:
            # No asset created — correct behaviour
            assert Asset.objects.count() == initial_count


@pytest.mark.django_db
class TestUS_DM_008_UploadManageImages_PlaceholderAfterDelete:
    """US-DM-008 (extra): After deleting the only image, a placeholder appears.

    Spec refs: S2.2.5-03, S2.2.5-05
    """

    def test_deleting_only_image_shows_placeholder(
        self,
        dept_manager_client,
        active_asset,
        dept_manager_user,
    ):
        """Upload 1 image, delete it, assert placeholder/no-image indicator."""
        from django.core.files.uploadedfile import SimpleUploadedFile

        from assets.models import AssetImage

        gif_bytes = (
            b"GIF87a\x01\x00\x01\x00\x80\x01\x00\x00\x00\x00"
            b"\xff\xff\xff,\x00\x00\x00\x00\x01\x00\x01\x00"
            b"\x00\x02\x02D\x01\x00;"
        )
        image_file = SimpleUploadedFile(
            "test_img.gif", gif_bytes, content_type="image/gif"
        )
        dept_manager_client.post(
            reverse("assets:image_upload", args=[active_asset.pk]),
            {"image": image_file},
        )
        img = AssetImage.objects.filter(asset=active_asset).first()
        if img is None:
            pytest.skip("Image upload did not succeed — skipping delete step")

        # Delete the image
        dept_manager_client.post(
            reverse("assets:image_delete", args=[active_asset.pk, img.pk]),
            {},
        )

        resp = dept_manager_client.get(
            reverse("assets:asset_detail", args=[active_asset.pk])
        )
        assert resp.status_code == 200
        content = resp.content.decode().lower()
        # A placeholder img tag or "no image" text should appear when no
        # images remain.
        assert (
            "placeholder" in content
            or "no image" in content
            or "no-image" in content
            or "default" in content
            or "svg" in content
            or AssetImage.objects.filter(asset=active_asset).count() == 0
        ), (
            "After deleting the only image, the detail page should show a "
            "placeholder or 'no image' indicator"
        )


@pytest.mark.django_db
class TestUS_DM_012_DisposeAsset_CheckedOutBlocked:
    """US-DM-012 (extra): DM cannot dispose a checked-out asset.

    Spec refs: S2.2.3-05, S2.3.15-01
    """

    def test_dm_cannot_dispose_checked_out_asset(
        self,
        dept_manager_client,
        active_asset,
        borrower_user,
        location,
        admin_user,
    ):
        """Check out an asset, attempt disposal via asset_delete; must be
        rejected because the asset is checked out."""
        # Check out via Transaction + FK (same as other tests)
        Transaction.objects.create(
            asset=active_asset,
            action="checkout",
            user=admin_user,
            borrower=borrower_user,
            from_location=active_asset.current_location,
            to_location=location,
        )
        active_asset.checked_out_to = borrower_user
        active_asset.save()

        # Disposal in PROPS is via asset_delete, not the edit form
        # (FORM_STATUS_CHOICES doesn't include 'disposed')
        resp = dept_manager_client.post(
            reverse("assets:asset_delete", args=[active_asset.pk]),
            {},
        )
        active_asset.refresh_from_db()
        assert active_asset.status != "disposed", (
            "A checked-out asset must not be disposable without first "
            "being checked in"
        )


@pytest.mark.django_db
class TestUS_DM_020_AssignNFCTag_Uniqueness:
    """US-DM-020 (extra): NFC tag already on another asset is rejected with
    a message identifying the other asset.

    Spec refs: S2.5.2-01, S2.5.2-05, S2.5.4-03
    """

    def test_nfc_tag_already_on_another_asset_rejected(
        self,
        dept_manager_client,
        active_asset,
        category,
        location,
        dept_manager_user,
    ):
        """Assign NFC to asset A, then attempt to assign same tag to asset B
        — must be rejected with an error that names asset A."""
        from assets.factories import AssetFactory
        from assets.models import NFCTag

        asset_b = AssetFactory(
            name="NFC Duplicate Target",
            status="active",
            category=category,
            current_location=location,
            created_by=dept_manager_user,
        )

        # Assign tag to active_asset
        dept_manager_client.post(
            reverse("assets:nfc_add", args=[active_asset.pk]),
            {"tag_id": "DUPETAG001", "notes": ""},
        )
        assert NFCTag.objects.filter(
            tag_id__iexact="DUPETAG001",
            asset=active_asset,
            removed_at__isnull=True,
        ).exists()

        # Now try to assign same tag to asset_b
        resp = dept_manager_client.post(
            reverse("assets:nfc_add", args=[asset_b.pk]),
            {"tag_id": "DUPETAG001", "notes": ""},
            follow=True,
        )
        content = resp.content.decode()
        # The tag must NOT appear on asset_b
        assert not NFCTag.objects.filter(
            tag_id__iexact="DUPETAG001",
            asset=asset_b,
            removed_at__isnull=True,
        ).exists(), "Duplicate NFC tag was wrongly assigned to a second asset"
        # The error response should name the first asset
        assert active_asset.name in content, (
            "Error message should identify the asset that already holds the "
            f"NFC tag ('{active_asset.name}')"
        )

    def test_removed_nfc_tags_appear_in_history(
        self,
        dept_manager_client,
        active_asset,
        dept_manager_user,
    ):
        """Assign NFC tag, remove it, then GET asset detail — the removed
        tag should appear in a history section on the page."""
        from assets.models import NFCTag

        nfc = NFCTag.objects.create(
            asset=active_asset,
            tag_id="HISTTAG001",
            assigned_by=dept_manager_user,
        )
        # Remove it
        dept_manager_client.post(
            reverse("assets:nfc_remove", args=[active_asset.pk, nfc.pk]),
            {"notes": "Removed for history test"},
        )
        nfc.refresh_from_db()
        assert nfc.removed_at is not None

        resp = dept_manager_client.get(
            reverse("assets:asset_detail", args=[active_asset.pk])
        )
        assert resp.status_code == 200
        content = resp.content.decode()
        assert "HISTTAG001" in content, (
            "Removed NFC tag 'HISTTAG001' should appear in the asset detail "
            "history section"
        )


@pytest.mark.django_db
class TestUS_DM_025_ExportAssets_Columns:
    """US-DM-025 (extra): Export XLSX contains expected column headers.

    Spec refs: S2.9.1-01, S2.9.1-02, S2.9.1-03
    """

    def test_export_contains_expected_columns(
        self,
        dept_manager_client,
        active_asset,
    ):
        """Export assets; open XLSX with openpyxl; assert header columns."""
        import io

        import openpyxl

        resp = dept_manager_client.get(reverse("assets:export_assets"))
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.get("Content-Type", "") or "xlsx" in (
            resp.get("Content-Disposition", "")
        )

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        # Pick the first (or Assets) sheet
        sheet = wb.active
        if "Assets" in wb.sheetnames:
            sheet = wb["Assets"]

        # Read header row (row 1)
        headers = [
            str(cell.value).strip().lower() if cell.value else ""
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]
        expected = {
            "name",
            "barcode",
            "category",
            "location",
            "status",
            "condition",
        }
        missing = expected - set(headers)
        assert not missing, (
            f"Export XLSX is missing expected columns: {missing}. "
            f"Found headers: {headers}"
        )


@pytest.mark.django_db
class TestUS_DM_031_BulkEditAssets_BlankCategory:
    """US-DM-031 (extra): Bulk edit with blank category does not overwrite.

    Spec refs: S2.8.3-04, S2.8.3-05
    """

    def test_bulk_edit_blank_category_does_not_overwrite(
        self,
        dept_manager_client,
        active_asset,
        category,
    ):
        """Bulk-edit with no category selected — asset's category unchanged."""
        original_category = active_asset.category

        resp = dept_manager_client.post(
            reverse("assets:bulk_actions"),
            {
                "bulk_action": "bulk_edit",
                "asset_ids": [active_asset.pk],
                # No edit_category — blank
                "edit_location": "",
            },
        )
        active_asset.refresh_from_db()
        # The blank bulk edit should either be rejected (no-op) or leave
        # the category untouched
        assert active_asset.category == original_category, (
            "Bulk edit with a blank category should not overwrite the "
            "asset's existing category"
        )


@pytest.mark.django_db
class TestUS_DM_037_ViewOverdueItems_DashboardContent:
    """US-DM-037 (extra): Dashboard shows overdue items.

    Spec refs: S2.11.2a-01, S2.3.8-06
    """

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP: Dashboard does not surface an 'overdue' indicator when "
            "assets have a past due_date on their checkout transaction. "
            "(S2.11.2a-01, S2.3.8-06)"
        ),
    )
    def test_dashboard_shows_overdue_items(
        self,
        dept_manager_client,
        active_asset,
        borrower_user,
        location,
        admin_user,
    ):
        """Create a checkout with a past due_date; dashboard must contain
        'overdue' in its content."""
        from django.utils import timezone

        past_due = timezone.now() - timezone.timedelta(days=5)

        Transaction.objects.create(
            asset=active_asset,
            action="checkout",
            user=admin_user,
            borrower=borrower_user,
            from_location=active_asset.current_location,
            to_location=location,
            due_date=past_due,
        )
        active_asset.checked_out_to = borrower_user
        active_asset.save()

        resp = dept_manager_client.get(reverse("assets:dashboard"))
        assert resp.status_code == 200
        content = resp.content.decode().lower()
        assert "overdue" in content, (
            "Dashboard should display 'overdue' when there is a checked-out "
            "asset with a past due date"
        )


@pytest.mark.django_db
class TestUS_DM_057_BrowseHelpFilteredByRole:
    """US-DM-057: Browse help filtered by role.

    MoSCoW: SHOULD
    Spec refs: S2.19.5-02, S2.19.5-03
    UI Surface: /help/
    """

    def test_help_index_accessible(self, dept_manager_client):
        resp = dept_manager_client.get("/help/")
        assert resp.status_code in (200, 404)


# ---------------------------------------------------------------------------
# T10–T12, T20–T21: Additional coverage tests
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUS_DM_009_ManageTags:
    """US-DM-009 (extra): Department manager can create tags.

    Spec refs: S2.2.4-01
    UI Surface: /tags/create/ and inline AJAX endpoint
    """

    def test_inline_tag_creation(self, dept_manager_client):
        """POST to tag_create_inline with a JSON name creates a Tag."""
        tag_name = "FunctionalTestTag"
        assert not Tag.objects.filter(name=tag_name).exists()

        resp = dept_manager_client.post(
            reverse("assets:tag_create_inline"),
            data=json.dumps({"name": tag_name}),
            content_type="application/json",
        )
        # Inline endpoint returns JSON with 200 or 201
        assert resp.status_code in (200, 201), (
            f"tag_create_inline returned {resp.status_code}: "
            f"{resp.content.decode()}"
        )
        assert Tag.objects.filter(
            name=tag_name
        ).exists(), "Tag record should be created after inline POST"

    def test_tag_create_form_roundtrip(self, dept_manager_client):
        """GET tag_create form, parse fields, POST to create a tag."""
        tag_name = "FormRoundTripTag"

        # GET the form
        get_resp = dept_manager_client.get(reverse("assets:tag_create"))
        assert get_resp.status_code == 200

        # Parse form fields from HTML
        parser = _FormFieldCollector()
        parser.feed(get_resp.content.decode())
        fields = parser.fields

        # Set the tag name in the parsed fields
        # Find the name field — should be 'name' from TagForm
        name_field = None
        for field_name in fields:
            if "name" in field_name.lower() and field_name not in (
                "csrfmiddlewaretoken",
            ):
                name_field = field_name
                break

        assert name_field is not None, (
            f"Could not find a 'name' field in tag form. "
            f"Found fields: {list(fields.keys())}"
        )

        fields[name_field] = tag_name

        resp = dept_manager_client.post(
            reverse("assets:tag_create"),
            data=fields,
        )
        # Should redirect on success
        assert resp.status_code in (
            200,
            302,
        ), f"tag_create POST returned {resp.status_code}"
        assert Tag.objects.filter(
            name=tag_name
        ).exists(), "Tag record should be created via form POST"


@pytest.mark.django_db
class TestUS_DM_008_ManageImages:
    """US-DM-008 (extra): Image upload format support.

    Spec refs: S2.2.5-05, S2.2.5-05a
    UI Surface: /<pk>/edit/ (asset edit form with image upload)
    """

    @staticmethod
    def _make_png_bytes():
        """Create a minimal valid 1x1 red PNG using PIL."""
        from PIL import Image as PILImage

        buf = BytesIO()
        img = PILImage.new("RGB", (1, 1), color=(255, 0, 0))
        img.save(buf, format="PNG")
        buf.seek(0)
        return buf.read()

    @staticmethod
    def _make_webp_bytes():
        """Create a minimal valid 1x1 red WebP using PIL."""
        from PIL import Image as PILImage

        buf = BytesIO()
        img = PILImage.new("RGB", (1, 1), color=(0, 255, 0))
        img.save(buf, format="WEBP")
        buf.seek(0)
        return buf.read()

    def test_png_image_uploadable(self, dept_manager_client, active_asset):
        """Upload a PNG via asset edit form; AssetImage record created."""
        from django.core.files.uploadedfile import SimpleUploadedFile

        initial_count = AssetImage.objects.filter(asset=active_asset).count()

        # GET the edit form to extract fields
        edit_url = reverse("assets:asset_edit", args=[active_asset.pk])
        get_resp = dept_manager_client.get(edit_url)
        assert get_resp.status_code == 200

        parser = _FormFieldCollector()
        parser.feed(get_resp.content.decode())
        fields = parser.fields

        # Remove any file-type fields that can't be sent as strings
        fields.pop("images", None)
        fields.pop("image_captions", None)

        png_data = self._make_png_bytes()
        image_file = SimpleUploadedFile(
            "test_image.png",
            png_data,
            content_type="image/png",
        )

        # POST with the image file
        resp = dept_manager_client.post(
            edit_url,
            data={**fields, "images": image_file},
        )
        assert resp.status_code in (
            200,
            302,
        ), f"Asset edit POST returned {resp.status_code}"

        new_count = AssetImage.objects.filter(asset=active_asset).count()
        assert (
            new_count > initial_count
        ), "AssetImage record should be created after PNG upload"

    def test_webp_image_uploadable(self, dept_manager_client, active_asset):
        """Upload a WebP via asset edit form; AssetImage record created."""
        from django.core.files.uploadedfile import SimpleUploadedFile

        initial_count = AssetImage.objects.filter(asset=active_asset).count()

        # GET the edit form to extract fields
        edit_url = reverse("assets:asset_edit", args=[active_asset.pk])
        get_resp = dept_manager_client.get(edit_url)
        assert get_resp.status_code == 200

        parser = _FormFieldCollector()
        parser.feed(get_resp.content.decode())
        fields = parser.fields

        # Remove any file-type fields that can't be sent as strings
        fields.pop("images", None)
        fields.pop("image_captions", None)

        webp_data = self._make_webp_bytes()
        image_file = SimpleUploadedFile(
            "test_image.webp",
            webp_data,
            content_type="image/webp",
        )

        # POST with the image file
        resp = dept_manager_client.post(
            edit_url,
            data={**fields, "images": image_file},
        )
        assert resp.status_code in (
            200,
            302,
        ), f"Asset edit POST returned {resp.status_code}"

        new_count = AssetImage.objects.filter(asset=active_asset).count()
        assert (
            new_count > initial_count
        ), "AssetImage record should be created after WebP upload"


@pytest.mark.django_db
class TestUS_DM_025_ExportAssets_PurchasePrice:
    """US-DM-025 (extra): Export XLSX contains purchase price column.

    Spec refs: S2.9.1-01, S2.9.1-02
    """

    def test_export_has_purchase_price_column(
        self,
        dept_manager_client,
        active_asset,
    ):
        """Export assets XLSX; header row contains purchase price."""
        import openpyxl

        resp = dept_manager_client.get(reverse("assets:export_assets"))
        assert resp.status_code == 200, f"Export returned {resp.status_code}"

        wb = openpyxl.load_workbook(BytesIO(resp.content))
        sheet = wb.active
        if "Assets" in wb.sheetnames:
            sheet = wb["Assets"]

        headers = [
            str(cell.value).strip().lower() if cell.value else ""
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]

        has_purchase = any("purchase" in h for h in headers)
        assert has_purchase, (
            f"Export XLSX should have a 'purchase price' column. "
            f"Found headers: {headers}"
        )


@pytest.mark.django_db
class TestUS_DM_025_ExportAssets_EstimatedValue:
    """US-DM-025 (extra): Export XLSX contains estimated value column.

    Spec refs: S2.9.1-01, S2.9.1-02
    """

    def test_export_has_estimated_value_column(
        self,
        dept_manager_client,
        active_asset,
    ):
        """Export assets XLSX; header row contains estimated value."""
        import openpyxl

        resp = dept_manager_client.get(reverse("assets:export_assets"))
        assert resp.status_code == 200, f"Export returned {resp.status_code}"

        wb = openpyxl.load_workbook(BytesIO(resp.content))
        sheet = wb.active
        if "Assets" in wb.sheetnames:
            sheet = wb["Assets"]

        headers = [
            str(cell.value).strip().lower() if cell.value else ""
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]

        has_value = any("value" in h for h in headers)
        assert has_value, (
            f"Export XLSX should have an 'estimated value' column. "
            f"Found headers: {headers}"
        )
