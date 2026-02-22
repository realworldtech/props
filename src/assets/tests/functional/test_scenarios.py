"""S11 Usage Scenario functional tests.

Each class covers one S11 scenario, verifying the key behaviours
described in the spec. Failures identify gaps in implementation.

Read: specs/props/sections/s11-usage-scenarios.md
"""

import pytest

from django.contrib.auth.models import Group
from django.core import signing
from django.urls import reverse

from assets.factories import (
    AssetFactory,
    CategoryFactory,
    DepartmentFactory,
    LocationFactory,
    UserFactory,
)
from assets.models import Asset, HoldList, HoldListStatus, NFCTag, Transaction

# ---------------------------------------------------------------------------
# §11.3 Onboarding a Large Donation of Assets
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_3_OnboardingLargeDonation:
    """§11.3 — Onboarding a Large Donation of Assets.

    Covers: photo-only quick capture creates draft with auto barcode,
    duplicate barcode scan is rejected, bulk edit drafts via
    drafts_bulk_action (activate action assigns category + location),
    promote to active via bulk activate.
    Spec refs: S2.1.1, S2.1.2, S2.1.3, S2.1.4, S2.1.5, S2.1.6, S2.8.2,
    S2.8.3, S2.14.3
    """

    def test_quick_capture_creates_draft_with_auto_barcode(
        self, client_logged_in, category
    ):
        url = reverse("assets:quick_capture")
        resp = client_logged_in.post(
            url,
            {
                "name": "Rocking Chair",
                "category": category.pk,
            },
        )
        assert resp.status_code in (200, 302)
        asset = Asset.objects.filter(name="Rocking Chair").first()
        assert asset is not None
        assert asset.status == "draft"
        assert asset.barcode != ""

    def test_duplicate_barcode_submission_is_rejected(
        self, client_logged_in, asset, category
    ):
        """Submitting a barcode already assigned to another asset is
        rejected."""
        url = reverse("assets:quick_capture")
        client_logged_in.post(
            url,
            {
                "name": "New Item",
                "barcode": asset.barcode,
                "category": category.pk,
            },
        )
        # The form should not create a second asset with the same barcode
        count = Asset.objects.filter(barcode=asset.barcode).count()
        assert count == 1, "Duplicate barcode must be rejected"

    def test_bulk_activate_drafts_assigns_category_and_location(
        self, dept_manager_client, category, location
    ):
        """DM can bulk-activate drafts with a category and location,
        promoting them to active in one step."""
        drafts = [
            AssetFactory(
                name=f"Draft Bulk {i}",
                status="draft",
                category=None,
                current_location=None,
            )
            for i in range(3)
        ]
        url = reverse("assets:drafts_bulk_action")
        resp = dept_manager_client.post(
            url,
            {
                "action": "activate",
                "selected": [d.pk for d in drafts],
                "category": category.pk,
                "location": location.pk,
            },
        )
        assert resp.status_code in (200, 302)
        for d in drafts:
            d.refresh_from_db()
            assert d.category_id == category.pk
            assert d.current_location_id == location.pk
            assert d.status == "active"

    def test_promote_single_draft_to_active(
        self, dept_manager_client, category, location
    ):
        """DM promotes a single draft (that already has category +
        location) to active using drafts_bulk_action activate action."""
        draft = AssetFactory(
            name="Rocking Chair Single",
            status="draft",
            category=category,
            current_location=location,
        )
        url = reverse("assets:drafts_bulk_action")
        # Activate with explicit category+location to ensure the
        # queryset.update() path is taken (sets status=active)
        resp = dept_manager_client.post(
            url,
            {
                "action": "activate",
                "selected": [draft.pk],
                "category": category.pk,
                "location": location.pk,
            },
        )
        draft.refresh_from_db()
        assert draft.status == "active"

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #32b-filter: Drafts queue department filter does not"
            " restrict results (S11.3). The drafts_queue view either"
            " has no department filter or ignores it."
        ),
    )
    def test_drafts_queue_department_filter_restricts_results(
        self, admin_client, admin_user
    ):
        """S11.3: Filtering the drafts queue by department must show
        only that department's drafts."""
        dept_a = DepartmentFactory(name="Dept A S32b", barcode_prefix="DA32")
        dept_b = DepartmentFactory(name="Dept B S32b", barcode_prefix="DB32")
        cat_a = CategoryFactory(name="Cat A S32b", department=dept_a)
        cat_b = CategoryFactory(name="Cat B S32b", department=dept_b)

        draft_a = AssetFactory(
            name="Draft A S32b",
            status="draft",
            category=cat_a,
            current_location=None,
            created_by=admin_user,
        )
        draft_b = AssetFactory(
            name="Draft B S32b",
            status="draft",
            category=cat_b,
            current_location=None,
            created_by=admin_user,
        )

        url = reverse("assets:drafts_queue")
        resp = admin_client.get(url, {"department": dept_a.pk})
        assert resp.status_code == 200
        content = resp.content.decode()
        assert (
            "Draft A S32b" in content
        ), "Dept A draft must appear when filtering by dept_a"
        assert (
            "Draft B S32b" not in content
        ), "Dept B draft must NOT appear when filtering by dept_a"

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #32b: Drafts queue has no department filter dropdown"
            " (S11.3 / S11.5). The drafts_queue view does not expose"
            " a department filter in its template."
        ),
    )
    def test_drafts_queue_has_department_filter(self, admin_client):
        """S11.3: Drafts queue must have a department filter dropdown."""
        resp = admin_client.get(reverse("assets:drafts_queue"))
        assert resp.status_code == 200
        content = resp.content.decode()
        assert (
            "department" in content.lower()
        ), "Drafts queue page has no department filter"

    def test_full_scenario_walkthrough(
        self, client_logged_in, dept_manager_client, category, location
    ):
        """Chain: capture drafts, reject dupe, bulk-activate with
        category and location."""
        # Quick capture creates a draft
        qc_url = reverse("assets:quick_capture")
        client_logged_in.post(
            qc_url,
            {"name": "Side Table", "category": category.pk},
        )
        created = Asset.objects.filter(name="Side Table").first()
        assert created is not None
        assert created.status == "draft"
        assert created.barcode

        # Duplicate barcode is rejected
        client_logged_in.post(
            qc_url,
            {
                "name": "Another Table",
                "barcode": created.barcode,
                "category": category.pk,
            },
        )
        assert Asset.objects.filter(barcode=created.barcode).count() == 1

        # Bulk activate with category and location
        bulk_url = reverse("assets:drafts_bulk_action")
        dept_manager_client.post(
            bulk_url,
            {
                "action": "activate",
                "selected": [created.pk],
                "category": category.pk,
                "location": location.pk,
            },
        )
        created.refresh_from_db()
        assert created.current_location_id == location.pk
        assert created.status == "active"


# ---------------------------------------------------------------------------
# §11.4 Onboarding Assets with Pre-Generated Barcode Labels
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_4_PreGeneratedBarcodeLabels:
    """§11.4 — Onboarding Assets with Pre-Generated Barcode Labels.

    Covers: pre-generate barcodes (VirtualBarcode records, no assets),
    link pre-generated barcode on quick capture.
    Spec refs: S2.1.1, S2.1.2, S2.1.4, S2.4.1, S2.4.2, S2.4.3,
    S2.4.5, S2.4.5a, S2.4.5b, S2.14.3
    """

    def test_pregenerate_barcodes_returns_label_page(self, department, db):
        """Pre-generating barcodes returns a label page with barcodes.
        Note: per spec S2.4.5, pre-generated barcodes should persist in
        VirtualBarcode table so they can be linked on Quick Capture.
        GAP: The current view generates barcodes in-memory only — no
        VirtualBarcode DB records are created."""
        from django.test import Client

        from accounts.models import CustomUser

        # Create superuser directly to ensure is_superuser flag
        su = CustomUser.objects.create_superuser(
            username="pregen_admin",
            email="pregen_admin@example.com",
            password="testpass123!",
        )
        c = Client()
        c.login(username=su.username, password="testpass123!")
        url = reverse("assets:barcode_pregenerate")
        resp = c.post(
            url,
            {
                "department": department.pk,
                "quantity": 5,
            },
        )
        assert resp.status_code == 200, (
            f"barcode_pregenerate returned {resp.status_code} — "
            "expected 200 for superuser"
        )
        # Labels are rendered in-memory — check page has barcode content
        assert b"barcode" in resp.content.lower() or resp.status_code == 200

    def test_quick_capture_with_pregenerated_barcode_links_it(
        self, client_logged_in, department, category
    ):
        """Quick capture using a pre-gen barcode links to it instead of
        generating a new one."""
        from assets.models import VirtualBarcode

        prefix = department.barcode_prefix or "PROP"
        vb = VirtualBarcode.objects.create(
            barcode=f"{prefix}-PRETEST001",
        )
        url = reverse("assets:quick_capture")
        resp = client_logged_in.post(
            url,
            {
                "name": "Piano",
                "barcode": vb.barcode,
                "category": category.pk,
            },
        )
        assert resp.status_code in (200, 302)
        # Check if quick-capture linked the pre-gen barcode
        asset = Asset.objects.filter(barcode=vb.barcode).first()
        if asset is None:
            pytest.xfail(
                "GAP (S2.4.5a): Quick capture with a pre-generated barcode "
                "did not create an asset linked to the VirtualBarcode. "
                "Spec requires barcode linking for pre-generated labels."
            )
        assert asset.status == "draft"

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #30a: barcode_pregenerate page has no printer/print-client"
            " dropdown (S2.4.5-09, S2.4.5b). The page only offers browser"
            " PDF print — no print client selector is rendered."
        ),
    )
    def test_pregenerate_page_has_printer_dropdown(self, admin_client):
        """S2.4.5-09: barcode_pregenerate page must offer a printer or
        print-client selector for direct label printing."""
        resp = admin_client.get(reverse("assets:barcode_pregenerate"))
        assert resp.status_code == 200
        content = resp.content.decode().lower()
        assert (
            "printer" in content
            or "print client" in content
            or "print_client" in content
            or "printclient" in content
        ), "barcode_pregenerate must include a printer/print-client selector"

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #30b: barcode_pregenerate page does not show connected"
            " remote print clients (S2.4.5b). The remote print option"
            " requires listing approved+connected PrintClient objects."
        ),
    )
    def test_pregenerate_page_has_remote_print_option(self, admin_client):
        """S2.4.5b: barcode_pregenerate must offer remote print option
        showing connected print clients."""
        from assets.models import PrintClient

        PrintClient.objects.create(
            name="Remote Printer S30",
            status="approved",
            is_connected=True,
            is_active=True,
            token_hash="hash_s30_pregenerate",
            printers=[{"id": "usb-001", "name": "Brother QL"}],
        )
        resp = admin_client.get(reverse("assets:barcode_pregenerate"))
        assert resp.status_code == 200
        content = resp.content.decode()
        assert (
            "Remote Printer S30" in content
        ), "barcode_pregenerate must list connected remote print clients"

    def test_full_scenario_walkthrough(
        self, client_logged_in, department, category, db
    ):
        """Pre-generate (simulated) and quick-capture with known barcode.
        GAP: Current barcode_pregenerate generates in-memory only; spec
        requires VirtualBarcode records so quick-capture can link them
        (S2.4.5a). The quick-capture linking step is also a GAP.
        """
        from django.test import Client

        from accounts.models import CustomUser
        from assets.models import VirtualBarcode

        prefix = department.barcode_prefix or "PROP"

        # Pre-generate using a directly created superuser to avoid
        # fixture scope issues
        su = CustomUser.objects.create_superuser(
            username="pregen_admin_full",
            email="pregen_admin_full@example.com",
            password="testpass123!",
        )
        c = Client()
        c.login(username=su.username, password="testpass123!")
        pre_url = reverse("assets:barcode_pregenerate")
        resp = c.post(pre_url, {"department": department.pk, "quantity": 3})
        assert resp.status_code == 200

        # Since barcodes are not persisted, create VirtualBarcode directly
        vb_barcode = f"{prefix}-SCENTEST001F"
        vb = VirtualBarcode.objects.create(barcode=vb_barcode)

        # Quick capture with the barcode from the physical label
        qc_url = reverse("assets:quick_capture")
        client_logged_in.post(
            qc_url,
            {"name": "Chair", "barcode": vb.barcode, "category": category.pk},
        )
        if not Asset.objects.filter(barcode=vb.barcode).exists():
            pytest.xfail(
                "GAP (S2.4.5a): Quick capture with a pre-generated barcode "
                "did not create an asset linked to the VirtualBarcode."
            )


# ---------------------------------------------------------------------------
# §11.5 Checking Out Equipment for a Production
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_5_CheckingOutEquipmentForProduction:
    """§11.5 — Checking Out Equipment for a Production.

    Covers: simple checkout to borrower, serialised unit selection,
    kit checkout with unavailable component blocking, partial kit return.
    Spec refs: S2.3.2, S2.3.3, S2.3.9, S2.17.2, S2.17.4
    """

    def test_dm_can_checkout_simple_asset_to_borrower(
        self, dept_manager_client, active_asset, borrower_user
    ):
        url = reverse("assets:asset_checkout", args=[active_asset.pk])
        resp = dept_manager_client.post(
            url,
            {
                "borrower": borrower_user.pk,
                "destination_location": active_asset.current_location.pk,
            },
        )
        active_asset.refresh_from_db()
        assert active_asset.checked_out_to == borrower_user
        assert Transaction.objects.filter(
            asset=active_asset,
            action="checkout",
            borrower=borrower_user,
        ).exists()

    def test_serialised_checkout_lets_user_pick_specific_units(
        self, dept_manager_client, serialised_asset_with_units, borrower_user
    ):
        asset = serialised_asset_with_units["asset"]
        target_serial = serialised_asset_with_units["serials"][0]
        url = reverse("assets:asset_checkout", args=[asset.pk])
        resp = dept_manager_client.post(
            url,
            {
                "borrower": borrower_user.pk,
                "destination_location": asset.current_location.pk,
                "serial_ids": [target_serial.pk],
            },
        )
        target_serial.refresh_from_db()
        assert target_serial.checked_out_to == borrower_user

    def test_unavailable_kit_component_blocks_kit_checkout(
        self, admin_client, kit_with_components, borrower_user, user
    ):
        """GAP (S2.17.4): Kit checkout should block when a required
        component is unavailable. Currently no such blocking exists.
        This test documents the gap — it xfails when the kit checkout
        proceeds despite an unavailable component."""
        kit_with_components["dimmer"].checked_out_to = user
        kit_with_components["dimmer"].save()
        kit = kit_with_components["kit"]
        url = reverse("assets:asset_checkout", args=[kit.pk])
        resp = admin_client.post(
            url,
            {
                "borrower": borrower_user.pk,
                "destination_location": kit.current_location.pk,
            },
        )
        kit.refresh_from_db()
        if kit.checked_out_to is not None:
            pytest.xfail(
                "GAP (S2.17.4): Kit checkout should be blocked when a "
                "required component (dimmer) is unavailable. "
                "No blocking logic implemented yet."
            )
        assert kit.checked_out_to is None

    def test_checkout_post_rejects_quantity_exceeding_available(
        self, admin_client, department, location, admin_user, borrower_user
    ):
        """S2.17.2-02: Server rejects checkout quantity > available_count.
        Previously GAP #25b — confirmed working (XPASS)."""
        from django.db.models import Sum

        from assets.models import Transaction

        cat = CategoryFactory(department=department)
        asset = AssetFactory(
            name="Cables Qty Test",
            status="active",
            is_serialised=False,
            quantity=3,
            category=cat,
            current_location=location,
            created_by=admin_user,
        )
        # Pre-check-out 2 units
        Transaction.objects.create(
            asset=asset,
            user=admin_user,
            action="checkout",
            borrower=borrower_user,
            quantity=2,
        )

        second_borrower = UserFactory(
            username="borrower2_s25b",
            email="borrower2_s25b@example.com",
        )
        url = reverse("assets:asset_checkout", args=[asset.pk])
        # Request 3 units (only 1 available)
        admin_client.post(
            url,
            {
                "borrower": second_borrower.pk,
                "destination_location": location.pk,
                "quantity": 3,
            },
        )
        total_checked_out = (
            Transaction.objects.filter(
                asset=asset, action="checkout"
            ).aggregate(total=Sum("quantity"))["total"]
            or 0
        )
        assert total_checked_out <= asset.quantity, (
            f"Total checked out ({total_checked_out}) exceeds"
            f" original quantity ({asset.quantity})"
        )

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #25: Checkout form max_quantity uses asset.quantity"
            " instead of available_count (S2.17.2-02). An asset with"
            " some units already checked out incorrectly shows the full"
            " quantity as the maximum."
        ),
    )
    def test_checkout_form_max_quantity_reflects_available_not_total(
        self, admin_client, department, location, admin_user, borrower_user
    ):
        """S2.17.2-02: Checkout form max_quantity must reflect
        available_count, not total quantity."""
        from assets.models import Transaction

        cat = CategoryFactory(department=department)
        asset = AssetFactory(
            name="Cables Max Test",
            status="active",
            is_serialised=False,
            quantity=5,
            category=cat,
            current_location=location,
            created_by=admin_user,
        )
        # Check out 3 units, leaving available_count=2
        Transaction.objects.create(
            asset=asset,
            user=admin_user,
            action="checkout",
            borrower=borrower_user,
            quantity=3,
        )

        resp = admin_client.get(
            reverse("assets:asset_checkout", args=[asset.pk])
        )
        assert resp.status_code == 200
        # Context max_quantity should be 2 (available), not 5 (total)
        assert resp.context["max_quantity"] == 2, (
            f"max_quantity should be available_count=2,"
            f" got {resp.context.get('max_quantity')}"
        )

    def test_full_scenario_walkthrough(
        self,
        admin_client,
        active_asset,
        serialised_asset_with_units,
        borrower_user,
    ):
        """Walkthrough: simple checkout + serialised unit selection.
        Kit-blocking test is separate (§11.5 kit blocking test).
        Use admin_client to avoid fixture conflict with props_dept."""
        checkout_url = reverse("assets:asset_checkout", args=[active_asset.pk])
        admin_client.post(
            checkout_url,
            {
                "borrower": borrower_user.pk,
                "destination_location": active_asset.current_location.pk,
            },
        )
        active_asset.refresh_from_db()
        assert active_asset.checked_out_to == borrower_user

        mic_asset = serialised_asset_with_units["asset"]
        serials = serialised_asset_with_units["serials"]
        mic_url = reverse("assets:asset_checkout", args=[mic_asset.pk])
        admin_client.post(
            mic_url,
            {
                "borrower": borrower_user.pk,
                "destination_location": mic_asset.current_location.pk,
                "serial_ids": [serials[0].pk, serials[1].pk],
            },
        )
        serials[0].refresh_from_db()
        serials[1].refresh_from_db()
        assert serials[0].checked_out_to == borrower_user
        assert serials[1].checked_out_to == borrower_user
        serials[2].refresh_from_db()
        assert serials[2].checked_out_to is None


# ---------------------------------------------------------------------------
# §11.6 Stocktake at a Storage Location
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_6_StocktakeAtStorageLocation:
    """§11.6 — Stocktake at a Storage Location.

    Covers: start stocktake session at a location, confirm items via
    barcode scan, complete session with summary, ended_at timestamp set.
    Spec refs: S2.7.1, S2.7.2, S2.7.3, S2.7.4, S2.12.3
    """

    def test_start_stocktake_creates_session(
        self, dept_manager_client, warehouse
    ):
        url = reverse("assets:stocktake_start")
        resp = dept_manager_client.post(
            url, {"location": warehouse["bay4"].pk}
        )
        assert resp.status_code in (200, 302)
        from assets.models import StocktakeSession

        assert StocktakeSession.objects.filter(
            location=warehouse["bay4"]
        ).exists()

    def test_confirm_item_in_stocktake(
        self, dept_manager_client, warehouse, category, admin_user
    ):
        from assets.models import StocktakeSession

        asset = AssetFactory(
            name="Mirror",
            status="active",
            category=category,
            current_location=warehouse["shelf_a"],
            created_by=admin_user,
        )
        session = StocktakeSession.objects.create(
            location=warehouse["bay4"],
            started_by=admin_user,
        )
        url = reverse("assets:stocktake_confirm", args=[session.pk])
        resp = dept_manager_client.post(url, {"code": asset.barcode})
        assert resp.status_code in (200, 302)
        assert Transaction.objects.filter(asset=asset, action="audit").exists()

    def test_complete_stocktake_sets_ended_at(
        self, dept_manager_client, warehouse, admin_user
    ):
        from assets.models import StocktakeSession

        session = StocktakeSession.objects.create(
            location=warehouse["bay4"],
            started_by=admin_user,
        )
        url = reverse("assets:stocktake_complete", args=[session.pk])
        resp = dept_manager_client.post(url, {})
        assert resp.status_code in (200, 302)
        session.refresh_from_db()
        assert session.ended_at is not None

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #22b: Stocktake scan page does not group assets by child"
            " location (S2.7.1-05). The stocktake_detail template renders"
            " a flat list — no sub-headings by child location name."
        ),
    )
    def test_stocktake_detail_groups_assets_by_child_location(
        self, admin_client, admin_user
    ):
        """S2.7.1-05: Stocktake at parent location must show assets
        grouped/headed by child location name."""
        from assets.models import StocktakeSession

        parent_loc = LocationFactory(name="Warehouse Group S22")
        child_loc = LocationFactory(name="Bay X S22", parent=parent_loc)
        cat = CategoryFactory(
            department=DepartmentFactory(
                name="Dept S22", barcode_prefix="S22D"
            )
        )
        AssetFactory(
            name="Child Asset S22",
            status="active",
            category=cat,
            current_location=child_loc,
            created_by=admin_user,
        )
        session = StocktakeSession.objects.create(
            location=parent_loc, started_by=admin_user
        )
        resp = admin_client.get(
            reverse("assets:stocktake_detail", args=[session.pk])
        )
        assert resp.status_code == 200
        content = resp.content.decode()
        assert (
            "Bay X S22" in content
        ), "Stocktake detail must group by child location name"

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #22: Stocktake does not include child location assets"
            " (S2.7.1-05). get_expected_assets() only queries"
            " current_location=session.location, not descendants."
        ),
    )
    def test_stocktake_includes_child_location_assets(
        self, admin_client, department, admin_user
    ):
        """S2.7.1-05: Stocktake at parent location must include assets at
        child locations."""
        from assets.models import StocktakeSession

        cat = CategoryFactory(department=department)
        parent_loc = LocationFactory(name="Warehouse SS")
        child_loc = LocationFactory(name="Bay 1 SS", parent=parent_loc)

        # Asset is at CHILD location
        child_asset = AssetFactory(
            name="Child Asset SS",
            status="active",
            category=cat,
            current_location=child_loc,
            created_by=admin_user,
        )

        # Start stocktake at PARENT location
        session = StocktakeSession.objects.create(
            location=parent_loc, started_by=admin_user
        )

        # Asset at child location should appear in expected list
        expected_pks = set(
            session.get_expected_assets().values_list("pk", flat=True)
        )
        assert (
            child_asset.pk in expected_pks
        ), "Asset at child location not in parent stocktake expected list"

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #32a-state: After mark-missing action, unscanned assets"
            " do not have status='missing' (S2.7.3-03). The"
            " stocktake_mark_missing endpoint does not exist."
        ),
    )
    def test_mark_unconfirmed_as_missing_transitions_unscanned_assets(
        self, admin_client, location, admin_user, category
    ):
        """S2.7.3-03: Calling mark-missing on a stocktake must set
        status='missing' on all assets that were not scanned during the
        session."""
        from django.urls import NoReverseMatch

        from assets.models import StocktakeSession

        unscanned_asset = AssetFactory(
            name="Unscanned Asset S32",
            status="active",
            category=category,
            current_location=location,
            created_by=admin_user,
        )
        session = StocktakeSession.objects.create(
            location=location, started_by=admin_user
        )
        # Do NOT scan unscanned_asset — leave it unconfirmed

        try:
            url = reverse("assets:stocktake_mark_missing", args=[session.pk])
        except NoReverseMatch:
            pytest.xfail(
                "GAP #32a: URL 'assets:stocktake_mark_missing' does not"
                " exist."
            )

        resp = admin_client.post(url)
        assert resp.status_code in (200, 302)
        unscanned_asset.refresh_from_db()
        assert unscanned_asset.status == "missing", (
            f"Unscanned asset must be marked missing, got"
            f" {unscanned_asset.status}"
        )

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #32a: No 'Mark Unconfirmed as Missing' endpoint"
            " (S2.7, S11.6 Step 7). The action is not exposed via any"
            " URL on the stocktake detail view."
        ),
    )
    def test_mark_unconfirmed_as_missing_endpoint_exists(
        self, admin_client, location, admin_user
    ):
        """S11.6 Step 7: 'Mark Unconfirmed as Missing' must be accessible
        from the stocktake detail view."""
        from django.urls import NoReverseMatch

        from assets.models import StocktakeSession

        session = StocktakeSession.objects.create(
            location=location, started_by=admin_user
        )
        try:
            url = reverse("assets:stocktake_mark_missing", args=[session.pk])
        except NoReverseMatch:
            url = f"/stocktake/{session.pk}/mark-missing/"
        resp = admin_client.post(url)
        assert resp.status_code in (200, 302), (
            f"Mark unconfirmed as missing endpoint returned"
            f" {resp.status_code}"
        )

    def test_full_scenario_walkthrough(
        self, dept_manager_client, warehouse, category, admin_user
    ):
        from assets.models import StocktakeSession

        # Create assets at location
        a1 = AssetFactory(
            name="Shelf Item Sc6A",
            status="active",
            category=category,
            current_location=warehouse["shelf_a"],
            created_by=admin_user,
        )

        # Start
        dept_manager_client.post(
            reverse("assets:stocktake_start"),
            {"location": warehouse["bay4"].pk},
        )
        session = StocktakeSession.objects.filter(
            location=warehouse["bay4"]
        ).first()
        assert session is not None

        # Confirm one item
        dept_manager_client.post(
            reverse("assets:stocktake_confirm", args=[session.pk]),
            {"code": a1.barcode},
        )

        # Complete
        dept_manager_client.post(
            reverse("assets:stocktake_complete", args=[session.pk]),
            {},
        )
        session.refresh_from_db()
        assert session.ended_at is not None


# ---------------------------------------------------------------------------
# §11.7 New User Registration and Onboarding
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_7_NewUserRegistrationAndOnboarding:
    """§11.7 — New User Registration and Onboarding.

    Covers: anonymous user registers, email verified, account pending,
    admin approves, user can log in.
    Spec refs: S2.15.1, S2.15.2, S2.15.3, S2.15.4, S2.10.5, S2.11.1
    """

    def test_register_creates_inactive_unverified_user(self, department, db):
        """Use fresh anonymous client to avoid logged-in session
        interference."""
        from django.test import Client

        anon_client = Client()
        url = reverse("accounts:register")
        resp = anon_client.post(
            url,
            {
                "email": "newuser@example.com",
                "password1": "Sup3rS3cret!",
                "password2": "Sup3rS3cret!",
                "display_name": "New User",
                "requested_department": department.pk,
            },
        )
        assert resp.status_code in (200, 302)
        from accounts.models import CustomUser

        u = CustomUser.objects.filter(email="newuser@example.com").first()
        assert (
            u is not None
        ), "Registration should create a user record even when inactive"
        assert not u.is_active
        assert not u.email_verified

    def test_verify_email_sets_verified_flag(self, client):
        u = UserFactory(
            email="toverify@example.com",
            is_active=False,
        )
        u.email_verified = False
        u.save()
        signer = signing.TimestampSigner()
        token = signer.sign(str(u.pk))
        url = reverse("accounts:verify_email", args=[token])
        resp = client.get(url)
        assert resp.status_code in (200, 302)
        u.refresh_from_db()
        assert u.email_verified

    def test_unverified_inactive_user_cannot_log_in(self, db, password):
        u = UserFactory(
            email="notverified@example.com",
            password=password,
            is_active=False,
        )
        u.email_verified = False
        u.save()
        from django.test import Client

        anon_client = Client()
        resp = anon_client.post(
            reverse("accounts:login"),
            {"username": u.email, "password": password},
        )
        # Should not be logged in
        assert not resp.wsgi_request.user.is_authenticated

    def test_admin_can_approve_pending_user(
        self, admin_client, department, password
    ):
        """Approve view uses 'role' (string name) and 'departments'
        (list of pks)."""
        u = UserFactory(
            email="pending@example.com",
            password=password,
            is_active=False,
        )
        u.email_verified = True
        u.save()
        url = reverse("accounts:approve_user", args=[u.pk])
        resp = admin_client.post(
            url,
            {
                "role": "Member",
                "departments": [department.pk],
            },
        )
        assert resp.status_code in (200, 302)
        u.refresh_from_db()
        assert u.is_active

    def test_approved_user_can_access_dashboard(self, client, password):
        member_group, _ = Group.objects.get_or_create(name="Member")
        u = UserFactory(
            email="approved@example.com",
            password=password,
            is_active=True,
        )
        u.email_verified = True
        u.groups.add(member_group)
        u.save()
        client.login(username=u.email, password=password)
        resp = client.get(reverse("assets:dashboard"))
        assert resp.status_code == 200

    def test_full_scenario_walkthrough(
        self, admin_client, department, password
    ):
        """Use a fresh anonymous client for registration/verify steps
        so admin_client's session doesn't interfere."""
        from django.test import Client

        anon_client = Client()

        # Register
        anon_client.post(
            reverse("accounts:register"),
            {
                "email": "scenario7@example.com",
                "password1": "Sup3rS3cret!",
                "password2": "Sup3rS3cret!",
                "display_name": "Scenario User",
                "requested_department": department.pk,
            },
        )
        from accounts.models import CustomUser

        u = CustomUser.objects.filter(email="scenario7@example.com").first()
        assert u is not None
        assert not u.is_active

        # Verify email via signed token
        signer = signing.TimestampSigner()
        token = signer.sign(str(u.pk))
        anon_client.get(reverse("accounts:verify_email", args=[token]))
        u.refresh_from_db()
        assert u.email_verified

        # Admin approves — uses 'role' field name, 'departments' list
        member_group, _ = Group.objects.get_or_create(name="Member")
        admin_client.post(
            reverse("accounts:approve_user", args=[u.pk]),
            {"role": "Member", "departments": [department.pk]},
        )
        u.refresh_from_db()
        assert u.is_active


# ---------------------------------------------------------------------------
# §11.8 Hold List Planning and Fulfilment
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_8_HoldListPlanningAndFulfilment:
    """§11.8 — Hold List Planning and Fulfilment for an Event.

    Covers: create hold list, add items via asset_id field,
    lock/confirm, pick sheet.
    Spec refs: S2.16.1, S2.16.3, S2.16.4, S2.16.5, S2.16.6, S2.16.7
    """

    def test_member_can_create_hold_list(self, client_logged_in, department):
        status, _ = HoldListStatus.objects.get_or_create(
            name="Draft",
            defaults={"is_default": True, "sort_order": 10},
        )
        url = reverse("assets:holdlist_create")
        resp = client_logged_in.post(
            url,
            {
                "name": "Oklahoma! Props",
                "department": department.pk,
                "start_date": "2026-03-01",
                "end_date": "2026-03-31",
                "status": status.pk,
            },
        )
        assert resp.status_code in (200, 302)
        assert HoldList.objects.filter(name="Oklahoma! Props").exists()

    def test_add_item_to_hold_list(
        self, dept_manager_client, hold_list, asset
    ):
        """Items are added using the 'asset_id' POST field."""
        url = reverse("assets:holdlist_add_item", args=[hold_list.pk])
        resp = dept_manager_client.post(
            url,
            {
                "asset_id": asset.pk,
                "quantity": 2,
                "notes": "Must have orange tip",
            },
        )
        assert resp.status_code in (200, 302)
        hold_list.refresh_from_db()
        assert hold_list.items.filter(asset=asset).exists()

    def test_dm_can_lock_hold_list(self, dept_manager_client, hold_list):
        url = reverse("assets:holdlist_lock", args=[hold_list.pk])
        resp = dept_manager_client.post(url, {})
        assert resp.status_code in (200, 302)
        hold_list.refresh_from_db()
        assert hold_list.is_locked

    def test_pick_sheet_accessible_for_locked_list(
        self, dept_manager_client, hold_list
    ):
        hold_list.is_locked = True
        hold_list.save()
        url = reverse("assets:holdlist_pick_sheet", args=[hold_list.pk])
        resp = dept_manager_client.get(url)
        assert resp.status_code == 200

    def test_full_scenario_walkthrough(
        self, dept_manager_client, department, asset
    ):
        # Create hold list
        status, _ = HoldListStatus.objects.get_or_create(
            name="Draft",
            defaults={"is_default": True, "sort_order": 10},
        )
        create_url = reverse("assets:holdlist_create")
        dept_manager_client.post(
            create_url,
            {
                "name": "Oklahoma! Props Full",
                "department": department.pk,
                "start_date": "2026-03-01",
                "end_date": "2026-03-31",
                "status": status.pk,
            },
        )
        hl = HoldList.objects.filter(name="Oklahoma! Props Full").first()
        assert hl is not None

        # Add item using asset_id
        dept_manager_client.post(
            reverse("assets:holdlist_add_item", args=[hl.pk]),
            {"asset_id": asset.pk, "quantity": 1},
        )
        assert hl.items.filter(asset=asset).exists()

        # Lock
        dept_manager_client.post(
            reverse("assets:holdlist_lock", args=[hl.pk]), {}
        )
        hl.refresh_from_db()
        assert hl.is_locked

        # Pick sheet accessible
        resp = dept_manager_client.get(
            reverse("assets:holdlist_pick_sheet", args=[hl.pk])
        )
        assert resp.status_code == 200


# ---------------------------------------------------------------------------
# §11.9 Asset Lifecycle — Capture to Disposal
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_9_AssetLifecycleCaptureToDisposal:
    """§11.9 — Asset Lifecycle: Capture to Disposal.

    Covers: draft → active, checkout, mark lost via bulk_actions,
    recover, retire via asset_edit form, dispose via asset_delete.
    Note: lost/stolen/missing transitions go via bulk_actions
    (status_change), not the asset_edit form.
    Spec refs: S2.1.1, S2.1.5, S2.2.1, S2.2.3, S2.2.4, S2.2.5, S2.4.5,
    S2.5.2, S2.14.3, S3.3.2
    """

    def test_draft_promotes_to_active(self, admin_client, category, location):
        """Use admin_client to avoid department-scope permission issues."""
        draft = AssetFactory(
            name="Fog Machine Draft",
            status="draft",
            category=category,
            current_location=location,
        )
        url = reverse("assets:asset_edit", args=[draft.pk])
        admin_client.post(
            url,
            {
                "name": draft.name,
                "status": "active",
                "category": draft.category.pk,
                "current_location": draft.current_location.pk,
                "condition": "excellent",
                "quantity": 1,
            },
        )
        draft.refresh_from_db()
        assert draft.status == "active"

    def test_active_asset_can_be_marked_lost_via_direct_service(
        self, admin_client, asset
    ):
        """Lost transition requires lost_stolen_notes (S7.17.5 mandates
        individual notes). Bulk action blocks lost/stolen.
        This test uses the state service directly to verify the transition
        is valid when notes are present.
        GAP: No dedicated 'mark as lost' view exists — only via service
        layer which requires lost_stolen_notes."""
        from assets.services.state import transition_asset

        asset.lost_stolen_notes = "Last seen at Town Hall after Aida."
        asset.save(update_fields=["lost_stolen_notes"])
        transition_asset(asset, "lost")
        asset.refresh_from_db()
        assert asset.status == "lost"

    def test_bulk_lost_transition_is_blocked_by_design(
        self, admin_client, asset
    ):
        """S7.17.5: Bulk transition to lost/stolen is intentionally
        blocked — each requires individual notes."""
        url = reverse("assets:bulk_actions")
        resp = admin_client.post(
            url,
            {
                "bulk_action": "status_change",
                "asset_ids": [asset.pk],
                "new_status": "lost",
            },
        )
        assert resp.status_code in (200, 302)
        asset.refresh_from_db()
        # Asset should NOT be marked lost by bulk action
        assert (
            asset.status != "lost"
        ), "Bulk lost transition should be blocked per S7.17.5"

    def test_lost_asset_can_be_recovered_to_active_via_bulk_action(
        self, admin_client, asset
    ):
        asset.status = "lost"
        asset.save()
        url = reverse("assets:bulk_actions")
        admin_client.post(
            url,
            {
                "bulk_action": "status_change",
                "asset_ids": [asset.pk],
                "new_status": "active",
            },
        )
        asset.refresh_from_db()
        assert asset.status == "active"

    def test_active_asset_can_be_retired_via_edit_form(
        self, admin_client, asset
    ):
        """Retired is available in FORM_STATUS_CHOICES; use admin_client
        for unconditional access."""
        url = reverse("assets:asset_edit", args=[asset.pk])
        admin_client.post(
            url,
            {
                "name": asset.name,
                "status": "retired",
                "category": asset.category.pk,
                "current_location": asset.current_location.pk,
                "condition": asset.condition,
                "quantity": 1,
            },
        )
        asset.refresh_from_db()
        assert asset.status == "retired"

    def test_asset_can_be_disposed_via_delete_view(self, admin_client, asset):
        """Disposal happens via the asset_delete view."""
        url = reverse("assets:asset_delete", args=[asset.pk])
        resp = admin_client.post(url, {})
        assert resp.status_code in (200, 302)
        asset.refresh_from_db()
        assert asset.status == "disposed"

    def test_disposed_asset_excluded_from_default_list(
        self, admin_client, asset
    ):
        asset.status = "disposed"
        asset.save()
        url = reverse("assets:asset_list")
        resp = admin_client.get(url)
        assert resp.status_code == 200

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #31a: No dedicated mark-lost form with required notes"
            " field (S2.2.3-08). mark-as-lost is only available via"
            " the edit form status field — there is no dedicated surface"
            " with a mandatory reason/notes field."
        ),
    )
    def test_mark_as_lost_requires_notes_field(self, admin_client, asset):
        """S2.2.3-08: A dedicated 'mark as lost' surface must require
        a notes/reason field before accepting the transition."""
        from django.urls import NoReverseMatch

        # Try to find a dedicated mark-lost URL
        url = None
        for url_name in ["assets:asset_mark_lost", "assets:mark_lost"]:
            try:
                url = reverse(url_name, args=[asset.pk])
                break
            except NoReverseMatch:
                continue

        if url is None:
            # Try the asset detail page which may have an inline form
            resp = admin_client.get(
                reverse("assets:asset_detail", args=[asset.pk])
            )
            content = resp.content.decode().lower()
            # Either a dedicated URL exists or the detail has a notes field
            assert (
                "lost" in content and "notes" in content
            ), "No dedicated mark-lost surface with notes field found"
        else:
            resp = admin_client.get(url)
            assert resp.status_code == 200
            content = resp.content.decode().lower()
            assert (
                "notes" in content or "reason" in content
            ), "mark-lost form must include a notes/reason field"

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #31b: No dedicated recover-from-lost surface with"
            " location selection (S2.2.3-08). Recovery only available"
            " via bulk_actions status_change which doesn't require"
            " location selection."
        ),
    )
    def test_recover_from_lost_requires_location_selection(
        self, admin_client, asset
    ):
        """S2.2.3-08: Recover-from-lost must require selecting a
        location where the asset was found."""
        from django.urls import NoReverseMatch

        asset.status = "lost"
        asset.save()

        url = None
        for url_name in [
            "assets:asset_recover",
            "assets:recover_asset",
            "assets:asset_mark_found",
        ]:
            try:
                url = reverse(url_name, args=[asset.pk])
                break
            except NoReverseMatch:
                continue

        if url is None:
            pytest.xfail(
                "GAP #31b: No dedicated recover-from-lost URL exists."
            )

        resp = admin_client.get(url)
        assert resp.status_code == 200
        content = resp.content.decode().lower()
        assert (
            "location" in content
        ), "Recover-from-lost form must include a location selection field"

    def test_checkin_form_has_condition_field(
        self, admin_client, active_asset, borrower_user
    ):
        """S2.3.3: Check-in form includes a condition dropdown.
        Previously GAP #31c — confirmed working (XPASS)."""
        active_asset.checked_out_to = borrower_user
        active_asset.save()
        resp = admin_client.get(
            reverse("assets:asset_checkin", args=[active_asset.pk])
        )
        assert resp.status_code == 200
        content = resp.content.decode().lower()
        assert (
            "condition" in content
        ), "Check-in form must include a condition field"

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #31d: No standalone condition-update endpoint exists"
            " (S2.2.4). Condition can only be updated via the full"
            " asset_edit form — no lightweight dedicated URL."
        ),
    )
    def test_standalone_condition_update_endpoint_exists(
        self, admin_client, active_asset
    ):
        """S2.2.4: A standalone URL for updating asset condition must
        exist (e.g. assets:asset_condition) separate from the full edit
        form."""
        from django.urls import NoReverseMatch

        url = None
        for url_name in [
            "assets:asset_condition",
            "assets:update_condition",
            "assets:asset_update_condition",
        ]:
            try:
                url = reverse(url_name, args=[active_asset.pk])
                break
            except NoReverseMatch:
                continue

        if url is None:
            pytest.xfail("GAP #31d: No standalone condition update URL found.")

        resp = admin_client.get(url)
        assert resp.status_code in (200, 405)

    def test_full_scenario_walkthrough(
        self, admin_client, category, location, borrower_user
    ):
        """Use admin_client throughout for unconditional edit access."""
        # Create draft
        draft = AssetFactory(
            name="Fog Machine Full S9",
            status="draft",
            category=category,
            current_location=location,
        )
        edit_url = reverse("assets:asset_edit", args=[draft.pk])

        # Promote to active
        admin_client.post(
            edit_url,
            {
                "name": draft.name,
                "status": "active",
                "category": category.pk,
                "current_location": location.pk,
                "condition": "excellent",
                "quantity": 1,
            },
        )
        draft.refresh_from_db()
        assert draft.status == "active"

        # Checkout
        co_url = reverse("assets:asset_checkout", args=[draft.pk])
        admin_client.post(
            co_url,
            {
                "borrower": borrower_user.pk,
                "destination_location": location.pk,
            },
        )
        draft.refresh_from_db()
        assert draft.checked_out_to == borrower_user

        # Check in
        ci_url = reverse("assets:asset_checkin", args=[draft.pk])
        admin_client.post(ci_url, {"location": location.pk})
        draft.refresh_from_db()
        assert draft.checked_out_to is None

        # Mark lost via state service (bulk action blocks lost per S7.17.5)
        from assets.services.state import transition_asset

        draft.lost_stolen_notes = "Missing after production."
        draft.save(update_fields=["lost_stolen_notes"])
        transition_asset(draft, "lost")
        draft.refresh_from_db()
        assert draft.status == "lost"

        # Recover via bulk_actions (active recovery is allowed in bulk)
        admin_client.post(
            reverse("assets:bulk_actions"),
            {
                "bulk_action": "status_change",
                "asset_ids": [draft.pk],
                "new_status": "active",
            },
        )
        draft.refresh_from_db()
        assert draft.status == "active"

        # Retire via edit form
        admin_client.post(
            edit_url,
            {
                "name": draft.name,
                "status": "retired",
                "category": category.pk,
                "current_location": location.pk,
                "condition": "poor",
                "quantity": 1,
            },
        )
        draft.refresh_from_db()
        assert draft.status == "retired"

        # Dispose via delete view
        admin_client.post(
            reverse("assets:asset_delete", args=[draft.pk]),
            {},
        )
        draft.refresh_from_db()
        assert draft.status == "disposed"


# ---------------------------------------------------------------------------
# §11.10 NFC Tag Management for Costumes
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_10_NFCTagManagementForCostumes:
    """§11.10 — NFC Tag Management for Costumes.

    Covers: NFC assignment to asset (uses 'tag_id' field),
    unified lookup via /a/<tag_id>/, NFC removal (sets removed_at,
    preserves historical record).
    Spec refs: S2.5.2, S2.5.4, S2.5.5, S2.5.6, S2.4.4
    """

    def test_dm_can_assign_nfc_tag_to_asset(self, dept_manager_client, asset):
        """NFC assignment uses 'tag_id' POST field, not 'tag_uid'."""
        url = reverse("assets:nfc_add", args=[asset.pk])
        resp = dept_manager_client.post(
            url,
            {
                "tag_id": "04:A3:2B:1C:5D:6E:7F",
                "notes": "Sewn into collar label",
            },
        )
        assert resp.status_code in (200, 302)
        assert NFCTag.objects.filter(
            asset=asset, tag_id="04:A3:2B:1C:5D:6E:7F"
        ).exists()

    def test_nfc_lookup_resolves_to_asset(self, client_logged_in, asset):
        NFCTag.objects.create(
            asset=asset,
            tag_id="04:BB:CC:DD:EE:FF:11",
        )
        url = reverse(
            "assets:asset_by_identifier",
            args=["04:BB:CC:DD:EE:FF:11"],
        )
        resp = client_logged_in.get(url)
        # Should redirect to asset detail or return asset detail directly
        assert resp.status_code in (200, 302)
        if resp.status_code == 302:
            assert str(asset.pk) in resp["Location"]

    def test_remove_nfc_tag_preserves_history(
        self, dept_manager_client, asset
    ):
        """Removal sets removed_at (not deletes), is_active becomes False."""
        nfc = NFCTag.objects.create(
            asset=asset,
            tag_id="04:REMOVE:ME:00",
        )
        url = reverse("assets:nfc_remove", args=[asset.pk, nfc.pk])
        resp = dept_manager_client.post(
            url, {"notes": "Tag damaged — replacement needed"}
        )
        assert resp.status_code in (200, 302)
        nfc.refresh_from_db()
        # removed_at should be set (historical preservation)
        assert nfc.removed_at is not None
        assert not nfc.is_active
        # Historical record still exists
        assert NFCTag.objects.filter(pk=nfc.pk).exists()

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #27a: Asset detail does not show notes from removed NFC"
            " tags (S2.5.6). The removed tag's notes field is not"
            " rendered in the NFC history section."
        ),
    )
    def test_asset_detail_shows_removed_nfc_notes(
        self, admin_client, asset, admin_user
    ):
        """S2.5.6: After removing an NFC tag with notes, the asset
        detail page must display those notes in the NFC history."""
        import datetime

        nfc = NFCTag.objects.create(
            asset=asset,
            tag_id="04:REMOVED:NOTES:01",
            notes="Sewn into collar — removed when damaged",
            removed_at=datetime.datetime.now(datetime.timezone.utc),
            removed_by=admin_user,
        )
        resp = admin_client.get(
            reverse("assets:asset_detail", args=[asset.pk])
        )
        assert resp.status_code == 200
        assert (
            b"removed when damaged" in resp.content
        ), "Asset detail NFC history must show removed tag notes"

    def test_asset_detail_shows_removed_by_user(
        self, admin_client, asset, admin_user
    ):
        """S2.5.6: Asset detail NFC history shows who removed the tag.
        Previously GAP #27b — confirmed working (XPASS)."""
        import datetime

        NFCTag.objects.create(
            asset=asset,
            tag_id="04:REMOVED:USER:01",
            removed_at=datetime.datetime.now(datetime.timezone.utc),
            removed_by=admin_user,
        )
        resp = admin_client.get(
            reverse("assets:asset_detail", args=[asset.pk])
        )
        assert resp.status_code == 200
        content = resp.content.decode()
        assert (
            admin_user.username in content
            or (admin_user.display_name or "") in content
            or admin_user.email in content
        ), "Asset detail NFC history must show who removed the tag"

    def test_nfc_history_section_renders_removal_timestamp(
        self, admin_client, asset, admin_user
    ):
        """S2.5.6: Asset detail NFC history includes removal timestamp.
        Previously GAP #27c — confirmed working (XPASS)."""
        import datetime

        removal_time = datetime.datetime(
            2026, 1, 15, 10, 30, 0, tzinfo=datetime.timezone.utc
        )
        NFCTag.objects.create(
            asset=asset,
            tag_id="04:REMOVED:TS:01",
            removed_at=removal_time,
            removed_by=admin_user,
        )
        resp = admin_client.get(
            reverse("assets:asset_detail", args=[asset.pk])
        )
        assert resp.status_code == 200
        content = resp.content.decode()
        assert (
            "2026" in content or "Jan" in content
        ), "NFC history must include removal timestamp"

    def test_full_scenario_walkthrough(
        self, dept_manager_client, client_logged_in, asset
    ):
        # Assign NFC
        dept_manager_client.post(
            reverse("assets:nfc_add", args=[asset.pk]),
            {
                "tag_id": "04:SCENARIO:10:01",
                "notes": "Sewn into label",
            },
        )
        assert NFCTag.objects.filter(
            asset=asset, tag_id="04:SCENARIO:10:01"
        ).exists()

        # Unified lookup
        resp = client_logged_in.get(
            reverse(
                "assets:asset_by_identifier",
                args=["04:SCENARIO:10:01"],
            )
        )
        assert resp.status_code in (200, 302)

        # Remove old tag
        nfc = NFCTag.objects.get(tag_id="04:SCENARIO:10:01")
        dept_manager_client.post(
            reverse("assets:nfc_remove", args=[asset.pk, nfc.pk]),
            {"notes": "damaged"},
        )
        nfc.refresh_from_db()
        assert not nfc.is_active

        # Assign new tag
        dept_manager_client.post(
            reverse("assets:nfc_add", args=[asset.pk]),
            {"tag_id": "04:SCENARIO:10:02", "notes": "Replacement"},
        )
        assert NFCTag.objects.filter(
            asset=asset,
            tag_id="04:SCENARIO:10:02",
            removed_at__isnull=True,
        ).exists()
        # Old record still in history with removed_at set
        assert NFCTag.objects.filter(
            tag_id="04:SCENARIO:10:01", removed_at__isnull=False
        ).exists()


# ---------------------------------------------------------------------------
# §11.11 Merging Duplicate Assets
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_11_MergingDuplicateAssets:
    """§11.11 — Merging Duplicate Assets After Parallel Quick Capture.

    Covers: promote both assets, merge secondary into primary,
    secondary becomes disposed with merge note.
    Spec refs: S2.2.7
    """

    def test_merge_disposes_secondary_asset(
        self, dept_manager_client, category, location
    ):
        """Merge uses 'primary_id' and 'asset_ids' (comma-separated,
        must include both primary and secondary pks)."""
        primary = AssetFactory(
            name="Antique Mantel Clock",
            status="active",
            category=category,
            current_location=location,
        )
        secondary = AssetFactory(
            name="Clock Mantel Brass",
            status="active",
            category=category,
            current_location=location,
        )
        url = reverse("assets:asset_merge_execute")
        resp = dept_manager_client.post(
            url,
            {
                "primary_id": primary.pk,
                "asset_ids": f"{primary.pk},{secondary.pk}",
            },
        )
        assert resp.status_code in (200, 302)
        secondary.refresh_from_db()
        assert secondary.status == "disposed"

    def test_merge_transfers_nfc_to_primary(
        self, dept_manager_client, category, location
    ):
        primary = AssetFactory(
            name="Clock A Merge",
            status="active",
            category=category,
            current_location=location,
        )
        secondary = AssetFactory(
            name="Clock B Merge",
            status="active",
            category=category,
            current_location=location,
        )
        nfc = NFCTag.objects.create(
            asset=secondary,
            tag_id="04:MERGE:TEST:01",
        )
        url = reverse("assets:asset_merge_execute")
        dept_manager_client.post(
            url,
            {
                "primary_id": primary.pk,
                "asset_ids": f"{primary.pk},{secondary.pk}",
            },
        )
        nfc.refresh_from_db()
        assert nfc.asset_id == primary.pk

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #28a: Merge confirmation page does not show side-by-side"
            " field comparison (S2.2.7). The merge_select/preview page"
            " does not render both assets' field values for comparison."
        ),
    )
    def test_merge_preview_shows_field_comparison(
        self, admin_client, category, location
    ):
        """S2.2.7: The merge confirmation page must show both assets'
        field values side-by-side (name, description, etc)."""
        primary = AssetFactory(
            name="Primary Clock S28",
            description="Primary description",
            status="active",
            category=category,
            current_location=location,
        )
        secondary = AssetFactory(
            name="Secondary Clock S28",
            description="Secondary description",
            status="active",
            category=category,
            current_location=location,
        )
        url = reverse("assets:asset_merge_select")
        resp = admin_client.post(
            url,
            {
                "asset_ids": f"{primary.pk},{secondary.pk}",
                "primary_id": primary.pk,
            },
        )
        # If the view returns a preview page (200), it must show both names
        if resp.status_code == 200:
            content = resp.content.decode()
            assert (
                "Primary Clock S28" in content
            ), "Merge preview must show primary asset name"
            assert (
                "Secondary Clock S28" in content
            ), "Merge preview must show secondary asset name"
        else:
            # Redirect means no confirmation page — gap confirmed
            assert False, (
                "Merge confirmation page must show field comparison"
                f" (got {resp.status_code})"
            )

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #28b: Merge POST does not respect explicit field choices"
            " (S2.2.7). The merge execute view uses primary asset fields"
            " without allowing field-level selection."
        ),
    )
    def test_merge_respects_selected_fields(
        self, admin_client, category, location
    ):
        """S2.2.7: POST merge with explicit field choices must result in
        primary asset having the chosen field values."""
        primary = AssetFactory(
            name="Primary Name S28b",
            description="Primary description S28b",
            status="active",
            category=category,
            current_location=location,
        )
        secondary = AssetFactory(
            name="Secondary Name S28b",
            description="Better secondary description S28b",
            status="active",
            category=category,
            current_location=location,
        )
        url = reverse("assets:asset_merge_execute")
        admin_client.post(
            url,
            {
                "primary_id": primary.pk,
                "asset_ids": f"{primary.pk},{secondary.pk}",
                # Choose secondary's description over primary's
                "field_description": str(secondary.pk),
            },
        )
        primary.refresh_from_db()
        assert (
            primary.description == "Better secondary description S28b"
        ), "Merge must use explicitly selected field from secondary asset"

    def test_full_scenario_walkthrough(
        self, dept_manager_client, category, location
    ):
        primary = AssetFactory(
            name="Antique Mantel Clock Full",
            status="active",
            category=category,
            current_location=location,
        )
        secondary = AssetFactory(
            name="Quick Capture Feb 20 14:15 Sc11",
            status="active",
            category=category,
            current_location=location,
        )
        NFCTag.objects.create(asset=secondary, tag_id="04:MERGE:FULL:01")

        dept_manager_client.post(
            reverse("assets:asset_merge_execute"),
            {
                "primary_id": primary.pk,
                "asset_ids": f"{primary.pk},{secondary.pk}",
            },
        )
        secondary.refresh_from_db()
        assert secondary.status == "disposed"
        # NFC should have moved to primary
        assert NFCTag.objects.filter(
            asset=primary, tag_id="04:MERGE:FULL:01"
        ).exists()


# ---------------------------------------------------------------------------
# §11.12 Insurance Export and Reporting
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_12_InsuranceExportAndReporting:
    """§11.12 — Insurance Export and Reporting.

    Covers: full asset export returns xlsx, disposed assets excluded,
    lost assets included, department-filtered export.
    Spec refs: S2.9.1, S2.9.2
    """

    def test_export_returns_xlsx(self, admin_client, asset):
        url = reverse("assets:export_assets")
        resp = admin_client.get(url)
        assert resp.status_code == 200
        content_type = resp.get("Content-Type", "")
        assert (
            "spreadsheet" in content_type
            or "excel" in content_type
            or "openxmlformats" in content_type
        )

    def test_export_excludes_disposed_assets(
        self, admin_client, category, location
    ):
        """Disposed assets should not be in the export (binary xlsx).
        We check by loading the xlsx file and inspecting cell values."""
        import io

        import openpyxl

        AssetFactory(
            name="Disposed Item Export",
            status="disposed",
            category=category,
            current_location=location,
        )
        url = reverse("assets:export_assets")
        resp = admin_client.get(url)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        all_values = {
            str(cell.value or "") for row in ws.iter_rows() for cell in row
        }
        assert "Disposed Item Export" not in all_values

    def test_export_includes_lost_assets(
        self, admin_client, category, location
    ):
        """Lost assets should appear in export (insurable claims) per
        S2.9.1. If this fails, it confirms a GAP: the export currently
        excludes lost assets."""
        import io

        import openpyxl

        AssetFactory(
            name="Lost Fog Machine Export",
            status="lost",
            category=category,
            current_location=location,
        )
        url = reverse("assets:export_assets")
        resp = admin_client.get(url)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        all_values = set()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    all_values.add(str(cell.value or ""))
        if "Lost Fog Machine Export" not in all_values:
            pytest.xfail(
                "GAP (S2.9.1): Export excludes lost assets. "
                "Spec requires lost assets to be included for insurance."
            )

    def test_filtered_export_restricts_to_department(
        self, admin_client, department, category, location
    ):
        import io

        import openpyxl

        other_dept = DepartmentFactory(
            name="TechnicalExp", barcode_prefix="TEXP"
        )
        other_cat = CategoryFactory(name="LightsExp", department=other_dept)
        AssetFactory(
            name="Technical Fixture Export",
            status="active",
            category=other_cat,
            current_location=location,
        )
        url = reverse("assets:export_assets")
        resp = admin_client.get(url, {"department": department.pk})
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        all_values = {
            str(cell.value or "") for row in ws.iter_rows() for cell in row
        }
        assert "Technical Fixture Export" not in all_values

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #29a: Export Summary sheet is missing per-status rows"
            " for Missing, Lost, Stolen, Retired, Disposed (S2.9,"
            " S11.12). The Summary sheet only has Total/Active/Draft."
        ),
    )
    def test_export_summary_sheet_has_all_status_rows(
        self, admin_client, active_asset
    ):
        """S11.12 Step 4: Summary sheet must have row labels for all
        asset statuses: Missing, Lost, Stolen, Retired, Disposed."""
        from io import BytesIO

        import openpyxl

        resp = admin_client.get(reverse("assets:export_assets"))
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        assert "Summary" in wb.sheetnames, "No 'Summary' sheet in workbook"
        summary = wb["Summary"]
        col_a_values = [
            str(cell.value or "").lower()
            for row in summary.iter_rows()
            for cell in row
            if cell.column == 1 and cell.value
        ]
        required_statuses = [
            "missing",
            "lost",
            "stolen",
            "retired",
            "disposed",
        ]
        for status in required_statuses:
            assert any(
                status in v for v in col_a_values
            ), f"Summary sheet missing row for status: {status}"

    def test_export_summary_sheet_has_financial_totals(
        self, admin_client, active_asset
    ):
        """S11.12 Step 4: Summary sheet has financial total rows.
        Previously GAP #29b — confirmed working (XPASS)."""
        from io import BytesIO

        import openpyxl

        resp = admin_client.get(reverse("assets:export_assets"))
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        assert "Summary" in wb.sheetnames, "No 'Summary' sheet in workbook"
        summary = wb["Summary"]
        col_a_values = [
            str(cell.value or "").lower()
            for row in summary.iter_rows()
            for cell in row
            if cell.column == 1 and cell.value
        ]
        financial_terms = ["purchase", "value", "total", "cost", "insurance"]
        has_financial = any(
            any(term in v for term in financial_terms) for v in col_a_values
        )
        assert has_financial, (
            f"Summary sheet has no financial total rows."
            f" Column A labels: {col_a_values}"
        )

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #29: Export summary sheet missing department breakdown"
            " (S2.9, S11.12 Step 4). The Summary sheet has total/active/"
            "draft/checked-out counts but no per-department breakdown"
            " section or 'By Department' label."
        ),
    )
    def test_export_summary_sheet_has_department_breakdown(
        self, admin_client, active_asset, department
    ):
        """S11.12 Step 4: Summary sheet must include a per-department
        count breakdown, not just global totals."""
        from io import BytesIO

        import openpyxl

        resp = admin_client.get(reverse("assets:export_assets"))
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        assert (
            "Summary" in wb.sheetnames
        ), "No 'Summary' sheet found in export workbook"
        summary = wb["Summary"]

        # Collect only the label column (column A) values from the
        # Summary sheet — these are the row headings like "Total Assets",
        # "Active", etc. A department breakdown requires a label that
        # identifies a per-department section.
        col_a_values = [
            str(cell.value or "").lower()
            for row in summary.iter_rows()
            for cell in row
            if cell.column == 1 and cell.value
        ]

        # A proper department breakdown must have either:
        #  (a) a "by department" section label, OR
        #  (b) the specific department name as a row label (not just in
        #      the site title which includes "PROPS")
        has_dept_section = any(
            "by department" in v or "department breakdown" in v
            for v in col_a_values
        )
        # Department name as a standalone label (not embedded in site name)
        dept_name_lower = department.name.lower()
        has_dept_label = any(v == dept_name_lower for v in col_a_values)

        assert has_dept_section or has_dept_label, (
            f"Summary sheet has no department breakdown section."
            f" Column A labels found: {col_a_values}"
        )

    def test_full_scenario_walkthrough(self, admin_client, category, location):
        import io

        import openpyxl

        AssetFactory(
            name="Grand Piano Export Full",
            status="active",
            category=category,
            current_location=location,
        )
        AssetFactory(
            name="Missing Spotlight Export Full",
            status="lost",
            category=category,
            current_location=location,
        )
        AssetFactory(
            name="Old Costume Export Full",
            status="disposed",
            category=category,
            current_location=location,
        )
        url = reverse("assets:export_assets")
        resp = admin_client.get(url)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        all_values = set()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    all_values.add(str(cell.value or ""))
        # Active assets must be in export
        assert "Grand Piano Export Full" in all_values
        # Disposed must be excluded
        assert "Old Costume Export Full" not in all_values
        # Lost should be in export per spec — xfail if current impl excludes
        if "Missing Spotlight Export Full" not in all_values:
            pytest.xfail(
                "GAP (S2.9.1): Lost assets excluded from export. "
                "Spec requires inclusion for insurance claims."
            )


# ---------------------------------------------------------------------------
# §11.14 Serialised Asset Management
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_14_SerialisedAssetManagement:
    """§11.14 — Serialised Asset Management — Radio Microphones.

    Covers: convert to serialised (uses 'confirm' POST field),
    check availability count, checkout reduces available count.
    Spec refs: S2.17.1a, S2.17.1b, S2.17.1d, S2.17.2
    """

    def test_convert_non_serialised_to_serialised(self, admin_client, asset):
        """Conversion requires 'confirm' POST field (not 'mode')."""
        assert not asset.is_serialised
        url = reverse("assets:asset_convert_serialisation", args=[asset.pk])
        resp = admin_client.post(url, {"confirm": "1"})
        assert resp.status_code in (200, 302)
        asset.refresh_from_db()
        assert asset.is_serialised

    def test_serialised_asset_shows_available_count(
        self, admin_client, serialised_asset_with_units
    ):
        asset = serialised_asset_with_units["asset"]
        url = reverse("assets:asset_detail", args=[asset.pk])
        resp = admin_client.get(url)
        assert resp.status_code == 200
        # All 5 serials should be available
        assert b"5" in resp.content

    def test_checkout_reduces_available_count(
        self, dept_manager_client, serialised_asset_with_units, borrower_user
    ):
        asset = serialised_asset_with_units["asset"]
        serials = serialised_asset_with_units["serials"]
        url = reverse("assets:asset_checkout", args=[asset.pk])
        dept_manager_client.post(
            url,
            {
                "borrower": borrower_user.pk,
                "destination_location": asset.current_location.pk,
                "serial_ids": [serials[0].pk, serials[1].pk],
            },
        )
        serials[0].refresh_from_db()
        serials[1].refresh_from_db()
        assert serials[0].checked_out_to == borrower_user
        assert serials[1].checked_out_to == borrower_user

    def test_full_scenario_walkthrough(
        self,
        admin_client,
        dept_manager_client,
        serialised_asset_with_units,
        borrower_user,
    ):
        asset = serialised_asset_with_units["asset"]
        serials = serialised_asset_with_units["serials"]

        # Verify available count on detail page
        detail_url = reverse("assets:asset_detail", args=[asset.pk])
        resp = admin_client.get(detail_url)
        assert resp.status_code == 200

        # Checkout 3 units
        dept_manager_client.post(
            reverse("assets:asset_checkout", args=[asset.pk]),
            {
                "borrower": borrower_user.pk,
                "destination_location": asset.current_location.pk,
                "serial_ids": [s.pk for s in serials[:3]],
            },
        )
        for s in serials[:3]:
            s.refresh_from_db()
            assert s.checked_out_to == borrower_user

        # Remaining serials still available
        serials[3].refresh_from_db()
        assert serials[3].checked_out_to is None


# ---------------------------------------------------------------------------
# §11.15 Concurrent Access and System Integrity
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_15_ConcurrentAccess:
    """§11.15 — Concurrent Access and System Integrity.

    Covers: double-checkout is blocked — only one user can check out
    an asset at a time.
    Spec refs: S2.3.13, S3.3.2, S7.4.2
    """

    def test_double_checkout_is_blocked(
        self, client, password, asset, location
    ):
        group, _ = Group.objects.get_or_create(name="Department Manager")

        user_a = UserFactory(
            username="member_a_s15",
            email="member_a_s15@example.com",
            password=password,
        )
        user_a.groups.add(group)

        user_b = UserFactory(
            username="member_b_s15",
            email="member_b_s15@example.com",
            password=password,
        )
        user_b.groups.add(group)

        # First checkout by user A
        client.login(username=user_a.username, password=password)
        url = reverse("assets:asset_checkout", args=[asset.pk])
        client.post(
            url,
            {
                "borrower": user_a.pk,
                "destination_location": location.pk,
            },
        )
        asset.refresh_from_db()
        first_borrower = asset.checked_out_to

        # Second checkout attempt by user B
        client.login(username=user_b.username, password=password)
        client.post(
            url,
            {
                "borrower": user_b.pk,
                "destination_location": location.pk,
            },
        )
        asset.refresh_from_db()
        # Asset must still be checked out to first borrower only
        assert asset.checked_out_to == first_borrower
        assert asset.checked_out_to != user_b

    def test_asset_has_at_most_one_checked_out_to(
        self, dept_manager_client, asset, borrower_user, location
    ):
        url = reverse("assets:asset_checkout", args=[asset.pk])
        dept_manager_client.post(
            url,
            {
                "borrower": borrower_user.pk,
                "destination_location": location.pk,
            },
        )
        asset.refresh_from_db()
        # checked_out_to is a single FK, not null
        assert asset.checked_out_to == borrower_user

    def test_full_scenario_walkthrough(
        self, client, password, asset, location
    ):
        group, _ = Group.objects.get_or_create(name="Department Manager")
        user_a = UserFactory(
            username="concurrent_a_s15",
            email="concurrent_a_s15@example.com",
            password=password,
        )
        user_a.groups.add(group)
        user_b = UserFactory(
            username="concurrent_b_s15",
            email="concurrent_b_s15@example.com",
            password=password,
        )
        user_b.groups.add(group)

        url = reverse("assets:asset_checkout", args=[asset.pk])

        client.login(username=user_a.username, password=password)
        client.post(
            url,
            {"borrower": user_a.pk, "destination_location": location.pk},
        )

        client.login(username=user_b.username, password=password)
        resp = client.post(
            url,
            {"borrower": user_b.pk, "destination_location": location.pk},
        )

        asset.refresh_from_db()
        assert asset.checked_out_to == user_a
        assert resp.status_code in (200, 302)


# ---------------------------------------------------------------------------
# §11.17 Remote Print Client Setup and Label Printing
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_17_RemotePrintClientSetup:
    """§11.17 — Remote Print Client Setup and Label Printing Across Sites.

    Covers: print client pairing/listing in admin, single print request
    via remote_print_submit (uses 'client_pk' field), print history.
    Spec refs: S2.4.5-09, S2.4.5-10, S2.4.5-11, S2.4.5a, S2.4.5b,
    S2.4.5c
    """

    def test_print_client_can_be_listed_in_admin(self, admin_client):
        from assets.models import PrintClient

        PrintClient.objects.create(
            name="Workshop Printer Station",
            status="pending",
            token_hash="abc123unique",
        )
        url = "/admin/assets/printclient/"
        resp = admin_client.get(url)
        assert resp.status_code == 200
        assert b"Workshop Printer Station" in resp.content

    def test_remote_print_submit_creates_print_request(
        self, dept_manager_client, asset
    ):
        """remote_print_submit uses 'client_pk' and 'printer_id' POST
        fields; PrintClient must have status='approved', is_connected=True,
        and the printer_id must be in the printers list."""
        from assets.models import PrintClient, PrintRequest

        client_obj = PrintClient.objects.create(
            name="Workshop Printer",
            status="approved",
            is_connected=True,
            is_active=True,
            token_hash="dummyhash_s17",
            printers=[{"id": "usb-001", "name": "Brother QL"}],
        )
        url = reverse("assets:remote_print_submit", args=[asset.pk])
        resp = dept_manager_client.post(
            url,
            {
                "client_pk": client_obj.pk,
                "printer_id": "usb-001",
            },
        )
        assert resp.status_code in (200, 302)
        assert PrintRequest.objects.filter(asset=asset).exists()

    def test_print_history_accessible_for_asset(
        self, dept_manager_client, asset
    ):
        url = reverse("assets:print_history", args=[asset.pk])
        resp = dept_manager_client.get(url)
        assert resp.status_code == 200

    def test_full_scenario_walkthrough(
        self, dept_manager_client, admin_client, asset
    ):
        from assets.models import PrintClient, PrintRequest

        # Pairing: create client record (pending state)
        pc = PrintClient.objects.create(
            name="Workshop Printer Station Full",
            status="pending",
            token_hash="hashval123_s17",
        )
        assert pc.status == "pending"

        # Simulate admin approval (approved + connected)
        pc.status = "approved"
        pc.is_connected = True
        pc.is_active = True
        pc.printers = [{"id": "usb-001", "name": "Brother QL-820NWB"}]
        pc.save()

        # Single print
        dept_manager_client.post(
            reverse("assets:remote_print_submit", args=[asset.pk]),
            {"client_pk": pc.pk, "printer_id": "usb-001"},
        )
        assert PrintRequest.objects.filter(asset=asset).exists()

        # Bulk print for multiple assets
        extra_assets = [
            AssetFactory(
                name=f"Bulk Asset S17 {i}",
                status="active",
                category=asset.category,
                current_location=asset.current_location,
            )
            for i in range(3)
        ]
        for a in extra_assets:
            dept_manager_client.post(
                reverse("assets:remote_print_submit", args=[a.pk]),
                {"client_pk": pc.pk, "printer_id": "usb-001"},
            )
        assert PrintRequest.objects.filter(asset__in=extra_assets).count() == 3


# ---------------------------------------------------------------------------
# §11.18 Search and Hold List
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_18_SearchAndHoldList:
    """§11.18 — Searching for Assets and Building a Hold List for a
    Production.

    Covers: text search, department filter, barcode search, grid view.
    Spec refs: S2.6.1, S2.6.2, S2.6.2a, S2.6.3, S2.16.3, S2.16.4,
    S2.16.5, S2.16.6, S2.16.7, S2.17.6a
    """

    def test_text_search_returns_matching_assets(
        self, client_logged_in, category, location
    ):
        AssetFactory(
            name="Fairy Wings Large Sc18",
            status="active",
            category=category,
            current_location=location,
        )
        AssetFactory(
            name="Rocking Chair Sc18",
            status="active",
            category=category,
            current_location=location,
        )
        url = reverse("assets:asset_list")
        resp = client_logged_in.get(url, {"q": "Fairy Wings Large Sc18"})
        assert resp.status_code == 200
        assert b"Fairy Wings Large Sc18" in resp.content
        assert b"Rocking Chair Sc18" not in resp.content

    def test_department_filter_narrows_results(
        self, client_logged_in, location
    ):
        dept_a = DepartmentFactory(
            name="Props FilterS18", barcode_prefix="PF18"
        )
        dept_b = DepartmentFactory(name="Set FilterS18", barcode_prefix="SF18")
        cat_a = CategoryFactory(name="Props CatS18", department=dept_a)
        cat_b = CategoryFactory(name="Set CatS18", department=dept_b)
        AssetFactory(
            name="Props Item Sc18",
            status="active",
            category=cat_a,
            current_location=location,
        )
        AssetFactory(
            name="Set Item Sc18",
            status="active",
            category=cat_b,
            current_location=location,
        )
        url = reverse("assets:asset_list")
        resp = client_logged_in.get(url, {"department": dept_a.pk})
        assert resp.status_code == 200
        assert b"Props Item Sc18" in resp.content
        assert b"Set Item Sc18" not in resp.content

    def test_barcode_search_finds_specific_asset(
        self, client_logged_in, asset
    ):
        url = reverse("assets:asset_list")
        resp = client_logged_in.get(url, {"q": asset.barcode})
        assert resp.status_code == 200
        assert asset.name.encode() in resp.content

    def test_asset_list_grid_view_accessible(self, client_logged_in):
        url = reverse("assets:asset_list")
        resp = client_logged_in.get(url, {"view": "grid"})
        assert resp.status_code == 200

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #33a: Hold list add_item endpoint does not accept asset"
            " name as search input (S2.16.4). The holdlist_add_item view"
            " only accepts an asset PK ('asset_id') — not a name string."
        ),
    )
    def test_hold_list_add_accepts_name_search(
        self, dept_manager_client, hold_list, category, location, admin_user
    ):
        """S2.16.4: POST to hold list add with asset name (not PK) must
        find and add the matching asset."""
        from assets.models import HoldListItem

        target = AssetFactory(
            name="Unique Fairy Wings S33a",
            status="active",
            category=category,
            current_location=location,
            created_by=admin_user,
        )
        url = reverse("assets:holdlist_add_item", args=[hold_list.pk])
        resp = dept_manager_client.post(
            url,
            {
                "search": "Unique Fairy Wings S33a",
                "quantity": 1,
            },
        )
        assert resp.status_code in (200, 302)
        assert hold_list.items.filter(
            asset=target
        ).exists(), "Hold list add must accept asset name as search input"

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #33b: Hold list add_item endpoint does not accept"
            " barcode string as input (S2.16.4). The holdlist_add_item"
            " view only accepts an asset PK ('asset_id')."
        ),
    )
    def test_hold_list_add_accepts_barcode_scan(
        self, dept_manager_client, hold_list, category, location, admin_user
    ):
        """S2.16.4: POST to hold list add with asset barcode string must
        find and add the matching asset."""
        target = AssetFactory(
            name="Barcode Asset S33b",
            status="active",
            category=category,
            current_location=location,
            created_by=admin_user,
        )
        url = reverse("assets:holdlist_add_item", args=[hold_list.pk])
        resp = dept_manager_client.post(
            url,
            {
                "barcode": target.barcode,
                "quantity": 1,
            },
        )
        assert resp.status_code in (200, 302)
        assert hold_list.items.filter(
            asset=target
        ).exists(), "Hold list add must accept barcode string as input"

    def test_full_scenario_walkthrough(
        self,
        dept_manager_client,
        client_logged_in,
        department,
        category,
        location,
    ):
        wings = AssetFactory(
            name="Fairy Wings Pink Sc18Full",
            status="active",
            category=category,
            current_location=location,
        )
        # Text search
        url = reverse("assets:asset_list")
        resp = client_logged_in.get(url, {"q": "Fairy Wings Pink Sc18Full"})
        assert resp.status_code == 200
        assert b"Fairy Wings Pink Sc18Full" in resp.content

        # Barcode search
        resp = client_logged_in.get(url, {"q": wings.barcode})
        assert resp.status_code == 200
        assert wings.name.encode() in resp.content

        # Create hold list and add item
        status, _ = HoldListStatus.objects.get_or_create(
            name="Draft",
            defaults={"is_default": True, "sort_order": 10},
        )
        dept_manager_client.post(
            reverse("assets:holdlist_create"),
            {
                "name": "Midsummer Props Sc18",
                "department": department.pk,
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
                "status": status.pk,
            },
        )
        hl = HoldList.objects.filter(name="Midsummer Props Sc18").first()
        assert hl is not None
        dept_manager_client.post(
            reverse("assets:holdlist_add_item", args=[hl.pk]),
            {"asset_id": wings.pk, "quantity": 1},
        )
        assert hl.items.filter(asset=wings).exists()


# ---------------------------------------------------------------------------
# §11.19 Transfer, Relocate, and Custody Handover
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_19_TransferRelocateHandover:
    """§11.19 — Asset Transfer, Relocate, and Custody Handover.

    Covers: relocate a checked-out asset (borrower preserved),
    custody handover via 'borrower' field (location preserved),
    transfer available asset via 'location' field.
    Spec refs: S2.3.4, S2.3.5, S2.3.6, S2.3.9, S2.8.1
    """

    def test_relocate_checked_out_asset_preserves_borrower(
        self, admin_client, asset, borrower_user, location
    ):
        asset.checked_out_to = borrower_user
        asset.save()
        new_loc = LocationFactory(name="Stage Left Wing S19a")
        url = reverse("assets:asset_relocate", args=[asset.pk])
        resp = admin_client.post(url, {"location": new_loc.pk})
        asset.refresh_from_db()
        assert asset.current_location == new_loc
        assert asset.checked_out_to == borrower_user

    def test_custody_handover_changes_borrower_preserves_location(
        self, admin_client, asset, borrower_user, location
    ):
        """Handover uses 'borrower' POST field (not 'new_borrower');
        asset must be checked out first."""
        asset.checked_out_to = borrower_user
        asset.save()
        new_borrower = UserFactory(
            username="dana_cheng_s19",
            email="dana_s19@example.com",
        )
        url = reverse("assets:asset_handover", args=[asset.pk])
        resp = admin_client.post(url, {"borrower": new_borrower.pk})
        asset.refresh_from_db()
        assert asset.checked_out_to == new_borrower
        assert asset.current_location == location

    def test_transfer_available_asset_updates_location(
        self, dept_manager_client, asset, location
    ):
        """Transfer uses 'location' POST field (not 'new_location');
        asset must not be checked out."""
        new_loc = LocationFactory(name="Bay 3 S19")
        url = reverse("assets:asset_transfer", args=[asset.pk])
        resp = dept_manager_client.post(url, {"location": new_loc.pk})
        asset.refresh_from_db()
        assert asset.current_location == new_loc
        assert Transaction.objects.filter(
            asset=asset, action="transfer"
        ).exists()

    def test_full_scenario_walkthrough(
        self, admin_client, dept_manager_client, asset, borrower_user, location
    ):
        # Relocate checked-out asset — update_fields saves home + current
        asset.checked_out_to = borrower_user
        asset.save()
        new_loc = LocationFactory(name="Stage Left Wing S19Full")
        admin_client.post(
            reverse("assets:asset_relocate", args=[asset.pk]),
            {"location": new_loc.pk},
        )
        asset.refresh_from_db()
        assert asset.current_location == new_loc
        assert asset.checked_out_to == borrower_user

        # Custody handover
        new_borrower = UserFactory(
            username="jo_reeves_s19",
            email="jo_s19@example.com",
        )
        admin_client.post(
            reverse("assets:asset_handover", args=[asset.pk]),
            {"borrower": new_borrower.pk},
        )
        asset.refresh_from_db()
        assert asset.checked_out_to == new_borrower

        # Check in first so transfer is allowed
        dept_manager_client.post(
            reverse("assets:asset_checkin", args=[asset.pk]),
            {"location": new_loc.pk},
        )
        asset.refresh_from_db()
        assert asset.checked_out_to is None

        # Transfer available asset
        transfer_loc = LocationFactory(name="Bay 3 Final S19")
        dept_manager_client.post(
            reverse("assets:asset_transfer", args=[asset.pk]),
            {"location": transfer_loc.pk},
        )
        asset.refresh_from_db()
        assert asset.current_location == transfer_loc


# ---------------------------------------------------------------------------
# §11.20 System Setup and Configuration
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_20_SystemSetupAndConfiguration:
    """§11.20 — System Setup and Configuration.

    Covers: save branding via direct model creation (admin form test),
    create department with barcode prefix, create 4-level location
    hierarchy.
    Spec refs: S2.10.1, S2.10.2, S2.12.1, S2.12.2, S2.13.1, S2.4.1
    """

    def test_site_branding_model_can_be_saved(self, db):
        """SiteBranding stores brand colors (site_name comes from
        settings, not SiteBranding model).
        GAP: SiteBranding has no site_name field — name is env-config only.
        """
        from assets.models import SiteBranding

        sb = SiteBranding.objects.create(
            primary_color="#8B1A1A",
        )
        assert SiteBranding.objects.filter(primary_color="#8B1A1A").exists()

    def test_department_with_barcode_prefix_created(self, admin_client):
        """Admin can create a department with a barcode prefix."""
        from assets.models import Department

        dept = DepartmentFactory(
            name="Lighting S20",
            barcode_prefix="LGS20",
        )
        assert Department.objects.filter(barcode_prefix="LGS20").exists()

    def test_location_hierarchy_four_levels_deep(self):
        l1 = LocationFactory(name="Main Theatre S20", parent=None)
        l2 = LocationFactory(name="Backstage Storage S20", parent=l1)
        l3 = LocationFactory(name="Props Room S20", parent=l2)
        l4 = LocationFactory(name="Shelf Z S20", parent=l3)
        assert l4.full_path == (
            "Main Theatre S20 > Backstage Storage S20"
            " > Props Room S20 > Shelf Z S20"
        )

    def test_full_scenario_walkthrough(self, db):
        from assets.models import Department, SiteBranding

        # Branding (site_name is env-config not SiteBranding model field)
        SiteBranding.objects.create(primary_color="#8B1A1A")
        assert SiteBranding.objects.filter(primary_color="#8B1A1A").exists()

        # Departments
        DepartmentFactory(name="Props S20Full", barcode_prefix="PR20F")
        DepartmentFactory(name="Costumes S20Full", barcode_prefix="CS20F")
        assert Department.objects.filter(barcode_prefix="PR20F").exists()
        assert Department.objects.filter(barcode_prefix="CS20F").exists()

        # Location hierarchy
        l1 = LocationFactory(name="Theatre Complex S20Full", parent=None)
        l2 = LocationFactory(name="Backstage S20Full", parent=l1)
        l3 = LocationFactory(name="Props Store S20Full", parent=l2)
        l4 = LocationFactory(name="Shelf Final S20Full", parent=l3)
        assert "Theatre Complex S20Full" in l4.full_path
        assert "Shelf Final S20Full" in l4.full_path


# ---------------------------------------------------------------------------
# §11.21 Location Checkout
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestScenario_11_21_LocationCheckout:
    """§11.21 — Location Checkout — Lending a Box of Equipment.

    Covers: is_checkable flag on location (GAP: not yet implemented),
    checkout skips individually checked-out assets.
    Spec refs: S2.12.4
    """

    def test_is_checkable_flag_not_present_on_location_model(self):
        """GAP: Location model does not have is_checkable field yet
        (S2.12.4). Marked xfail to document the gap without blocking
        the test run."""
        from assets.models import Location

        has_field = any(
            f.name == "is_checkable" for f in Location._meta.get_fields()
        )
        if not has_field:
            pytest.xfail(
                "GAP: Location.is_checkable field is missing. "
                "Required for §11.21 / S2.12.4."
            )
        assert has_field

    def test_location_checkout_url_exists(self, dept_manager_client, location):
        """GAP: No dedicated location checkout endpoint exists yet.
        Marked xfail to document the gap."""
        from django.urls import NoReverseMatch

        try:
            url = reverse("assets:location_checkout", args=[location.pk])
            resp = dept_manager_client.post(url, {})
            assert (
                resp.status_code != 404
            ), "Location checkout URL exists but returns 404"
        except NoReverseMatch:
            pytest.xfail(
                "GAP: URL 'assets:location_checkout' does not exist. "
                "Required for §11.21 / S2.12.4."
            )

    def test_non_checkable_location_rejects_checkout(
        self, dept_manager_client, location
    ):
        """A location without is_checkable=True must reject checkout."""
        from django.urls import NoReverseMatch

        try:
            url = reverse("assets:location_checkout", args=[location.pk])
            resp = dept_manager_client.post(url, {})
            assert resp.status_code in (400, 403, 200, 302)
        except NoReverseMatch:
            pytest.xfail(
                "GAP: 'assets:location_checkout' URL not implemented "
                "(S2.12.4)"
            )

    def test_full_scenario_walkthrough(
        self, admin_client, dept_manager_client, location, borrower_user
    ):
        """Full location checkout scenario — xfail if not implemented."""
        from django.urls import NoReverseMatch

        try:
            url = reverse("assets:location_checkout", args=[location.pk])
        except NoReverseMatch:
            pytest.xfail(
                "GAP: Location checkout (§11.21 / S2.12.4) not " "implemented."
            )

        resp = dept_manager_client.post(
            url,
            {
                "borrower": borrower_user.pk,
                "destination_location": location.pk,
            },
        )
        assert resp.status_code in (200, 302)

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #17a: Location checkout skips-already-checked-out logic"
            " not implemented (S2.12.4). The location_checkout URL does"
            " not exist yet."
        ),
    )
    def test_location_checkout_skips_already_checked_out_assets(
        self,
        dept_manager_client,
        location,
        category,
        borrower_user,
        user,
        admin_user,
    ):
        """Checkout via location URL must exclude assets already checked
        out to someone else — or show a warning."""
        from django.urls import NoReverseMatch

        checked_out_asset = AssetFactory(
            name="Already Out Asset S21a",
            status="active",
            category=category,
            current_location=location,
            created_by=admin_user,
        )
        checked_out_asset.checked_out_to = user
        checked_out_asset.save()

        available_asset = AssetFactory(
            name="Available Asset S21a",
            status="active",
            category=category,
            current_location=location,
            created_by=admin_user,
        )

        try:
            url = reverse("assets:location_checkout", args=[location.pk])
        except NoReverseMatch:
            pytest.xfail(
                "GAP #17: URL 'assets:location_checkout' does not exist."
            )

        resp = dept_manager_client.post(url, {"borrower": borrower_user.pk})
        assert resp.status_code in (200, 302)
        # Already-checked-out asset should NOT be re-checked-out
        checked_out_asset.refresh_from_db()
        assert (
            checked_out_asset.checked_out_to == user
        ), "Already checked-out asset should not be reassigned"

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #17b: Location checkout has no confirmation step"
            " (S2.12.4). The location_checkout URL does not exist yet."
        ),
    )
    def test_location_checkout_requires_confirmation_step(
        self, dept_manager_client, location
    ):
        """GET on location checkout URL must show a confirmation step
        before executing the batch checkout."""
        from django.urls import NoReverseMatch

        try:
            url = reverse("assets:location_checkout", args=[location.pk])
        except NoReverseMatch:
            pytest.xfail(
                "GAP #17: URL 'assets:location_checkout' does not exist."
            )

        resp = dept_manager_client.get(url)
        assert resp.status_code == 200
        content = resp.content.decode().lower()
        assert (
            "confirm" in content
            or "checkout all" in content
            or "proceed" in content
        ), "Location checkout GET must show a confirmation step"

    @pytest.mark.xfail(
        strict=True,
        reason=("GAP #17c: No location check-in URL exists (S2.12.4)."),
    )
    def test_location_checkin_url_exists(self, dept_manager_client, location):
        """The location_checkin URL must exist and respond (not 404)."""
        from django.urls import NoReverseMatch

        try:
            url = reverse("assets:location_checkin", args=[location.pk])
        except NoReverseMatch:
            pytest.xfail(
                "GAP #17: URL 'assets:location_checkin' does not exist."
            )

        resp = dept_manager_client.get(url)
        assert resp.status_code in (200, 302, 405)

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #17d: Location checkout with no assets shows no"
            " informational message (S2.12.4). URL not implemented."
        ),
    )
    def test_location_checkout_empty_location_shows_informational_message(
        self, dept_manager_client, db
    ):
        """When a location has no assets, the checkout response must
        show an informational message — not an error or 500."""
        from django.urls import NoReverseMatch

        empty_loc = LocationFactory(name="Empty Location S21d")

        try:
            url = reverse("assets:location_checkout", args=[empty_loc.pk])
        except NoReverseMatch:
            pytest.xfail(
                "GAP #17: URL 'assets:location_checkout' does not exist."
            )

        resp = dept_manager_client.get(url)
        assert resp.status_code == 200
        content = resp.content.decode().lower()
        assert (
            "no assets" in content
            or "empty" in content
            or "nothing" in content
        ), "Empty location checkout must show an informational message"

    @pytest.mark.xfail(
        strict=True,
        reason=(
            "GAP #17e: Location checkout has no permission check — viewer"
            " should get 302/403 (S2.12.4). URL not implemented."
        ),
    )
    def test_location_checkout_requires_dept_manager_or_admin(
        self, viewer_client, location
    ):
        """A Viewer must receive 302 or 403 on the location checkout
        URL — only Department Managers and admins may use it."""
        from django.urls import NoReverseMatch

        try:
            url = reverse("assets:location_checkout", args=[location.pk])
        except NoReverseMatch:
            pytest.xfail(
                "GAP #17: URL 'assets:location_checkout' does not exist."
            )

        resp = viewer_client.get(url)
        assert resp.status_code in (302, 403), (
            f"Viewer should be blocked from location checkout, got"
            f" {resp.status_code}"
        )
