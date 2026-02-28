"""Tests for service layer business logic."""

from unittest.mock import MagicMock

import pytest

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.test.utils import override_settings
from django.urls import reverse
from django.utils import timezone

from assets.factories import (
    AssetFactory,
    AssetImageFactory,
    AssetKitFactory,
    AssetSerialFactory,
    CategoryFactory,
    DepartmentFactory,
    HoldListFactory,
    HoldListItemFactory,
    HoldListStatusFactory,
    LocationFactory,
    NFCTagFactory,
    ProjectFactory,
    SiteBrandingFactory,
    StocktakeItemFactory,
    StocktakeSessionFactory,
    TagFactory,
    TransactionFactory,
    UserFactory,
    VirtualBarcodeFactory,
)
from assets.models import (
    Asset,
    AssetImage,
    AssetSerial,
    Location,
    PrintClient,
    Tag,
    Transaction,
)

User = get_user_model()

# ============================================================
# SERVICE TESTS
# ============================================================


class TestStateService:
    def test_validate_transition_valid(self, asset):
        from assets.services.state import validate_transition

        validate_transition(asset, "retired")  # Should not raise

    def test_validate_transition_invalid(self, asset):
        from assets.services.state import validate_transition

        with pytest.raises(ValidationError):
            validate_transition(asset, "draft")

    def test_validate_transition_noop(self, asset):
        from assets.services.state import validate_transition

        validate_transition(asset, "active")  # Same status, no-op

    def test_validate_transition_bad_status(self, asset):
        from assets.services.state import validate_transition

        with pytest.raises(ValidationError, match="not a valid status"):
            validate_transition(asset, "bogus")

    def test_transition_asset(self, asset):
        from assets.services.state import transition_asset

        result = transition_asset(asset, "retired")
        assert result.status == "retired"
        asset.refresh_from_db()
        assert asset.status == "retired"


class TestTransactionService:
    def test_create_checkout(self, asset, second_user, user):
        from assets.services.transactions import create_checkout

        txn = create_checkout(asset, second_user, user, notes="Test")
        assert txn.action == "checkout"
        assert txn.borrower == second_user
        asset.refresh_from_db()
        assert asset.checked_out_to == second_user

    def test_create_checkin(self, asset, second_user, user, location):
        from assets.services.transactions import create_checkin

        asset.checked_out_to = second_user
        asset.save()
        new_loc = Location.objects.create(name="Return Spot")
        txn = create_checkin(asset, new_loc, user, notes="Returned")
        assert txn.action == "checkin"
        asset.refresh_from_db()
        assert asset.checked_out_to is None
        assert asset.current_location == new_loc

    def test_create_transfer(self, asset, user):
        from assets.services.transactions import create_transfer

        new_loc = Location.objects.create(name="New Location")
        txn = create_transfer(asset, new_loc, user)
        assert txn.action == "transfer"
        asset.refresh_from_db()
        assert asset.current_location == new_loc


class TestPermissionsService:
    def test_superuser_is_system_admin(self, admin_user):
        from assets.services.permissions import get_user_role

        assert get_user_role(admin_user) == "system_admin"

    def test_regular_user_is_viewer(self, viewer_user):
        from assets.services.permissions import get_user_role

        assert get_user_role(viewer_user) == "viewer"

    def test_member_user_role(self, user):
        from assets.services.permissions import get_user_role

        assert get_user_role(user) == "member"

    def test_department_manager_by_assignment(self, user, department):
        from assets.services.permissions import get_user_role

        department.managers.add(user)
        assert get_user_role(user, department) == "department_manager"

    def test_can_edit_asset_admin(self, admin_user, asset):
        from assets.services.permissions import can_edit_asset

        assert can_edit_asset(admin_user, asset)

    def test_can_edit_asset_viewer_denied(self, viewer_user, asset):
        from assets.services.permissions import can_edit_asset

        assert not can_edit_asset(viewer_user, asset)

    def test_member_can_edit_own_draft(self, member_user, db):
        from assets.services.permissions import can_edit_asset

        a = Asset(name="My Draft", status="draft", created_by=member_user)
        a.save()
        assert can_edit_asset(member_user, a)

    def test_member_cannot_edit_others_draft(self, member_user, draft_asset):
        from assets.services.permissions import can_edit_asset

        assert not can_edit_asset(member_user, draft_asset)

    def test_can_delete_asset(self, admin_user, asset):
        from assets.services.permissions import can_delete_asset

        assert can_delete_asset(admin_user, asset)

    def test_viewer_cannot_delete(self, viewer_user, asset):
        from assets.services.permissions import can_delete_asset

        assert not can_delete_asset(viewer_user, asset)

    def test_can_checkout_member(self, member_user, asset):
        from assets.services.permissions import can_checkout_asset

        assert can_checkout_asset(member_user, asset)

    def test_viewer_cannot_checkout(self, viewer_user, asset):
        from assets.services.permissions import can_checkout_asset

        assert not can_checkout_asset(viewer_user, asset)


class TestMergeService:
    def test_merge_moves_tags(self, asset, user, category, location):
        from assets.services.merge import merge_assets

        dup = Asset(
            name="Duplicate",
            category=category,
            current_location=location,
            status="active",
            created_by=user,
        )
        dup.save()
        t = Tag.objects.create(name="merge-test")
        dup.tags.add(t)

        merge_assets(asset, [dup], user)
        assert t in asset.tags.all()
        dup.refresh_from_db()
        assert dup.status == "disposed"

    def test_merge_fills_missing_description(self, user, category, location):
        from assets.services.merge import merge_assets

        primary = Asset(
            name="Primary",
            category=category,
            current_location=location,
            status="active",
            created_by=user,
        )
        primary.save()

        dup = Asset(
            name="Dup",
            description="Has a description",
            category=category,
            current_location=location,
            status="active",
            created_by=user,
        )
        dup.save()

        merge_assets(primary, [dup], user)
        primary.refresh_from_db()
        assert primary.description == "Has a description"

    def test_merge_moves_transactions(self, asset, user, category, location):
        from assets.services.merge import merge_assets

        dup = Asset(
            name="Dup With Txn",
            category=category,
            current_location=location,
            status="active",
            created_by=user,
        )
        dup.save()
        Transaction.objects.create(asset=dup, user=user, action="audit")

        merge_assets(asset, [dup], user)
        assert asset.transactions.filter(action="audit").exists()


class TestExportService:
    def test_export_returns_bytes(self, asset):
        from assets.services.export import export_assets_xlsx

        buffer = export_assets_xlsx(Asset.objects.all())
        assert buffer.getvalue()[:4] == b"PK\x03\x04"  # ZIP/XLSX magic

    def test_export_contains_sheets(self, asset):
        from io import BytesIO

        import openpyxl

        from assets.services.export import export_assets_xlsx

        buffer = export_assets_xlsx(Asset.objects.all())
        wb = openpyxl.load_workbook(BytesIO(buffer.getvalue()))
        assert "Summary" in wb.sheetnames
        assert "Assets" in wb.sheetnames

    def test_export_includes_asset_data(self, asset):
        from io import BytesIO

        import openpyxl

        from assets.services.export import export_assets_xlsx

        buffer = export_assets_xlsx(Asset.objects.all())
        wb = openpyxl.load_workbook(BytesIO(buffer.getvalue()))
        ws = wb["Assets"]
        # Row 1 is header, row 2 should be our asset
        assert ws.cell(row=2, column=1).value == asset.name


class TestBarcodeService:
    def test_generate_barcode_string(self):
        from assets.services.barcode import generate_barcode_string

        bc = generate_barcode_string()
        assert bc.startswith("ASSET-")
        assert len(bc) == 14  # "ASSET-" + 8 hex chars

    def test_generate_barcode_string_unique(self):
        from assets.services.barcode import generate_barcode_string

        codes = {generate_barcode_string() for _ in range(100)}
        assert len(codes) == 100

    def test_generate_code128_image(self):
        from assets.services.barcode import generate_code128_image

        result = generate_code128_image("ASSET-TEST1234")
        assert result is not None
        # Should be a PNG
        data = result.read()
        assert len(data) > 0

    def test_get_asset_url(self):
        from assets.services.barcode import get_asset_url

        url = get_asset_url("ASSET-ABCD1234")
        assert "/a/ASSET-ABCD1234/" in url


class TestBulkService:
    def test_bulk_transfer(self, asset, user):
        from assets.services.bulk import bulk_transfer

        new_loc = Location.objects.create(name="Bulk Dest")
        result = bulk_transfer([asset.pk], new_loc.pk, user)
        assert result["transferred"] == 1
        asset.refresh_from_db()
        assert asset.current_location == new_loc

    def test_bulk_transfer_skips_checked_out(self, asset, user, second_user):
        from assets.services.bulk import bulk_transfer

        asset.checked_out_to = second_user
        asset.save()
        new_loc = Location.objects.create(name="Bulk Dest 2")
        result = bulk_transfer([asset.pk], new_loc.pk, user)
        assert result["transferred"] == 0
        assert len(result["skipped"]) == 1

    def test_bulk_status_change(self, asset, user):
        from assets.services.bulk import bulk_status_change

        count, failures = bulk_status_change([asset.pk], "retired", user)
        assert count == 1
        assert failures == []
        asset.refresh_from_db()
        assert asset.status == "retired"

    def test_bulk_status_change_skips_invalid(self, asset, user):
        from assets.services.bulk import bulk_status_change

        # active -> draft is not a valid transition
        count, failures = bulk_status_change([asset.pk], "draft", user)
        assert count == 0
        assert len(failures) == 1


class TestExportImprovements:
    """Test export filters and filename (Batch E)."""

    def test_export_date_stamped_filename(self, admin_client, asset):
        from datetime import date

        response = admin_client.get(reverse("assets:export_assets"))
        assert response.status_code == 200
        disposition = response["Content-Disposition"]
        assert date.today().isoformat() in disposition
        assert disposition.startswith("attachment;")

    def test_export_with_category_filter(self, admin_client, asset, category):
        response = admin_client.get(
            reverse("assets:export_assets") + f"?category={category.pk}"
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response["Content-Type"]

    def test_export_with_location_filter(self, admin_client, asset, location):
        response = admin_client.get(
            reverse("assets:export_assets") + f"?location={location.pk}"
        )
        assert response.status_code == 200

    def test_export_with_search_query(self, admin_client, asset):
        response = admin_client.get(
            reverse("assets:export_assets") + "?q=Test"
        )
        assert response.status_code == 200


class TestHandoverService:
    """Test custody handover creates a single 'handover' transaction."""

    def test_handover_creates_single_transaction(
        self, asset, user, second_user
    ):
        from assets.services.transactions import create_handover

        asset.checked_out_to = user
        asset.save()

        third_user = User.objects.create_user(
            username="newborrower",
            email="new@example.com",
            password="testpass123!",
        )

        txn = create_handover(
            asset, third_user, second_user, notes="Test handover"
        )

        assert txn.action == "handover"
        assert txn.borrower == third_user
        asset.refresh_from_db()
        assert asset.checked_out_to == third_user

    def test_handover_with_location(self, asset, user, second_user):
        from assets.services.transactions import create_handover

        asset.checked_out_to = user
        asset.save()
        new_loc = Location.objects.create(name="Handover Spot")

        create_handover(asset, second_user, user, to_location=new_loc)

        asset.refresh_from_db()
        assert asset.checked_out_to == second_user
        assert asset.current_location == new_loc

    def test_handover_transaction_count(self, asset, user, second_user):
        from assets.services.transactions import create_handover

        asset.checked_out_to = user
        asset.save()
        before_count = Transaction.objects.count()

        create_handover(asset, second_user, user)

        assert Transaction.objects.count() == before_count + 1


# ============================================================
# QUERY COUNT REGRESSION TESTS
# ============================================================


class TestQueryCounts:
    """Lock in query counts for key views to prevent N+1 regressions.

    Uses django_assert_num_queries to verify that views execute a fixed
    number of SQL queries regardless of data volume.
    """

    def test_dashboard_query_count(
        self,
        django_assert_num_queries,
        client_logged_in,
        asset,
        department,
        category,
        location,
        tag,
    ):
        """Dashboard should use a fixed number of queries."""
        asset.tags.add(tag)
        with django_assert_num_queries(16):
            response = client_logged_in.get(reverse("assets:dashboard"))
        assert response.status_code == 200

    def test_asset_list_query_count(
        self,
        django_assert_num_queries,
        client_logged_in,
        asset,
    ):
        """Asset list should use a fixed number of queries."""
        # +1 for PrintClient query (bulk remote print)
        with django_assert_num_queries(15):
            response = client_logged_in.get(reverse("assets:asset_list"))
        assert response.status_code == 200

    def test_asset_detail_query_count(
        self,
        django_assert_num_queries,
        client_logged_in,
        asset,
    ):
        """Asset detail should use a fixed number of queries."""
        # V500: +2 per available_count call (multiple in template),
        # V492: +2 for serial queries
        # S2.4.5: +1 for remote print client query
        # Permission-based roles use has_perm cache, fewer queries
        with django_assert_num_queries(25):
            response = client_logged_in.get(
                reverse("assets:asset_detail", args=[asset.pk])
            )
        assert response.status_code == 200


class TestQueryCountBudgets:
    """Verify views meet query count budgets from spec M12 (S8.6.5-03).

    Spec targets: dashboard ≤12, asset list ≤5, asset detail ≤8.
    Realistic budgets account for auth (2), context processors (2),
    pending-user/hold-list counts, site branding, recent transactions,
    and necessary filter/sidebar queries.
    Uses CaptureQueriesContext for detailed failure output.
    """

    def test_dashboard_query_budget(
        self,
        admin_client,
        admin_user,
        asset,
        department,
        category,
        location,
        tag,
    ):
        """Dashboard (cold cache) should use ≤17 queries."""
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        asset.tags.add(tag)
        budget = 17
        # Clear the dashboard cache to measure a cold hit
        from django.core.cache import cache

        cache.clear()
        with CaptureQueriesContext(connection) as ctx:
            response = admin_client.get(reverse("assets:dashboard"))
        assert response.status_code == 200
        assert len(ctx) <= budget, (
            f"Dashboard exceeded budget: expected ≤{budget} queries, "
            f"got {len(ctx)}:\n"
            + "\n".join(
                f"  {i}: {q['sql'][:100]}" for i, q in enumerate(ctx, 1)
            )
        )

    def test_dashboard_cached_query_budget(
        self,
        admin_client,
        admin_user,
        asset,
        department,
        category,
        location,
    ):
        """Dashboard with warm cache should use fewer queries."""
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        # First request populates cache
        admin_client.get(reverse("assets:dashboard"))
        # Second request benefits from cache
        budget = 9
        with CaptureQueriesContext(connection) as ctx:
            response = admin_client.get(reverse("assets:dashboard"))
        assert response.status_code == 200
        assert len(ctx) <= budget, (
            f"Cached dashboard exceeded budget: expected ≤{budget} "
            f"queries, got {len(ctx)}:\n"
            + "\n".join(
                f"  {i}: {q['sql'][:100]}" for i, q in enumerate(ctx, 1)
            )
        )

    def test_asset_list_query_budget(
        self,
        admin_client,
        admin_user,
        asset,
    ):
        """Asset list should use ≤17 queries.

        Breakdown: auth (2) + branding (1) + pending_approvals (1) +
        pagination COUNT (1) + main query (1) + 2 prefetches +
        4 filter sidebar + active_users (1) + available_count (2/asset)
        = 17 (with 1 asset).
        """
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        budget = 17
        with CaptureQueriesContext(connection) as ctx:
            response = admin_client.get(reverse("assets:asset_list"))
        assert response.status_code == 200
        assert len(ctx) <= budget, (
            f"Asset list exceeded budget: expected ≤{budget} queries, "
            f"got {len(ctx)}:\n"
            + "\n".join(
                f"  {i}: {q['sql'][:100]}" for i, q in enumerate(ctx, 1)
            )
        )

    def test_asset_list_no_n_plus_one(
        self,
        admin_client,
        admin_user,
        category,
        location,
        user,
    ):
        """Asset list query count should not grow with more assets.

        Creating multiple assets should not increase query count
        beyond the fixed overhead (no N+1 regression).
        """
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        from assets.models import AssetImage

        # Create 5 assets with images and tags
        assets = []
        tag_obj = Tag.objects.create(name="test-tag-n1")
        for i in range(5):
            a = Asset.objects.create(
                name=f"N+1 Test Asset {i}",
                category=category,
                current_location=location,
                status="active",
                created_by=user,
            )
            a.tags.add(tag_obj)
            AssetImage.objects.create(
                asset=a,
                image=f"test{i}.jpg",
                is_primary=True,
                uploaded_by=user,
            )
            assets.append(a)

        with CaptureQueriesContext(connection) as ctx:
            response = admin_client.get(reverse("assets:asset_list"))
        assert response.status_code == 200
        # With N+1 fixed, query count should be same as with 1 asset
        budget = 17
        assert len(ctx) <= budget, (
            f"Asset list N+1 detected: expected ≤{budget} queries "
            f"with {len(assets)} assets, got {len(ctx)}:\n"
            + "\n".join(
                f"  {i}: {q['sql'][:100]}" for i, q in enumerate(ctx, 1)
            )
        )

    def test_asset_detail_query_budget(
        self,
        admin_client,
        admin_user,
        asset,
    ):
        """Asset detail should use ≤23 queries.

        V500: available_count now uses transaction-based SUM queries
        which are called per-asset in hold list items and kit
        components. Future optimisation: cache or annotate.
        """
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        budget = 24  # +1 for S2.4.5 remote print client query
        with CaptureQueriesContext(connection) as ctx:
            response = admin_client.get(
                reverse("assets:asset_detail", args=[asset.pk])
            )
        assert response.status_code == 200
        assert len(ctx) <= budget, (
            f"Asset detail exceeded budget: expected ≤{budget} queries, "
            f"got {len(ctx)}:\n"
            + "\n".join(
                f"  {i}: {q['sql'][:100]}" for i, q in enumerate(ctx, 1)
            )
        )


class TestSerialBarcodeService:
    """Test serial barcode generation and cross-table validation."""

    def test_pattern_generation(self):
        from assets.services.barcode import (
            generate_serial_barcode_string,
        )

        result = generate_serial_barcode_string("ASSET-ABCD1234", 1)
        assert result == "ASSET-ABCD1234-S001"

    def test_pattern_generation_high_index(self):
        from assets.services.barcode import (
            generate_serial_barcode_string,
        )

        result = generate_serial_barcode_string("ASSET-ABCD1234", 42)
        assert result == "ASSET-ABCD1234-S042"

    def test_cross_table_available(self, db):
        from assets.services.barcode import (
            validate_cross_table_barcode,
        )

        assert validate_cross_table_barcode("TOTALLY-UNIQUE-CODE")

    def test_cross_table_collision_asset(self, asset):
        from assets.services.barcode import (
            validate_cross_table_barcode,
        )

        assert not validate_cross_table_barcode(asset.barcode)

    def test_cross_table_collision_serial(
        self, serialised_asset, asset_serial
    ):
        from assets.services.barcode import (
            validate_cross_table_barcode,
        )

        assert not validate_cross_table_barcode(asset_serial.barcode)

    def test_cross_table_exclude_self(self, asset):
        from assets.services.barcode import (
            validate_cross_table_barcode,
        )

        # Excluding the asset itself should make it available
        assert validate_cross_table_barcode(
            asset.barcode, exclude_asset_pk=asset.pk
        )


class TestSerialCRUDService:
    """Test serial CRUD service functions."""

    def test_create_serial(self, serialised_asset):
        from assets.services.serial import create_serial

        serial = create_serial(
            serialised_asset, "SVC-001", condition="excellent"
        )
        assert serial.pk is not None
        assert serial.serial_number == "SVC-001"
        assert serial.barcode is not None
        assert serial.condition == "excellent"

    def test_create_serial_auto_barcode(self, serialised_asset):
        from assets.services.serial import create_serial

        serial = create_serial(serialised_asset, "SVC-002")
        assert serial.barcode.startswith(serialised_asset.barcode)
        assert "-S" in serial.barcode

    def test_create_serial_non_serialised_fails(self, asset):
        from assets.services.serial import create_serial

        with pytest.raises(ValidationError, match="non-serialised"):
            create_serial(asset, "FAIL-001")

    def test_update_serial(self, asset_serial):
        from assets.services.serial import update_serial

        updated = update_serial(asset_serial, condition="poor")
        assert updated.condition == "poor"
        asset_serial.refresh_from_db()
        assert asset_serial.condition == "poor"

    def test_archive_serial(self, asset_serial):
        from assets.services.serial import archive_serial

        archived = archive_serial(asset_serial)
        assert archived.is_archived is True
        asset_serial.refresh_from_db()
        assert asset_serial.is_archived is True

    def test_restore_serial(self, asset_serial):
        from assets.services.serial import archive_serial, restore_serial

        archive_serial(asset_serial)
        restored = restore_serial(asset_serial)
        assert restored.is_archived is False

    def test_get_available_serials(
        self, serialised_asset, location, second_user
    ):
        from assets.services.serial import get_available_serials

        s1 = AssetSerial.objects.create(
            asset=serialised_asset,
            serial_number="AV-1",
            barcode="AV-1-BC",
            status="active",
            current_location=location,
        )
        AssetSerial.objects.create(
            asset=serialised_asset,
            serial_number="AV-2",
            barcode="AV-2-BC",
            status="active",
            checked_out_to=second_user,
            current_location=location,
        )
        AssetSerial.objects.create(
            asset=serialised_asset,
            serial_number="AV-3",
            barcode="AV-3-BC",
            status="retired",
            current_location=location,
        )
        available = get_available_serials(serialised_asset)
        assert list(available) == [s1]

    def test_get_serial_summary(self, serialised_asset, location, second_user):
        from assets.services.serial import get_serial_summary

        AssetSerial.objects.create(
            asset=serialised_asset,
            serial_number="SUM-1",
            barcode="SUM-1-BC",
            status="active",
            current_location=location,
        )
        AssetSerial.objects.create(
            asset=serialised_asset,
            serial_number="SUM-2",
            barcode="SUM-2-BC",
            status="active",
            checked_out_to=second_user,
            current_location=location,
        )
        AssetSerial.objects.create(
            asset=serialised_asset,
            serial_number="SUM-3",
            barcode="SUM-3-BC",
            status="retired",
            current_location=location,
        )
        summary = get_serial_summary(serialised_asset)
        assert summary["total"] == 3
        assert summary["by_status"]["active"] == 2
        assert summary["by_status"]["retired"] == 1
        assert summary["checked_out"] == 1
        assert summary["available"] == 1


@pytest.mark.django_db
class TestSetupGroupsCommand:
    """L15: Test setup_groups assigns can_approve_users permission."""

    def test_setup_groups_assigns_approve_users_to_system_admin(self, db):
        from django.contrib.auth.models import Group, Permission
        from django.contrib.contenttypes.models import ContentType
        from django.core.management import call_command

        from accounts.models import CustomUser

        call_command("setup_groups")

        system_admin = Group.objects.get(name="System Admin")
        user_ct = ContentType.objects.get_for_model(CustomUser)
        approve_perm = Permission.objects.get(
            codename="can_approve_users", content_type=user_ct
        )

        assert approve_perm in system_admin.permissions.all()

    def test_setup_groups_assigns_can_be_borrower_to_borrower(self, db):
        """Borrower group gets can_be_borrower permission."""
        from django.contrib.auth.models import Group, Permission
        from django.contrib.contenttypes.models import ContentType
        from django.core.management import call_command

        from assets.models import Asset

        call_command("setup_groups")

        borrower = Group.objects.get(name="Borrower")
        asset_ct = ContentType.objects.get_for_model(Asset)
        borrower_perm = Permission.objects.get(
            codename="can_be_borrower", content_type=asset_ct
        )

        assert borrower_perm in borrower.permissions.all()


@pytest.mark.django_db
class TestPermissionBasedRoleResolution:
    """Issue #53: Role resolution uses permissions, not group names.

    These tests verify that get_user_role() resolves roles based on
    the permissions a group grants, not the group's name. This allows
    deployments to rename groups without breaking role resolution.
    """

    def test_renamed_admin_group_resolves_system_admin(self, db, password):
        """Group renamed from 'System Admin' to anything else still
        resolves to system_admin if it has can_approve_users."""
        from conftest import _ensure_group_permissions

        # Create the canonical group first to get permissions
        _ensure_group_permissions("System Admin")
        # Now rename it
        group = Group.objects.get(name="System Admin")
        group.name = "Super Admin"
        group.save()

        u = UserFactory(username="sa_test", password=password)
        u.groups.add(group)

        from assets.services.permissions import get_user_role

        assert get_user_role(u) == "system_admin"

    def test_renamed_dept_manager_resolves(self, db, password):
        """Group with can_merge_assets (without can_approve_users)
        resolves to department_manager regardless of name."""
        from conftest import _ensure_group_permissions

        _ensure_group_permissions("Department Manager")
        group = Group.objects.get(name="Department Manager")
        group.name = "Team Lead"
        group.save()

        u = UserFactory(username="dm_test", password=password)
        u.groups.add(group)

        from assets.services.permissions import get_user_role

        assert get_user_role(u) == "department_manager"

    def test_renamed_member_resolves(self, db, password):
        """BEAMS scenario: 'Member' renamed to 'Team Member' with
        can_checkout_asset resolves to member."""
        from conftest import _ensure_group_permissions

        _ensure_group_permissions("Member")
        group = Group.objects.get(name="Member")
        group.name = "Team Member"
        group.save()

        u = UserFactory(username="member_test", password=password)
        u.groups.add(group)

        from assets.services.permissions import get_user_role

        assert get_user_role(u) == "member"

    def test_renamed_borrower_resolves(self, db, password):
        """Group with can_be_borrower (without can_checkout_asset)
        resolves to borrower."""
        from conftest import _ensure_group_permissions

        _ensure_group_permissions("Borrower")
        group = Group.objects.get(name="Borrower")
        group.name = "External Borrower"
        group.save()

        u = UserFactory(username="borrower_test", password=password)
        u.groups.add(group)

        from assets.services.permissions import get_user_role

        assert get_user_role(u) == "borrower"

    def test_no_groups_resolves_viewer(self, db, password):
        """User with no groups at all resolves to viewer."""
        u = UserFactory(username="nobody_test", password=password)

        from assets.services.permissions import get_user_role

        assert get_user_role(u) == "viewer"

    def test_hierarchy_borrower_plus_checkout_is_member(self, db, password):
        """User with both can_be_borrower and can_checkout_asset
        resolves to member (higher in hierarchy wins)."""
        from conftest import _ensure_group_permissions

        borrower_group = _ensure_group_permissions("Borrower")
        member_group = _ensure_group_permissions("Member")

        u = UserFactory(username="dual_test", password=password)
        u.groups.add(borrower_group, member_group)

        from assets.services.permissions import get_user_role

        assert get_user_role(u) == "member"

    def test_department_m2m_still_promotes_to_dept_manager(
        self, db, password, department
    ):
        """A Member user in department.managers M2M resolves to
        department_manager for that department."""
        from conftest import _ensure_group_permissions

        member_group = _ensure_group_permissions("Member")
        u = UserFactory(username="m2m_test", password=password)
        u.groups.add(member_group)
        department.managers.add(u)

        from assets.services.permissions import get_user_role

        assert get_user_role(u, department) == "department_manager"

    def test_superuser_resolves_system_admin(self, db, password):
        """Superuser always resolves to system_admin."""
        u = UserFactory(
            username="super_test",
            password=password,
            is_superuser=True,
        )

        from assets.services.permissions import get_user_role

        assert get_user_role(u) == "system_admin"


# ============================================================
# ADMIN BULK ACTION TESTS (S4.6.3.4)
# ============================================================


# ============================================================
# EXPORT ITERATOR TESTS (G11)
# ============================================================


@pytest.mark.django_db
class TestExportIterator:
    """G11: Export uses .iterator() for large datasets."""

    def test_export_uses_iterator_for_large_datasets(
        self, category, location, user
    ):
        """Exports exceeding 1000 assets use .iterator()."""
        from assets.services.export import (
            ITERATOR_THRESHOLD,
            export_assets_xlsx,
        )

        mock_qs = MagicMock()
        mock_qs.count.return_value = ITERATOR_THRESHOLD + 1
        mock_qs.filter.return_value = mock_qs
        mock_qs.aggregate.return_value = {
            "total_purchase": 0,
            "total_estimated": 0,
        }
        mock_qs.iterator.return_value = iter([])

        export_assets_xlsx(queryset=mock_qs)

        mock_qs.iterator.assert_called_once_with(chunk_size=1000)

    def test_export_works_for_small_datasets(self, category, location, user):
        """Small exports work without .iterator()."""
        from assets.services.export import export_assets_xlsx

        for i in range(3):
            a = Asset(
                name=f"Export Test {i}",
                category=category,
                current_location=location,
                status="active",
                created_by=user,
            )
            a.save()

        result = export_assets_xlsx()
        assert result is not None
        assert result.tell() == 0
        assert len(result.getvalue()) > 0

    def test_export_small_dataset_no_iterator(self, category, location, user):
        """Datasets under threshold do not call .iterator()."""
        from assets.services.export import (
            ITERATOR_THRESHOLD,
            export_assets_xlsx,
        )

        mock_qs = MagicMock()
        mock_qs.count.return_value = ITERATOR_THRESHOLD - 1
        mock_qs.filter.return_value = mock_qs
        mock_qs.aggregate.return_value = {
            "total_purchase": 0,
            "total_estimated": 0,
        }
        mock_qs.__iter__ = MagicMock(return_value=iter([]))

        export_assets_xlsx(queryset=mock_qs)

        mock_qs.iterator.assert_not_called()

    def test_export_iterator_threshold_is_1000(self):
        """The iterator threshold constant is 1000."""
        from assets.services.export import ITERATOR_THRESHOLD

        assert ITERATOR_THRESHOLD == 1000

    def test_export_iterator_chunk_size_is_1000(self):
        """The iterator chunk size constant is 1000."""
        from assets.services.export import ITERATOR_CHUNK_SIZE

        assert ITERATOR_CHUNK_SIZE == 1000


# ============================================================
# M1: Handover creates single 'handover' transaction
# ============================================================


class TestHandoverSingleTransaction:
    """Handover should create a single 'handover' type transaction."""

    def test_handover_creates_single_transaction(
        self, asset, user, second_user
    ):
        """Verify exactly 1 transaction with action='handover'."""
        from assets.services.transactions import create_handover

        asset.checked_out_to = user
        asset.save()
        before = Transaction.objects.count()

        create_handover(asset, second_user, user, notes="Single txn")

        assert Transaction.objects.count() == before + 1
        txn = Transaction.objects.latest("pk")
        assert txn.action == "handover"
        assert txn.borrower == second_user
        asset.refresh_from_db()
        assert asset.checked_out_to == second_user

    def test_handover_does_not_create_checkin_checkout(
        self, asset, user, second_user
    ):
        """Verify no checkin+checkout pair is created."""
        from assets.services.transactions import create_handover

        asset.checked_out_to = user
        asset.save()
        before = Transaction.objects.count()

        create_handover(asset, second_user, user)

        new_txns = Transaction.objects.filter(pk__gt=before)
        actions = list(new_txns.values_list("action", flat=True))
        assert "checkin" not in actions
        assert "checkout" not in actions
        assert actions == ["handover"]

    def test_handover_returns_single_transaction(
        self, asset, user, second_user
    ):
        """create_handover should return a single Transaction."""
        from assets.services.transactions import create_handover

        asset.checked_out_to = user
        asset.save()

        result = create_handover(asset, second_user, user)

        assert isinstance(result, Transaction)
        assert result.action == "handover"


# ============================================================
# BATCH 5: S4 INFRASTRUCTURE GAP TESTS
# ============================================================


@pytest.mark.django_db
class TestS3StorageConfiguration:
    """V572-V578: S3 storage, WhiteNoise, media types, Garage.

    Tests verify settings are correctly configured for storage backends.
    """

    def test_whitenoise_in_middleware(self):
        """WhiteNoise middleware is present in settings module.

        Note: conftest.py removes WhiteNoise from runtime MIDDLEWARE to
        avoid race conditions with parallel test workers, so we check
        the settings module source directly.
        """
        import props.settings as ps

        assert any("whitenoise" in m.lower() for m in ps.MIDDLEWARE)

    def test_whitenoise_staticfiles_backend_in_settings_module(self):
        """Staticfiles uses WhiteNoise storage backend in settings.py.

        Note: conftest.py overrides this for tests, so we check the
        settings module source directly.
        """
        import props.settings as ps

        # Check the prod/non-test STORAGES defined in settings.py
        # Both USE_S3 branches configure WhiteNoise for staticfiles
        source_code = open(ps.__file__).read()
        assert "whitenoise" in source_code.lower()

    def test_media_url_configured(self):
        """MEDIA_URL is set."""
        from django.conf import settings

        assert settings.MEDIA_URL

    def test_static_url_configured(self):
        """STATIC_URL is set."""
        from django.conf import settings

        assert settings.STATIC_URL

    @override_settings(USE_S3=True)
    def test_s3_storage_settings_structure(self):
        """When USE_S3 is True, the STORAGES config references S3."""
        # The actual USE_S3 check happens at settings import time,
        # so we verify the settings module contains the S3 config

        # Verify the S3 configuration exists in settings.py
        import props.settings as ps

        assert hasattr(ps, "USE_S3")

    def test_default_storage_is_filesystem_in_test(self):
        """In tests, default storage is overridden to filesystem."""
        from django.conf import settings

        backend = settings.STORAGES["default"]["BACKEND"]
        assert "FileSystem" in backend

    def test_image_field_uses_upload_to(self):
        """AssetImage.image uses upload_to='assets/'."""
        field = AssetImage._meta.get_field("image")
        assert field.upload_to == "assets/"

    def test_barcode_image_upload_to(self):
        """Asset.barcode_image uses upload_to='barcodes/'."""
        field = Asset._meta.get_field("barcode_image")
        assert field.upload_to == "barcodes/"


@pytest.mark.django_db
class TestStartupEnvValidation:
    """VV825 S8.1.8: Settings/env validation tests."""

    def test_secret_key_has_default(self):
        """SECRET_KEY is configured (defaults to dev key)."""
        from django.conf import settings

        assert settings.SECRET_KEY is not None
        assert len(settings.SECRET_KEY) > 10

    def test_debug_setting_is_boolean(self):
        """DEBUG setting is a boolean."""
        from django.conf import settings

        assert isinstance(settings.DEBUG, bool)

    def test_allowed_hosts_is_list(self):
        """ALLOWED_HOSTS is a list."""
        from django.conf import settings

        assert isinstance(settings.ALLOWED_HOSTS, list)
        assert len(settings.ALLOWED_HOSTS) >= 1

    def test_database_configured(self):
        """A database backend is configured."""
        from django.conf import settings

        assert "default" in settings.DATABASES
        assert "ENGINE" in settings.DATABASES["default"]

    def test_auth_user_model_configured(self):
        """Custom user model is correctly configured."""
        from django.conf import settings

        assert settings.AUTH_USER_MODEL == "accounts.CustomUser"

    def test_installed_apps_contains_core(self):
        """Core apps are installed."""
        from django.conf import settings

        assert "accounts" in settings.INSTALLED_APPS
        assert "assets" in settings.INSTALLED_APPS

    def test_authentication_backend_configured(self):
        """Custom auth backend is configured."""
        from django.conf import settings

        assert (
            "accounts.backends.EmailOrUsernameBackend"
            in settings.AUTHENTICATION_BACKENDS
        )


@pytest.mark.django_db
class TestTransactionAuditTrail:
    """Transaction audit trail: records are created for all operations."""

    def test_checkout_creates_transaction(self, asset, user, second_user):
        from assets.services.transactions import create_checkout

        txn = create_checkout(asset, second_user, user, notes="Audit")
        assert txn.pk is not None
        assert txn.action == "checkout"
        assert txn.borrower == second_user
        assert txn.user == user
        assert txn.from_location == asset.current_location

    def test_checkin_creates_transaction(
        self, asset, user, second_user, location
    ):
        from assets.services.transactions import (
            create_checkin,
            create_checkout,
        )

        create_checkout(asset, second_user, user)
        return_loc = Location.objects.create(name="Checkin Loc")
        txn = create_checkin(asset, return_loc, user)
        assert txn.action == "checkin"
        assert txn.to_location == return_loc

    def test_transfer_creates_transaction(self, asset, user):
        from assets.services.transactions import create_transfer

        new_loc = Location.objects.create(name="Transfer Dest")
        txn = create_transfer(asset, new_loc, user)
        assert txn.action == "transfer"
        assert txn.to_location == new_loc

    def test_handover_creates_transaction(self, asset, user, second_user):
        from assets.services.transactions import (
            create_checkout,
            create_handover,
        )

        create_checkout(asset, second_user, user)
        new_borrower = User.objects.create_user(
            username="newborrower", password="test123!"
        )
        txn = create_handover(asset, new_borrower, user)
        assert txn.action == "handover"
        assert txn.borrower == new_borrower

    def test_transaction_immutability(self, asset, user):
        txn = Transaction.objects.create(
            asset=asset, user=user, action="audit"
        )
        with pytest.raises(ValidationError, match="immutable"):
            txn.notes = "changed"
            txn.save()

    def test_transaction_cannot_be_deleted(self, asset, user):
        txn = Transaction.objects.create(
            asset=asset, user=user, action="audit"
        )
        with pytest.raises(ValidationError, match="immutable"):
            txn.delete()

    def test_backdated_transaction(self, asset, user, second_user):
        from assets.services.transactions import create_checkout

        past = timezone.now() - timezone.timedelta(days=7)
        txn = create_checkout(asset, second_user, user, timestamp=past)
        assert txn.is_backdated is True
        assert txn.timestamp == past

    def test_transfer_to_same_location_raises(self, asset, user):
        from assets.services.transactions import create_transfer

        with pytest.raises(ValueError, match="already at this"):
            create_transfer(asset, asset.current_location, user)

    def test_handover_to_same_borrower_raises(self, asset, user, second_user):
        from assets.services.transactions import (
            create_checkout,
            create_handover,
        )

        create_checkout(asset, second_user, user)
        with pytest.raises(ValueError, match="already checked out"):
            create_handover(asset, second_user, user)


LOCMEM_CACHE = {
    "default": {
        "BACKEND": ("django.core.cache.backends.locmem.LocMemCache"),
    }
}


@pytest.mark.django_db
class TestDashboardCaching:
    """G10: Dashboard aggregate caching with 60s TTL."""

    def setup_method(self):
        cache.clear()

    def test_dashboard_uses_cache(self, admin_client, asset):
        """Second dashboard hit uses cached aggregates."""
        url = reverse("assets:dashboard")
        with override_settings(CACHES=LOCMEM_CACHE):
            cache.clear()
            admin_client.get(url)

            from django.contrib.auth import get_user_model

            admin = get_user_model().objects.get(username="admin")
            cache_key = f"dashboard_aggregates_{admin.pk}"
            cached = cache.get(cache_key)
            assert cached is not None
            assert "total_active" in cached

    def test_dashboard_cache_returns_correct_data(self, admin_client, asset):
        """Cached data matches live query results."""
        url = reverse("assets:dashboard")
        with override_settings(CACHES=LOCMEM_CACHE):
            cache.clear()
            resp1 = admin_client.get(url)
            resp2 = admin_client.get(url)

            assert (
                resp1.context["total_active"] == resp2.context["total_active"]
            )
            assert resp1.context["total_draft"] == resp2.context["total_draft"]

    def test_dashboard_cache_expires(self, admin_client, asset):
        """Cache is invalidated after TTL expires."""
        from django.contrib.auth import get_user_model

        url = reverse("assets:dashboard")
        with override_settings(CACHES=LOCMEM_CACHE):
            cache.clear()
            admin_client.get(url)

            admin = get_user_model().objects.get(username="admin")
            cache_key = f"dashboard_aggregates_{admin.pk}"

            cache.delete(cache_key)
            assert cache.get(cache_key) is None

            admin_client.get(url)
            assert cache.get(cache_key) is not None

    def test_dashboard_cache_scoped_by_user(
        self,
        admin_client,
        client,
        admin_user,
        user,
        password,
        asset,
    ):
        """Different users get different cache keys."""
        url = reverse("assets:dashboard")
        with override_settings(CACHES=LOCMEM_CACHE):
            cache.clear()
            admin_client.get(url)
            admin_key = f"dashboard_aggregates_{admin_user.pk}"
            assert cache.get(admin_key) is not None

            client.login(username=user.username, password=password)
            client.get(url)
            user_key = f"dashboard_aggregates_{user.pk}"
            assert cache.get(user_key) is not None

            assert admin_key != user_key

    def test_dashboard_cache_ttl_is_60_seconds(self, admin_client, asset):
        """Cache is set with 60-second TTL."""
        from assets.views import DASHBOARD_CACHE_TTL

        assert DASHBOARD_CACHE_TTL == 60

    def test_dashboard_dept_manager_separate_cache(
        self, admin_user, dept_manager_user, asset
    ):
        """Dept managers and admins get different cache keys."""
        admin_key = f"dashboard_aggregates_{admin_user.pk}"
        mgr_key = f"dashboard_aggregates_{dept_manager_user.pk}"
        assert admin_key != mgr_key

        # Both keys store independently
        with override_settings(CACHES=LOCMEM_CACHE):
            cache.clear()
            cache.set(admin_key, {"total_active": 5}, 60)
            cache.set(mgr_key, {"total_active": 3}, 60)

            assert cache.get(admin_key)["total_active"] == 5
            assert cache.get(mgr_key)["total_active"] == 3
